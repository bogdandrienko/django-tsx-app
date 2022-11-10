import collections

from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User, Group, update_last_login
from django.http import HttpRequest
from django.shortcuts import render
from django.views import View
from django.views.decorators.cache import cache_page
from django.core.cache import caches
from django.db import connection, transaction
from django.http import JsonResponse

import datetime
import random
import re
import time
import openpyxl
from django_settings.celery import app as celery_instance
from celery.result import AsyncResult
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status

from django_app import models as django_models, serializers as django_serializers, utils as django_utils, \
    signals as django_signals, celery as django_celery

django_signals.register_all_signals()
http_method_names = ["GET", "POST", "PUT", "PATCH", "DELETE", "HEAD", "OPTIONS"]

LocMemCache = caches["default"]
DatabaseCache = caches["special"]
# RedisCache = caches["extra"]


# Create your views here.

# @cache_page(timeout=1)
def index_f(request):
    try:
        context = {}
        return render(request, "build/index.html", context=context)
    except Exception as error:
        if settings.DEBUG:
            print(f"error {error}")
        return django_utils.DjangoClass.DRFClass.RequestOldClass.return_global_error(request=request, error=error)


# @cache_page(timeout=120)
@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def captcha_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> any:
    if request.method == "GET":

        def caching(key: str, value: any, cache_instance: any, timeout: int) -> any:
            if value is not None:
                cache_instance.set(key, value, timeout=timeout)
            return cache_instance.get(key)

        users_list = caching(
            key="users_list",
            value=[{"username": f"{user.username}", "email": f"{user.email}"} for user in User.objects.all()],
            cache_instance=LocMemCache,
            timeout=3
        )
        users_list = [{"username": f"{user.username}", "email": f"{user.email}"} for user in User.objects.all()]
        print("cache_f users_list: ", users_list)

        return "Вы не робот!"


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def token_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> any:
    if request.method == "POST":
        # TODO {"username": "admin", "password": "admin"}
        # TODO {"username": "user", "password": "12345Qwerty!"}
        username = request.get(key="username", _type=str, default="", is_file=False)
        password = request.get(key="password", _type=str, default="", is_file=False)

        if username and password:
            user_obj = authenticate(username=username, password=password)
            if user_obj is None:
                raise Exception("Username or password incorrect!")
            if user_obj.user_model.is_active_account is False:
                raise Exception("Attention, your account is banned!")
            update_last_login(sender=None, user=user_obj)
            token_str = django_models.TokenModel.create_or_update_token(user=user_obj)
            return {"token": token_str}
        raise Exception("Username or password not fulled!")


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def todo_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> any:
    if request.pk:
        if request.method == "GET":
            if request.action == "_":
                todo_obj = django_models.Todo.objects.get(id=request.pk)
                return django_serializers.TodoSerializer(todo_obj, many=False).data
        elif request.method == "PUT" or request.method == "PATCH":
            if request.action == "_":
                title = request.get(key="title", _type=str, default=None, is_file=False)
                description = request.get(key="description", _type=str, default=None, is_file=False)
                avatar = request.get(key="avatar", _type=any, default=None, is_file=True)
                is_completed = request.get(key="is_completed", _type=bool, default=None, is_file=True)

                todo_obj = django_models.Todo.objects.get(id=request.pk)
                if title is not None and todo_obj.title != title:
                    todo_obj.title = title
                if description is not None and todo_obj.description != description:
                    todo_obj.description = description
                if avatar is not None and todo_obj.avatar != avatar:
                    todo_obj.avatar = avatar
                if is_completed is not None and todo_obj.is_completed != is_completed:
                    todo_obj.is_completed = is_completed
                todo_obj.save()
                return "Successfully update"
        elif request.method == "DELETE":
            if request.action == "":
                django_models.Todo.objects.get(id=request.pk).delete()
                return "Successfully delete"
    else:
        if request.method == "GET":
            if request.action == "_":
                page = request.get(key="page", _type=int, default=None, is_file=False)
                limit = request.get(key="limit", _type=int, default=None, is_file=False)

                todos_obj = django_models.Todo.objects.all()
                if page and limit:
                    todos_obj = django_utils.DjangoClass.PaginationClass.paginate(
                        page_number=page, object_list=todos_obj, limit_per_page=limit
                    )
                return django_serializers.TodoSerializer(todos_obj, many=True).data
        elif request.method == "POST":
            if request.action == "_":
                title = request.get(key="title", _type=str, default=None, is_file=False)
                description = request.get(key="description", _type=str, default=None, is_file=False)
                avatar = request.get(key="avatar", _type=any, default=None, is_file=True)
                is_completed = request.get(key="is_completed", _type=bool, default=False, is_file=False)

                if title:
                    django_models.Todo.objects.create(
                        title=title, description=description, avatar=avatar, is_completed=is_completed
                    )
                    return "Successfully create"
                else:
                    raise Exception({"detail": "Not have data"})


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def report_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> any:
    if request.pk:
        pass
    else:
        if request.method == "GET":
            if request.action == "monitoring":
                data = [
                    {"id": x, "type": "Автосамосвал", "speed": random.randint(0, 20), "mass": random.randint(87, 102),
                     "status": random.choice(["Норма", "Простой", "Движение", "Погрузка", "Ремонт", "ППР"]),
                     "time": django_utils.DateTimeUtils.get_current_time()}
                    for x in range(201, 230)
                ]
                return {"data": data}
            elif request.action == "report":
                try:
                    # ?page=1&limit=10&sort_by=name%20(asc)&filter_by=pay%20(true)&search=95
                    page = request.GET.get("page", 1)
                    limit = request.GET.get("limit", 10)
                    search = request.GET.get("search", "")
                    filter_by = request.GET.get("filter_by", "")
                    sort_by = request.GET.get("sort_by", "")

                    workbook = openpyxl.load_workbook('static/media/vehtrips.xlsx')
                    worksheet = workbook.active
                    matrix = worksheet.iter_rows(
                        min_col=1, min_row=1, max_col=worksheet.max_column, max_row=300, values_only=True
                    )
                    # matrix = []
                    # for i in range(1, worksheet.max_row+1):
                    #     row_data = []
                    #     for j in range(1, worksheet.max_column+1):
                    #         print(i, j)
                    #         row_data.append(worksheet.cell(i, j).value)
                    #     matrix.append(row_data)
                    matrix = [
                        {"Автосамосвал": x[0], "Экскаватор": x[1], "Зона разгрузки": x[2], "Тип породы": x[3],
                         "Зона погрузки": x[4], "Ном.": x[5], "Расстояние": x[6], "Масса": x[7], "Объём": x[8],
                         "Частота по массе": x[9], "Частота по объёму": x[10], "Сред. скорость": x[11],
                         "Часы": x[12], "Высота погрузки": x[13], "Высота разгрузки": x[14], "Время погрузки": x[15],
                         "Время выезда на дорогу": x[16], "Время разгрузки": x[17], "Время рейса": x[18]}
                        for x in list(matrix)
                    ]
                    matrix = [
                        x for x in matrix
                        if isinstance(x["Время разгрузки"], datetime.datetime) and
                           (filter_by == "все" or filter_by == x["Тип породы"]) and len(x["Зона разгрузки"]) > 0
                    ]
                    if sort_by == "времени погрузки (свежие в начале)":
                        matrix.sort(key=lambda x: x["Время погрузки"], reverse=True)
                    else:
                        matrix.sort(key=lambda x: x["Время погрузки"])
                    return {"data": matrix}
                except Exception as error:
                    print(error)
                    return {"data": []}
