import datetime
import openpyxl
from django.core.cache import caches
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.request import Request
from django_app import utils as django_utils, queries as django_queries

# from django_app import serializers as django_serializers
# from django_app import signals as django_signals

DatabaseCache = caches["default"]
LocMemCache = caches["ram_cache"]


def index(request: HttpRequest) -> HttpResponse:
    try:
        return render(request=request, template_name="index.html", context={})
    except Exception as error:
        return render(request=request, template_name="django_app/404.html", context={})


@django_utils.drf_decorator(auth=False)
def f_events_drainage(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        time_diff = request.GET["timeDiff"]

        def __query() -> list | dict | str:
            __rows: tuple = django_utils.request_to_oracle(query=django_queries.query_drainage_status(), args={"timeDiff": time_diff}, many=False)
            __data = {
                "maxtime": __rows[0],
                "mintime": __rows[1],
                "maxfuel": __rows[2],
                "minfuel": __rows[3],
                "diffuel": int(__rows[4]),
                "difval": int(__rows[5]),
            }
            return __data

        data = django_utils.caching(LocMemCache, f"{request.path}_{request.method}_{time_diff}", timeout=9, query=__query)

        return {"data": data}


@django_utils.drf_decorator(auth=False)
def f_events_dumptrucks(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        def __query() -> list | dict | str:
            rows_raw = django_utils.request_to_oracle(query=django_queries.query_vehtrips_status(), args={}, many=True)
            rows_instances = [{"vehid": i[0], "time": i[1], "x": i[2], "y": i[3], "weight": i[4], "fuel": i[5], "speed": i[6]} for i in rows_raw]
            return rows_instances

        r_data = django_utils.caching(LocMemCache, f"{request.path}_{request.method}", timeout=5, query=__query)

        return {"data": r_data}


def f_rational(request: Request):
    if request.method == "GET":
        return render(request, "rational_post.html", context={})


@django_utils.drf_decorator(auth=False)
def f_analyse_predictive(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        def __query() -> list | tuple | dict | str:
            trips: list[tuple] = django_utils.request_to_oracle(query=django_queries.query_analyse_predictive(), args={}, many=True)

            workbook: Workbook = openpyxl.load_workbook("static_external/table_norms.xlsx")
            worksheet: Worksheet = workbook["Самосвалы"]
            matrix_dumptrucks = {}
            for i in range(1 + 1, worksheet.max_row + 1):
                matrix_dumptrucks[int(worksheet.cell(row=i, column=1).value)] = {
                    "Состояние, %": worksheet.cell(row=i, column=2).value,
                    "Ср. скорость гружённый": worksheet.cell(row=i, column=3).value,
                    "Ср. скорость порожний": worksheet.cell(row=i, column=4).value,
                    "Время разгрузки": worksheet.cell(row=i, column=5).value,
                    "Время ожидания": worksheet.cell(row=i, column=6).value,
                    "Погода, %": worksheet.cell(row=i, column=7).value,
                }

            worksheet: Worksheet = workbook["Экскаваторы"]
            matrix_shovels = {}
            for i in range(1 + 1, worksheet.max_row + 1):
                matrix_shovels[str(worksheet.cell(row=i, column=1).value)] = {
                    "Вскрыша скальная": worksheet.cell(row=i, column=2).value,
                    "Вскрыша рыхлая": worksheet.cell(row=i, column=3).value,
                    "Вскрыша транзитная": worksheet.cell(row=i, column=4).value,
                    "Руда скальная": worksheet.cell(row=i, column=5).value,
                    "ВКП скала": worksheet.cell(row=i, column=6).value,
                    "Среднее кол-во ковшей": worksheet.cell(row=i, column=7).value,
                }

            def kpd_match(j):
                movetime: datetime.datetime = j[7]
                elapsed_time1 = round((movetime - datetime.datetime(2000, 1, 1)).total_seconds() / 60, 1)

                veh = matrix_dumptrucks[int(j[
                                                1])]  # {'Состояние, %': 100, 'Ср. скорость гружённый': 20.5, 'Ср. скорость порожний': 22.5, 'Время разгрузки': 0.5, 'Время ожидания': 0}
                # shov = matrix_shovels[str(j[2])]  # {'Вскрыша скальная': 2.5, 'Вскрыша рыхлая': 2.5, 'Вскрыша транзитная': 2.5, 'Руда скальная': 2.5, 'ВКП Скала': 2.5}

                path = round(j[11] / veh["Ср. скорость гружённый"] * 60, 1)

                full = round(path * (100 / veh["Состояние, %"]) * (100 / veh["Погода, %"]), 1)
                kpd = round((full / elapsed_time1) * 100, 0)

                return kpd, f"({path}[путь]) * {veh['Состояние, %']}[состояние самосвала] * {veh['Погода, %']}[погода] = {full}[итого норма] | {round(elapsed_time1, 1)}[факт]", \
                    0

            def kpd_match_old(j):
                elapsed_time1 = round((j[6] - j[5]).total_seconds() / 60, 1)

                veh = matrix_dumptrucks[int(j[
                                                1])]  # {'Состояние, %': 100, 'Ср. скорость гружённый': 20.5, 'Ср. скорость порожний': 22.5, 'Время разгрузки': 0.5, 'Время ожидания': 0}
                shov = matrix_shovels[
                    str(j[2])]  # {'Вскрыша скальная': 2.5, 'Вскрыша рыхлая': 2.5, 'Вскрыша транзитная': 2.5, 'Руда скальная': 2.5, 'ВКП Скала': 2.5}

                # load = (j[9] * shov[j[4]]) / 60 if j[9] > 0 else (shov["Среднее кол-во ковшей"] * shov[j[4]]) / 60
                # path = j[11] / veh["Ср. скорость гружённый"] * 60
                # unload = veh["Время разгрузки"]
                # return_path = j[11] / veh["Ср. скорость порожний"] * 60
                # full = round(load + path + unload, 1)
                # full = full * (100 / veh["Состояние, %"])
                # kpd = round((full / elapsed_time1) * 100, 0)

                load = round(shov[j[4]], 1)
                path = round(j[11] / veh["Ср. скорость гружённый"] * 60, 1)
                unload = round(veh["Время разгрузки"], 1)
                return_path = round(j[12] / veh["Ср. скорость порожний"] * 60, 1)

                elapsed_time1 += round(return_path, 1)

                full = round(load + path + unload + return_path, 1)
                full = round(full * (100 / veh["Состояние, %"]) * (100 / veh["Погода, %"]), 1)
                kpd = round((full / elapsed_time1) * 100, 0)

                return kpd, f"({load}[погрузка] + {path}[путь] + {unload}[разгрузка] + {return_path}[возврат]) * " \
                            f"{veh['Состояние, %']}[состояние самосвала] = {full}[итого норма] | {round(elapsed_time1, 1)}[факт]", return_path

            trips_raw = [
                {
                    "vehid": i[1],
                    "shovid": i[2],
                    "unloadid": i[3],
                    "worktype": i[4],
                    "timeload": i[5],
                    "timeunload": i[6],
                    "movetime": i[7],
                    "weigth": i[8],
                    "bucketcount": i[9],
                    "avspeed": i[10],
                    "length": i[11],
                    "unloadlength": i[12],
                    "loadheight": i[13],
                    "unloadheight": i[13],
                    "kpd": kpd_match(i)[0],
                    "detail": kpd_match(i)[1],
                }
                for i in trips
            ]

            trips_raw = sorted(trips_raw, key=lambda x: x["timeload"], reverse=True)

            # todo сгруппировать по каждому самосвалу
            data1 = {}
            for trips in trips:
                tech_id = f"{trips[1]}"
                try:
                    data1[tech_id] = [*data1[tech_id], trips]
                except Exception as error:
                    data1[tech_id] = [trips]

            # todo сгруппировать по трём категориям
            data2 = {}
            for tech_id, trips in data1.items():
                data2[tech_id] = {}
                data2[tech_id]["last_trip"] = [trips[-1]]
                for trip in trips:
                    # todo CORE
                    if (datetime.datetime.now() - trip[5]).total_seconds() < 1 * 60 * 60:
                        try:
                            data2[tech_id]["last_hour"] = [*data2[tech_id]["last_hour"], trip]
                        except Exception as error:
                            data2[tech_id]["last_hour"] = [trip]
                    try:
                        data2[tech_id]["last_shift"] = [*data2[tech_id]["last_shift"], trip]
                    except Exception as error:
                        data2[tech_id]["last_shift"] = [trip]

            # todo оценить каждый рейс и записать в новые поля
            data3 = {}
            for tech_id, trips in data2.items():
                data3[tech_id] = {}
                data3[tech_id]["ratings"] = {}
                for type_par in ["last_trip", "last_hour", "last_shift"]:
                    try:
                        count = len(data2[tech_id][type_par])
                        if count > 0:
                            res = 0
                            elapsed_time = 0
                            for trip in data2[tech_id][type_par]:
                                res += kpd_match(trip)[0]
                                elapsed_time = ((trip[6] - trip[5]).total_seconds() / 60) + kpd_match(trip)[2]
                            if type_par == "last_trip":
                                data3[tech_id]["ratings"][type_par] = {
                                    "rating": int(res / count),
                                    "count": int(elapsed_time),
                                }
                            else:
                                data3[tech_id]["ratings"][type_par] = {"rating": int(res / count), "count": count}
                        else:
                            data3[tech_id]["ratings"][type_par] = {"rating": 0, "count": 0}
                    except Exception as error:
                        data3[tech_id]["ratings"][type_par] = {"rating": 0, "count": 0}

            # todo конвертация словаря в массив
            data4 = []
            for key, value in data3.items():
                data4.append({"tech_id": key, **value})

            return data4, trips_raw

        r_data, r_trips = django_utils.caching(LocMemCache, f"{request.path}_{request.method}", timeout=1, query=__query)

        return {"data": r_data, "trips": r_trips}


@django_utils.drf_decorator(auth=False)
def f_reports_time_wait_to_load(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        param_date = datetime.datetime.strptime(request.GET["date"], "%Y-%m-%d")
        param_shift = request.GET["shift"]
        param_select_tech_id = request.GET["selectTechId"]
        param_target = request.GET["target"]

        def __query() -> list | tuple | dict | str:
            trips: list[tuple] = django_utils.request_to_oracle(
                query=django_queries.query_time_wait_to_load(),
                args=dict(param_date=param_date, param_shift=param_shift, param_select_tech_id=param_select_tech_id, param_target=param_target),
                many=True
            )

            trips_raw = [
                {
                    "vehid": i[0],
                    "timeFrom": i[2],
                    "timeTo": i[3],
                    "time": i[4],
                }
                for i in trips
            ]

            return trips_raw

        def __query2() -> list | tuple | dict | str:
            idles: tuple = django_utils.request_to_oracle(
                query=django_queries.query_time_wait_to_load_avg(),
                args=dict(param_date=param_date, param_shift=param_shift, param_target=2),
                many=False
            )
            idle = {
                "sum_idles": idles[0],
                "trips_idels": idles[1],
                "trips_all": idles[2],
                "avg_wait_all": idles[3],
                "avg_wait": idles[4],
            }
            return idle

        try:
            workbook: Workbook = openpyxl.load_workbook("grafick.xlsx")
            worksheet: Worksheet = workbook.active

            matrix = [x for x in worksheet.iter_rows(values_only=True, min_row=2)]
            name = "Не назначено"
            for i in matrix[::-1]:
                if int(i[1]) == int(param_shift) and (i[0] <= param_date):
                    name = f'{i[2]} [от {i[0].strftime("%d.%m.%Y")} | {i[1]}]'
                    break
        except Exception as error:
            name = str(error)

        r_trips = django_utils.caching(
            LocMemCache,
            f"{request.path}_{request.method}_{request.GET['date']}_{param_shift}_{'All' if param_select_tech_id == 'Все' else param_select_tech_id}_"
            f"{param_target}",
            timeout=10, query=__query
        )
        r_avg = django_utils.caching(
            LocMemCache,
            f"{request.path}_{request.method}_{request.GET['date']}_{param_shift}__avg_"
            f"{param_target}",
            timeout=10, query=__query2
        )

        return {"data": r_trips, "idle": r_avg, "query": {"date": param_date, "shift": param_shift, "name": name}}


@django_utils.drf_decorator(auth=False)
def f_reports_errors_asd(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        param_date = datetime.datetime.strptime(request.GET["date"], "%Y-%m-%d")
        param_shift = request.GET["shift"]
        # param_select_tech_id = request.GET["selectTechId"]
        param_target = request.GET["target"]

        def __query() -> list | tuple | dict | str:
            errors: list[tuple] = django_utils.request_to_oracle(
                query=django_queries.query_errors_asd(),
                args=dict(param_date=param_date, param_shift=param_shift, param_target=param_target),
                many=True
            )

            # errors = [
            #     ("ошибка 1", "описание 1"),
            #     ("ошибка 2", "описание 2"),
            #     ("ошибка 3", "описание 3"),
            # ]

            trips_raw = [
                {
                    "title": i[0],
                    "description": i[1],
                }
                for i in errors
            ]
            return trips_raw

        r_errors = django_utils.caching(
            LocMemCache,
            f"{request.path}_{request.method}_{request.GET['date']}_{param_shift}_{param_target}_"
            f"{param_target}",
            timeout=10, query=__query
        )

        return {"data": r_errors, "query": {"date": param_date, "shift": param_shift}}


@django_utils.drf_decorator(auth=False)
def f_reports_avg_speed(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        param_date = datetime.datetime.strptime(request.GET["date"], "%Y-%m-%d")
        param_shift = request.GET["shift"]
        param_select_tech_id = request.GET["selectTechId"]

        _args = dict(param_date=param_date, param_shift=param_shift, param_select_tech_id=param_select_tech_id)

        def __query() -> list | tuple | dict | str:
            trips: list[tuple] = django_utils.request_to_oracle(query=django_queries.query_analyse_avg_speed(), args={**_args}, many=True)

            trips_raw = [
                {
                    "vehid": i[0],
                    "shovid": i[1],
                    "fio": i[2],
                    "taskdate": i[3],
                    "shift": i[4],
                    "trip": i[5],
                    "worktype": i[6],
                    "timeload": i[7],
                    "timeunload": i[8],
                    "avgloadspeed": i[9],
                    "avgemptyspeed": i[10],
                    "avgspeed": i[11],
                }
                for i in trips
            ]

            # todo сгруппировать по каждому самосвалу
            data1 = {}
            for trip in trips_raw:
                tech_id = f"{trip['vehid']}"
                try:
                    data1[tech_id] = [*data1[tech_id], trip]
                except Exception as error:
                    data1[tech_id] = [trip]

            def kpd(_trips: list[dict]) -> dict:
                _avgloadspeed = 0.0
                _avgemptyspeed = 0.0
                _avgspeed = 0.0

                count = len(_trips[:-1:])

                for i in _trips[:-1:]:
                    _avgloadspeed += i["avgloadspeed"] if i["avgloadspeed"] else 0.0
                    _avgemptyspeed += i["avgemptyspeed"] if i["avgemptyspeed"] else 0.0
                    _avgspeed += i["avgspeed"] if i["avgspeed"] else 0.0

                _avgloadspeed = round(_avgloadspeed / count, 1)
                _avgemptyspeed = round(_avgemptyspeed / count, 1)
                _avgspeed = round(_avgspeed / count, 1)

                return {"avgloadspeed": _avgloadspeed, "avgemptyspeed": _avgemptyspeed, "avgspeed": _avgspeed,
                        "last_loadspeed": _trips[-1]["avgloadspeed"], "last_emptyspeed": _trips[-1]["avgemptyspeed"], "last_speed":
                            _trips[-1]["avgspeed"], "count": count + 1, "fio": _trips[-1]["fio"]}

            data2 = []
            for k, v in data1.items():
                val = kpd(v)
                data2.append({"tech": f"{k} ({val['count']} р.)", "value": val})

            data3 = sorted(data2, key=lambda x: x["tech"], reverse=False)

            return data3

        r_trips = django_utils.caching(
            LocMemCache,
            f"{request.path}_{request.method}_{request.GET['date']}_{param_shift}_{'All' if param_select_tech_id == 'Все' else param_select_tech_id}",
            timeout=10, query=__query
        )

        return {"data": r_trips, "query": {"date": param_date, "shift": param_shift}}


@django_utils.drf_decorator(auth=False)
def f_reports_operuchet_dumptrucks(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        param_date_from = datetime.datetime.strptime(request.GET["dateFrom"], "%Y-%m-%d")
        param_shift_from = request.GET["shiftFrom"]
        param_date_to = datetime.datetime.strptime(request.GET["dateTo"], "%Y-%m-%d")
        param_shift_to = request.GET["shiftTo"]
        param_select_tech_id = request.GET["selectTechId"]
        round_point = int(request.GET["roundPoint"])

        def __query() -> list | dict | str:
            rows_raw = django_utils.request_to_oracle(
                query=django_queries.query_reports_operuchet(),
                args={
                    "paramDateFrom": param_date_from,
                    "paramShiftFrom": param_shift_from,
                    "paramDateTo": param_date_to,
                    "paramShiftTo": param_shift_to,
                    "paramSelectTechId": param_select_tech_id,
                },
                many=True,
            )
            row_instances = []
            for i in rows_raw:
                tr_mass = round(i[4] / 2.1, round_point)
                tr_gruz = i[5] * tr_mass
                sk_mass = round(i[6] / 2.8, round_point)
                sk_gruz = i[7] * sk_mass
                rih_mass = round(i[8] / 1.8, round_point)
                rih_gruz = i[9] * rih_mass
                prs_mass = round(i[10] / 1.4, round_point)
                prs_gruz = i[11] * prs_mass
                rud_mass = round(i[12] / 2.8, round_point)
                rud_gruz = i[13] * rud_mass
                summ_mass = round(tr_mass + sk_mass + rih_mass + prs_mass + rud_mass, round_point)
                summ_proiz_gruz = round(tr_gruz + sk_gruz + rih_gruz + prs_gruz + rud_gruz, round_point)
                if summ_mass > 0:
                    km = round(summ_proiz_gruz / summ_mass, round_point)
                else:
                    km = 0
                temp_row_instances = [
                    [i[3], "тр,км", "тр", round(i[4], round_point), round(i[5], round_point), 2.1, tr_mass, summ_proiz_gruz, summ_mass, km],
                    [i[1], "ск,км", "ск", round(i[6], round_point), round(i[7], round_point), 2.8, sk_mass, "", "", ""],
                    ["", "рых,км", "рых", round(i[8], round_point), round(i[9], round_point), 1.8, rih_mass, "", "", ""],
                    ["", "ПРС,км", "прс", round(i[10], round_point), round(i[11], round_point), 1.4, prs_mass, "", "", ""],
                    ["", "руда,км", "руд", round(i[12], round_point), round(i[13], round_point), 2.8, rud_mass, "", "", ""],
                ]
                row_instances.extend(temp_row_instances)
            return row_instances

        r_data = django_utils.caching(
            LocMemCache,
            f"{request.path}_{request.method}_{param_date_from.strftime('%H_%M_%S')}_{param_shift_from}_"
            f"{param_date_to.strftime('%H_%M_%S')}_{param_shift_to}_{param_select_tech_id}_{round_point}",
            timeout=5,
            query=__query,
        )

        return {"data": r_data}


# todo #################################################################################################################


@django_utils.drf_decorator(auth=False)
def f_users_captcha(request: Request, pk=0) -> dict | str | None:
    if request.method == "GET":
        if pk != 0:
            return None

        return "Вы не робот!"
