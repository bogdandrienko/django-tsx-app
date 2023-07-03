import datetime
import cx_Oracle
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

query = """
WITH b
     AS (SELECT GETPREDEFINEDTIMEFROM('за указанную смену', :param_shift, :param_date) SDATEFROM,
                GETPREDEFINEDTIMETO('за указанную смену', :param_shift, :param_date)   SDATETO,
                :param_select_tech_id                                                                  TECHID
         FROM   DUAL)
SELECT q.VEHID,
       q.SHOVID,
       TRIM(d.FAMNAME)
       || ' '
       || TRIM(d.FIRSTNAME)
       || ' '
       || TRIM(d.SECNAME)                            fio,
       KEM_DATETODDMMYYYY(q.TASKDATE)                TASKDATE,
       q.SHIFT,
       q.TRIP                                        trips,
       q.WORKTYPE                                    worktype,
       q.TIMELOAD,
       q.TIMEUNLOAD,
       NVL(ROUND(q.AVSPEED, 2), -1)                           avgloadspeed,
       NVL(ROUND(q.AVSPEED_EMPTY, 2), -1)                     avgemptyspeed,
       NVL(ROUND(( q.AVSPEED + q.AVSPEED_EMPTY ) / 2, 2), -1) avspeed
FROM   (SELECT s.VEHID,
               s.VEHCODE,
               s.SHOVID,
               s.WORKTYPE,
               s.TIMELOAD,
               s.TIMEUNLOAD,
               s.TIMELOAD_NEXT,
               s.AVSPEED,
               s.TASKDATE,
               s.SHIFT,
               TRIP,
               SUM(st.MOVELENGTH / 1000) / SUM(( st.MOVELENGTH / 1000 ) / st.AVGSPEED) AVSPEED_EMPTY
        FROM   (SELECT VEHID,
                       VEHCODE,
                       SHOVID,
                       WORKTYPE,
                       TIMELOAD,
                       TIMEUNLOAD,
                       NVL(LEAD(TIMELOAD)
                             over (
                               PARTITION BY VEHCODE
                               ORDER BY TIMELOAD), B.SDATETO) TIMELOAD_NEXT,
                       AVSPEED,
                       GETCURSHIFTDATE(0, TIMELOAD)           taskdate,
                       GETCURSHIFTNUM(0, TIMELOAD)            shift,
                       1                                      trip
                FROM   VEHTRIPS
                       inner join b
                               ON TIMELOAD BETWEEN B.SDATEFROM AND B.SDATETO
                                  AND TIMEUNLOAD BETWEEN B.SDATEFROM AND B.SDATETO
                WHERE  SHOVID NOT LIKE '%Неопр.%' AND AVSPEED BETWEEN 10 AND 50
                       AND ( TRIM(UPPER(WORKTYPE)) NOT LIKE ( '%ВКП СКАЛА%' )
                             AND TRIM(UPPER(WORKTYPE)) NOT LIKE ( '%ВКП ЩЕБЕНЬ%' )
                             AND TRIM(UPPER(WORKTYPE)) NOT LIKE ( '%ВКП%' )
                             AND TRIM(UPPER(WORKTYPE)) NOT LIKE ( '%ПСП%' )
                             AND TRIM(UPPER (UNLOADID)) NOT LIKE ( '%АВТОДОРОГА%' )
                             AND TRIM(UPPER (UNLOADID)) NOT LIKE ( '%ВНЕ ОТВАЛА%' )
                             AND TRIM(UPPER (UNLOADID)) NOT LIKE ( '%ДОРОГА ОБЩЕГО ПОЛЬЗОВАНИЯ%' )
                             AND TRIM(UPPER (WORKTYPE)) NOT LIKE ( '%ВСП%' )
                             AND TRIM(UPPER (WORKTYPE)) NOT LIKE ( '%СНЕГ%' ) )
                       AND ( B.TECHID = 'Все'
                              OR VEHID = B.TECHID )
                ORDER  BY LENGTH(VEHID),
                          VEHID,
                          TIMELOAD) s
               left join SIMPLETRANSITIONS st
                      ON st.VEHCODE = s.VEHCODE
                         AND st.AVGSPEED > 5
                         AND ( st.TIMEGO BETWEEN s.TIMEUNLOAD AND s.TIMELOAD_NEXT )
                         AND st.MOVELENGTH > 0
        GROUP  BY s.VEHID,
                  s.VEHCODE,
                  s.WORKTYPE,
                  s.SHOVID,
                  s.TIMELOAD,
                  s.TIMEUNLOAD,
                  s.TIMELOAD_NEXT,
                  s.AVSPEED,
                  s.TASKDATE,
                  s.SHIFT,
                  s.TRIP)q
       left join SHIFTTASKS stk
              ON stk.TASKDATE = q.TASKDATE
                 AND stk.SHIFT = q.SHIFT
                 AND stk.VEHID = q.VEHID
       left join DRIVERS d
              ON stk.TABELNUM = d.TABELNUM
ORDER  BY q.TASKDATE,
          TIMELOAD 
"""


def test():

    def f_query(_args):
        try:
            cx_Oracle.init_oracle_client(lib_dir=r"../static_external/instantclient_21_9_lite")
        except Exception as error:
            pass

        with cx_Oracle.connect("DISPATCHER/disp@172.30.23.16/PITENEW") as __connection:
            with __connection.cursor() as __cursor:
                __cursor.execute(query, _args)
                trips: list[tuple] = __cursor.fetchall()

        trips_raw = [
            {
                "vehid": i[0],
                "shovid": i[1],
                "fio": i[2],
                "taskdate": i[3],
                "shift": i[4],
                "trip": i[5],
                "worktype": i[6],
                "timeload": i[7],
                "timeunload": i[8],
                "avgloadspeed": i[9],
                "avgemptyspeed": i[10],
                "avgspeed": i[11],
            }
            for i in trips
        ]
        # print(trips_raw)

        # todo сгруппировать по каждому самосвалу
        data1 = {}
        for trip in trips_raw:
            tech_id = f"{trip['vehid']}"
            try:
                data1[tech_id] = [*data1[tech_id], trip]
            except Exception as error:
                data1[tech_id] = [trip]
        # print(data1)

        # print(data2)

        data2 = []
        for k_1, v_1 in data1.items():
            if len(v_1) <= 1:
                continue
            avg_speed_all = 0
            # print(v_1)
            avg_speed_last = v_1[-1]["avgloadspeed"]
            for i in v_1[:-1:]:
                avg_speed_all += i["avgloadspeed"]
            avg_speed_all = round(avg_speed_all / len(v_1[:-1:]), 2)
            data2.append({
                "taskdate": v_1[0]["taskdate"], "shift": v_1[0]["shift"],
                "vehid": k_1, "fio": v_1[0]["fio"],
                "avg_speed_all": avg_speed_all, "avg_speed_last": avg_speed_last, "count": len(v_1),
                "timeload": v_1[-1]["timeload"], "timeunload": v_1[-1]["timeunload"],
            })
        # print(data2)

        data3 = sorted(data2, key=lambda x: x["vehid"], reverse=False)

        return data3

    param_date_from = datetime.datetime.strptime("2023-05-01", "%Y-%m-%d")
    param_date_to = datetime.datetime.strptime("2023-06-09", "%Y-%m-%d")

    dt = (param_date_to - param_date_from).days
    # print(dt)
    data = []
    for i in range(0, dt + 1):
        param_date = param_date_from + datetime.timedelta(days=i)

        night = f_query(dict(param_date=param_date, param_shift=1, param_select_tech_id="Все"))
        data.extend(night)

        day = f_query(dict(param_date=param_date, param_shift=2, param_select_tech_id="Все"))
        data.extend(day)

        print(f"{param_date} completed")

    for j in data:
        print(j)

    print("sql completed")

    workbook: Workbook = openpyxl.Workbook()
    worksheet: Worksheet = workbook.active

    for row in range(1, len(data) + 1):
        worksheet.cell(row=row, column=1, value=data[row-1]["taskdate"])
        worksheet.cell(row=row, column=2, value=data[row-1]["shift"])
        worksheet.cell(row=row, column=3, value=data[row-1]["vehid"])
        worksheet.cell(row=row, column=4, value=data[row-1]["fio"])
        worksheet.cell(row=row, column=5, value=data[row-1]["avg_speed_all"])
        worksheet.cell(row=row, column=6, value=data[row-1]["avg_speed_last"])
        worksheet.cell(row=row, column=7, value=data[row-1]["count"])
        worksheet.cell(row=row, column=8, value=data[row-1]["timeload"])
        worksheet.cell(row=row, column=9, value=data[row-1]["timeunload"])

    workbook.save("data.xlsx")

test()
