"""Web-platform"""
import random
import re
import socket
import threading
from collections import Counter
from functools import wraps
import time
import sqlite3
import datetime
import os
import hashlib
import asyncio
import aiohttp
import aiofiles
import openpyxl
import oracledb
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from fastapi.responses import Response
from fastapi import FastAPI, Request
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import HTMLResponse, RedirectResponse, JSONResponse
import queries


#


class Constants:
    DEBUG = True
    IS_SERVER = False if str(socket.gethostname()).strip() == "KGPPC-ABN" else True
    LOGGING = True
    LOGGING_TO_CONSOLE = True
    LOGGING_TO_FILE = True
    LOGGING_TO_DATABASE = False
    LOGGING_RESPONSE = False
    PORT = 8000
    jwt_token_lifetime_seconds = 24 * 60 * 60
    oracle_db_connection_string = "DISPATCHER/disp@172.30.23.16/PITENEW"

    base_folder: str = os.path.join(os.path.dirname(__file__), "..")
    templates_path: str = base_folder
    static_path: str = os.path.join(base_folder, "backend/react/build/static")
    static_root: str = "/static"
    media_path: str = os.path.join(base_folder, "backend/media")
    media_root: str = "/media"


class CacheServer:
    def __init__(self):
        self.data: dict[str, any] = {}
        self.clear_tasks: dict[str, datetime.datetime] = {}

    def get(self, key: str = None, request: Request = None, query: callable = None, timeout: float = 1.0):
        if key is None:
            key = f"{request.url}_{request.method}_{''.join([str(x) for x in request.query_params.values()])}"
        value = self.data.get(key, None)
        if value is None and query:
            value = query()()
            self.set(key=key, value=value, timeout=timeout)

        return value

    def set(self, key: str, value: any, timeout: float = 1.0):
        # Задача очистки кэша не останавливается при установке нового кэша

        self.data[key] = value
        threading.Thread(
            target=self.clear_cache,
            args=(),
            kwargs={
                "key": key,
                "timeout": timeout,
            },
        ).start()

    def clear(self, key: str):
        if self.data.get(key, None) is not None:
            del self.data[key]

    def reset(self):
        self.data = {}

    def clear_cache(self, key: str, timeout: float = 1.0):
        if 0.0 >= timeout > 99999:
            return

        # запоминаю время начала этой задачи
        date_time = datetime.datetime.now()
        self.clear_tasks[key] = date_time

        time.sleep(timeout)

        # отменяю задачу очистки кэша, если он был перезаписан
        if self.clear_tasks.get(key, None) != date_time:
            return

        if self.data.get(key, None) is not None:
            del self.data[key]


cache_client = CacheServer()
app = FastAPI(
    title="ChimichangApp",
    description="""
ChimichangApp API helps you do awesome stuff. 🚀

## Items

You can **read items**.

## Users

You will be able to:

* **Create users** (_not implemented_).
* **Read users** (_not implemented_).
""",
    summary="Deadpool's favorite app. Nuff said.",
    version="0.0.1",
    terms_of_service="https://example.com/terms/",
    contact={
        "name": "Deadpoolio the Amazing",
        "url": "https://x-force.example.com/contact/",
        "email": "dp@x-force.example.com",
    },
    license_info={
        "name": "Apache 2.0",
        "url": "https://www.apache.org/licenses/LICENSE-2.0.html",
    },
)
app.mount(
    Constants.static_root,
    StaticFiles(directory=Constants.static_path),
    name="static",
)
app.mount(
    Constants.media_root,
    StaticFiles(directory=Constants.media_path),
    name="media",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "localhost:3000",
        "http://127.0.0.1:3000",
        "127.0.0.1:3000",
        "http://127.0.0.1:8000",
        "127.0.0.1:8000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
templates = Jinja2Templates(directory=Constants.templates_path)


@app.on_event("startup")
async def startup_event():
    print(f"\n.....server port={Constants.PORT} {datetime.datetime.now()} started.....\n\n\n")


@app.on_event("shutdown")
def shutdown_event():
    print(f"\n\n\n.....server port={Constants.PORT} {datetime.datetime.now()} stopped.....\n")


class Utils:
    @staticmethod
    def decorator(need_auth=False):
        def decorator(function: callable):
            @wraps(function)
            async def decorated_function(request: Request, *args, **kwargs):
                _error = None
                response = None
                try:
                    time_start_func = time.perf_counter()
                    response = await function(request, *args, **kwargs)
                    elapsed_time = round((time.perf_counter() - time_start_func), 2)
                    user_id = -1
                    if need_auth:
                        token_access: str | None = request.cookies.get("token_access", None)
                        if token_access is None:
                            raise Exception("Need Authorization!")
                        token: tuple = Utils.db_query_sqlite(
                            query="""SELECT user_id, datetime_elapsed FROM tokens WHERE token_access = ?""",
                            args=(token_access,),
                            many=False,
                        )
                        user_id: int = token[0]
                        datetime_elapsed: datetime.datetime = datetime.datetime.strptime(token[1], "%Y-%m-%d %H:%M:%S")
                        if datetime.datetime.now() > datetime_elapsed + datetime.timedelta(
                            hours=6, seconds=Constants.jwt_token_lifetime_seconds
                        ):
                            raise Exception("Token is terminate")
                        request.user_id = user_id
                    text = f"{str(datetime.datetime.now())[0:-5:1]}({elapsed_time}s) ({request.method})/{'/'.join(str(request.url).split('/')[3:])} || {response if Constants.LOGGING_RESPONSE else '[successfully response disabled]'}"
                except Exception as error:
                    _error = str(error)
                    user_id = -1
                    text = (
                        f"{str(datetime.datetime.now())[0:-5:1]} ({request.method})/{'/'.join(str(request.url).split('/')[3:])} || {error}"
                    )

                if Constants.LOGGING_TO_CONSOLE:
                    print("\n" + text)
                if Constants.LOGGING_TO_FILE:
                    async with aiofiles.open("static/log.txt", mode="a", encoding="utf-8") as file:
                        await file.write(text + "\n")
                if Constants.LOGGING_TO_DATABASE:
                    Utils.db_query_sqlite(
                        query="""INSERT INTO logs (user_id, ip, path, method, data) VALUES (?, ?, ?, ?, ?)""",
                        args=(
                            user_id,
                            request.headers.get("referer", "-"),
                            "/".join(str(request.url).split("/")[3:]),
                            request.method,
                            text.split(" || ")[-1],
                        ),
                    )
                if _error is not None:
                    raise Exception(_error)
                return response

            return decorated_function

        return decorator

    @staticmethod
    def db_query_sqlite(query: str, args=(), many=True, source="database.db") -> tuple | list[tuple] | None:
        with sqlite3.connect(source, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as connection:
            cursor = connection.cursor()
            cursor.execute(query, args)
            try:
                if many:
                    return cursor.fetchall()
                return cursor.fetchone()
            except Exception as _:
                return None

    @staticmethod
    def request_to_oracle(query: str, args: dict = None, many: bool = True) -> tuple | list[tuple] | None:
        _ = """
sudo su
mkdir -p /opt/oracle
cd /opt/oracle
wget https://download.oracle.com/otn_software/linux/instantclient/214000/instantclient-basic-linux.x64-21.4.0.0.0dbru.zip
unzip instantclient_21_4
apt install libaio1
echo /opt/oracle/instantclient_21_4 > /etc/old.so.conf.d/oracle-instantclient.conf
ldconfig
pip install cx_Oracle
exit
"""

        try:
            oracledb.init_oracle_client(
                lib_dir=r"C:\ADDITIONAL\web_platform\instantclient_21_9_lite"
                if Constants.IS_SERVER
                else r"D:\Program Files\projects\instantclient_21_9_lite"
            )
        except Exception as err:
            print(err)
            pass
        try:
            with oracledb.connect(Constants.oracle_db_connection_string) as connection:
                with connection.cursor() as cursor:
                    if args is None:
                        args = {}
                    cursor.execute(query, args)
                    if many:
                        return cursor.fetchall()
                    return cursor.fetchone()
        except Exception as err:
            raise err

    @staticmethod
    def get_selected_shift(date_time: datetime.datetime) -> int:
        if 8 <= date_time.hour < 20:
            return 2
        return 1

    @staticmethod
    def get_shift_datetime_begin(
        date_time: datetime.datetime,
    ) -> datetime.datetime:
        if Utils.get_selected_shift(date_time=date_time) == 1:
            date_time = date_time - datetime.timedelta(days=0 if date_time.hour > 19 else 1)
            date_time = date_time.replace(hour=20, minute=0, second=0, microsecond=0)
        else:
            date_time = date_time.replace(hour=8, minute=0, second=0, microsecond=0)
        return date_time

    @staticmethod
    def get_shift_datetime_end(
        date_time: datetime.datetime,
    ) -> datetime.datetime:
        if Utils.get_selected_shift(date_time=date_time) == 1:
            date_time = date_time + datetime.timedelta(days=1 if date_time.hour > 19 else 0)
            date_time = date_time.replace(hour=8, minute=0, second=0, microsecond=0)
        else:
            date_time = date_time.replace(hour=20, minute=0, second=0, microsecond=0)
        return date_time

    @staticmethod
    def get_selected_taskdate(
        date_time: datetime.datetime,
    ) -> datetime.datetime:
        date_time = Utils.get_shift_datetime_begin(date_time=date_time)

        # dt1 = datetime.datetime.strptime("2023-07-18T10:52:07.666666", "%Y-%m-%dT%H:%M:%S.%f")
        # dt1 = datetime.datetime.strptime("2023-07-18T07:52:07.666666", "%Y-%m-%dT%H:%M:%S.%f")

        return date_time.replace(hour=0, minute=0, second=0, microsecond=0)

    @staticmethod
    def get_all_dumptrucks(exclude: list[str] | None = None) -> list[str]:
        __rows_raw: list[tuple] = Utils.request_to_oracle(
            query="""SELECT VEHID FROM DUMPTRUCKS""",
            args={},
            many=True,
        )
        if exclude is None:
            return [str(x[0]) for x in __rows_raw]
        return [str(x[0]) for x in __rows_raw if str(x[0]) not in exclude]
        # with open(
        #     "../frontend/src/data/dumptrucks.json", mode="r", encoding="utf-8"
        # ) as f:
        #     _file: dict = json.load(f)
        #     _list: list[str] = [
        #         str(x["tech"]) for x in _file["list"] if str(x["type"]) == "самосвал"
        #     ]
        # if exclude is None:
        #     return _list
        # return [x for x in _list if x not in exclude]

    @staticmethod
    def get_all_shovels(exclude: list[str] | None = None) -> list[str]:
        __rows_raw: list[tuple] = Utils.request_to_oracle(
            query="""SELECT SHOVID FROM SHOVELS""",
            args={},
            many=True,
        )
        if exclude is None:
            return [str(x[0]) for x in __rows_raw]
        return [str(x[0]) for x in __rows_raw if str(x[0]) not in exclude]
        # with open("../frontend/src/data/shovels.json", mode="r", encoding="utf-8") as f:
        #     _file: dict = json.load(f)
        #     _list: list[str] = [
        #         str(x["tech"])
        #         for x in _file["list"]
        #         if str(x["type"]) == "экскаватор" or x["type"] == "экскаватор ЭШ"
        #     ]
        # if exclude is None:
        #     return _list
        # return [x for x in _list if x not in exclude]

    @staticmethod
    def get_all_auxes(exclude: list[str] | None = None) -> list[str]:
        __rows_raw: list[tuple] = Utils.request_to_oracle(
            query="""SELECT AUXID FROM AUXTECHNICS""",
            args={},
            many=True,
        )
        if exclude is None:
            return [str(x[0]) for x in __rows_raw]
        return [str(x[0]) for x in __rows_raw if str(x[0]) not in exclude]
        # with open("../frontend/src/data/auxes.json", mode="r", encoding="utf-8") as f:
        #     _file: dict = json.load(f)
        #     _list: list[str] = [str(x["tech"]) for x in _file["list"]]
        # if exclude is None:
        #     return _list
        # return [x for x in _list if x not in exclude]

    @staticmethod
    def create_encrypted_password(_random_chars="abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890", _length=8):
        password = ""
        for i in range(1, _length + 1):
            password += random.choice(_random_chars)
        return password


class DataBase:
    @staticmethod
    def create_tables():
        DataBase.create_users_table()
        DataBase.create_tokens_table()
        DataBase.create_logs_table()

    @staticmethod
    def create_users_table():
        #         Utils.db_query_sqlite(
        #             query="""
        # CREATE TABLE IF NOT EXISTS users
        # (
        # id INTEGER PRIMARY KEY AUTOINCREMENT,
        # email TEXT UNIQUE,
        # password TEXT,
        # position TEXT DEFAULT 'user',
        # datetime_joined TEXT DEFAULT CURRENT_TIMESTAMP
        # )
        # """
        #         )

        print("users: ", Utils.db_query_sqlite(query="SELECT * FROM users", many=True))
        # Utils.db_query_sqlite(
        #     query="""INSERT INTO users (email, password) VALUES (?, ?)""",
        #     args=("admin@gmail.com", "admin123"),
        # )
        # print("users: ", Utils.db_query_sqlite(query="SELECT * FROM users", many=True))

    @staticmethod
    def create_tokens_table():
        #         Utils.db_query_sqlite(
        #             query="""
        # CREATE TABLE IF NOT EXISTS tokens
        # (
        # id INTEGER PRIMARY KEY AUTOINCREMENT,
        # user_id INTEGER,
        # token_access TEXT,
        # datetime_elapsed TEXT DEFAULT CURRENT_TIMESTAMP
        # )
        # """
        #         )

        print("tokens: ", Utils.db_query_sqlite(query="SELECT * FROM tokens", many=True))

    @staticmethod
    def create_logs_table():
        #         Utils.db_query_sqlite(
        #             query="""
        # CREATE TABLE IF NOT EXISTS logs
        # (
        # id INTEGER PRIMARY KEY AUTOINCREMENT,
        # user_id INTEGER,
        # ip TEXT,
        # path TEXT,
        # method TEXT,
        # data TEXT,
        # created TEXT DEFAULT CURRENT_TIMESTAMP
        # )
        # """
        #         )

        print("logs: ", Utils.db_query_sqlite(query="SELECT * FROM logs", many=True))


class Views:
    class Api:
        class Target:
            @staticmethod
            @app.get("/api/target/monitoring/weight_loads", response_class=JSONResponse)
            @Utils.decorator(need_auth=False)
            async def target_monitoring_weight_loads(request: Request):
                now = datetime.datetime.now() - datetime.timedelta(days=1)
                param_shift: int = Utils.get_selected_shift(date_time=now)
                param_date: datetime.datetime = Utils.get_selected_taskdate(date_time=now)
                param_date_shift_begin: datetime.datetime = Utils.get_shift_datetime_begin(date_time=now)
                param_date_shift_end: datetime.datetime = Utils.get_shift_datetime_end(date_time=now)

                __parameters = {
                    "param_shift": param_shift,
                    "param_date": param_date,
                    "param_date_shift_begin": param_date_shift_begin,
                    "param_date_shift_end": param_date_shift_end,
                }

                def __query() -> list | tuple | dict | str:
                    list_of_shovels: list[dict] = []
                    for shov_id in [
                        "001",
                        "002",
                        "003",
                        "201",
                        "202",
                        "203",
                        "204",
                        "205",
                        "206",
                        "207",
                        "208",
                        "255",
                        "330",
                        "401",
                        "402",
                        "403",
                    ]:
                        # хоз номер экскаватора
                        dict_of_shovel: dict = {"shov_id": shov_id}

                        # фио машиниста
                        fio_shov_r: tuple = Utils.request_to_oracle(
                            args={
                                "param_date": param_date,
                                "param_shift": param_shift,
                                "param_shov_id": shov_id,
                            },
                            many=False,
                            query=queries.query_get_shovel_driver_by_taskdate_and_shift(),
                        )
                        fio_shov = "-" if fio_shov_r is None else fio_shov_r[0]

                        # рейсы с перегрузами и недогрузами
                        list_peregruz_and_nedogruz_r: list[tuple] = Utils.request_to_oracle(
                            args={
                                "param_shov_id": shov_id,
                                "param_date": param_date,
                                "param_shift": param_shift,
                                "param_shift_begin": param_date_shift_begin,
                                "param_shift_end": param_date_shift_end,
                                "param_low": 91,
                                "param_high": 100,
                            },
                            many=True,
                            query=queries.Target.monitoring_weight_loads(),
                        )
                        if fio_shov == "-" or len(list_peregruz_and_nedogruz_r) < 1:
                            continue
                        dict_of_shovel["fio_shov"] = fio_shov
                        list_peregruz_and_nedogruz = [
                            {
                                "vehid": x[0],
                                "worktype": x[1],
                                "timeload": x[2],
                                "weight": x[3],
                                "volume": x[4],
                                "hour_load": x[5],
                                "bucket_count": x[6],
                                "fio": x[7],
                            }
                            for x in list_peregruz_and_nedogruz_r
                        ]
                        dict_of_shovel["list_peregruz_and_nedogruz"] = list_peregruz_and_nedogruz
                        list_of_shovels.append(dict_of_shovel)

                    return {"data": list_of_shovels, "parameters": __parameters}

                return {"response": cache_client.get(query=lambda: __query, request=request, timeout=5)}

            @staticmethod
            @app.get("/api/target/report/weight_loads", response_class=JSONResponse)
            @Utils.decorator(need_auth=False)
            async def target_report_weight_loads(request: Request):
                param_date_from: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDateFrom"], "%Y-%m-%d")
                param_date_to: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDateTo"], "%Y-%m-%d")
                param_shift_from: int = int(request.query_params["paramShiftFrom"])
                param_shift_to: int = int(request.query_params["paramShiftTo"])
                __parameters = {
                    "param_date_from": param_date_from,
                    "param_date_to": param_date_to,
                    "param_shift_from": param_shift_from,
                    "param_shift_to": param_shift_to,
                    "param_weight_low": 91,
                    "param_weight_high": 100,
                }

                def __query() -> dict:
                    _trips_raw: list[tuple] = Utils.request_to_oracle(
                        args=__parameters,
                        many=True,
                        query=queries.Target.report_weight_loads(),
                    )
                    _trips_dict: list[dict] = [
                        {
                            "date": i[0],
                            "shift": i[1],
                            "shov_id": i[2],
                            "shov_driver": i[3],
                            "veh_id": i[4],
                            "veh_driver": i[5],
                            "area": i[6],
                            "worktype": i[7],
                            "unloadid": i[8],
                            "timeload": i[9],
                            "avspeed": i[10],
                            "weigth": i[11],
                        }
                        for i in _trips_raw
                    ]

                    def export_to_excel(_data: list[dict]) -> str:
                        key = Utils.create_encrypted_password(
                            _random_chars="abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890", _length=8
                        )
                        date = f"{time.strftime('%Y-%m-%d')}"
                        path = "media/data/temp/target_report_weight_loads"
                        file_name = f"недогрузы_и_перегрузы_{date}_{key}.xlsx"
                        workbook: Workbook = openpyxl.Workbook()
                        worksheet: Worksheet = workbook.active

                        # Delete old files
                        for root, dirs, files in os.walk(path, topdown=True):
                            for file in files:
                                try:
                                    date_file = str(file).strip().split(".")[0].strip().split("_")[-1]
                                    if date != date_file:
                                        os.remove(f"{path}/{file}")
                                except Exception as error:
                                    pass
                        # worksheet.cell(row=1, column=1, value="РАБОЧИЙ ЛИСТ")
                        # worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

                        titles = [
                            "Дата",
                            "Смена",
                            "Экскаватор",
                            "Машинист",
                            "Самосвал	",
                            "Оператор",
                            "Горизонт",
                            "Материал",
                            "Место разгрузки",
                            "Время погрузки",
                            "Ср. скорость",
                            "Масса",
                        ]
                        for col_idx, value in enumerate(titles, 1):
                            worksheet.cell(row=1, column=col_idx, value=value)

                        for row_idx, row in enumerate(_data, 2):
                            for col_idx, value in enumerate(row.values(), 1):
                                worksheet.cell(row=row_idx, column=col_idx, value=value)

                        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{len(_data)+1}"

                        side_think = Side(border_style="thin", color="FF808080")
                        # {'mediumDashDotDot', 'thin', 'dashed', 'mediumDashed', 'dotted', 'double', 'thick',
                        # 'medium', 'dashDot','dashDotDot', 'hair', 'mediumDashDot', 'slantDashDot'}
                        border_think = Border(top=side_think, left=side_think, right=side_think, bottom=side_think)
                        aligm_think = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
                        font_header = Font(name="Arial", size=12, bold=True)

                        # headers
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.font = font_header
                            cell.border = border_think

                        font_body = Font(name="Arial", size=10, bold=False)
                        green_fill = PatternFill(fgColor="F04141", fill_type="solid")

                        # body
                        for row_idx in range(2, worksheet.max_row + 1):
                            for col_idx in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.font = font_body
                                cell.alignment = aligm_think
                                cell.border = border_think
                                if col_idx == 12 and cell.value >= 100:
                                    cell.fill = green_fill

                        # set column width
                        for col_idx in range(1, worksheet.max_column + 1):
                            width: int = 1
                            for row_idx in range(1, worksheet.max_row + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                value: int = len(str(cell.value))
                                if value > width:
                                    width = value
                            worksheet.column_dimensions[get_column_letter(col_idx)].height = 1
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = round((width * 1.25), 3)
                            # worksheet.column_dimensions[get_column_letter(col_idx)].auto_size = True
                            worksheet.column_dimensions[get_column_letter(col_idx)].bestFit = True

                        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
                        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LETTER
                        worksheet.page_margins.left = 0.05
                        worksheet.page_margins.right = 0.05
                        worksheet.page_margins.header = 0.1
                        worksheet.page_margins.bottom = 0.2
                        worksheet.page_margins.footer = 0.2
                        worksheet.page_margins.top = 0.1
                        worksheet.print_options.horizontalCentered = True
                        # sheet.print_options.verticalCentered = True
                        worksheet.page_setup.fitToHeight = 1
                        worksheet.page_setup.scale = 100
                        worksheet.page_setup.fitToHeight = 1
                        worksheet.page_setup.fitToWidth = 1
                        # worksheet.protection.password = key + "_1"
                        # worksheet.protection.sheet = True
                        # worksheet.protection.enable()
                        full_path = f"{path}/{file_name}"
                        workbook.save(full_path)
                        return full_path

                    return {
                        "data": _trips_dict,
                        "path_to_excel_file": export_to_excel(_trips_dict),
                        "parameters": __parameters,
                    }

                return {"response": cache_client.get(query=lambda: __query, request=request, timeout=5)}

            @staticmethod
            @app.get("/api/target/report/avg_speed", response_class=JSONResponse)
            @Utils.decorator(need_auth=False)
            async def target_report_avg_speed(request: Request):
                param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDate"], "%Y-%m-%d")
                __parameters = {
                    "param_date": param_date,
                }

                def __query() -> list | dict | str:
                    def __query_get_day_and_night() -> list | tuple | dict | str:
                        night: list[dict] = __query_by_shift(shift=1)
                        day: list[dict] = __query_by_shift(shift=2)

                        night_dict = {}
                        for i in night:
                            night_dict[i["tech"].split("(")[0]] = i

                        day_dict = {}
                        for i in day:
                            day_dict[i["tech"].split("(")[0]] = i

                        night_day = []
                        for key, value_night in night_dict.items():
                            try:
                                value_day = day_dict[key]
                                night_day_dict = {
                                    "tech": f'{value_night["tech"]} | {value_day["tech"]}',
                                    "value": {
                                        "avgloadspeed": f'{value_night["value"]["avgloadspeed"]} | {value_day["value"]["avgloadspeed"]}',
                                        "avgemptyspeed": f'{value_night["value"]["avgemptyspeed"]} | {value_day["value"]["avgemptyspeed"]}',
                                        "avgspeed": f'{value_night["value"]["avgspeed"]} | {value_day["value"]["avgspeed"]}',
                                        "last_loadspeed": f'{value_night["value"]["last_loadspeed"]} | {value_day["value"]["last_loadspeed"]}',
                                        "last_emptyspeed": f'{value_night["value"]["last_emptyspeed"]} | {value_day["value"]["last_emptyspeed"]}',
                                        "last_speed": f'{value_night["value"]["last_speed"]} | {value_day["value"]["last_speed"]}',
                                        "count": f'{value_night["value"]["count"]} | {value_day["value"]["count"]}',
                                        "fio": f'{value_night["value"]["fio"]} | {value_day["value"]["fio"]}',
                                        "shovid": f'{value_night["value"]["shovid"]} | {value_day["value"]["shovid"]}',
                                    },
                                    "key": key,
                                }
                                night_day.append(night_day_dict)
                            except Exception as _:
                                night_day_dict = {
                                    "tech": f'{value_night["tech"]} | -',
                                    "value": {
                                        "avgloadspeed": f'{value_night["value"]["avgloadspeed"]} | -',
                                        "avgemptyspeed": f'{value_night["value"]["avgemptyspeed"]} | -',
                                        "avgspeed": f'{value_night["value"]["avgspeed"]} | -',
                                        "last_loadspeed": f'{value_night["value"]["last_loadspeed"]} | -',
                                        "last_emptyspeed": f'{value_night["value"]["last_emptyspeed"]} | -',
                                        "last_speed": f'{value_night["value"]["last_speed"]} | -',
                                        "count": f'{value_night["value"]["count"]} | -',
                                        "fio": f'{value_night["value"]["fio"]} | -',
                                        "shovid": f'{value_night["value"]["shovid"]} | -',
                                    },
                                    "key": key,
                                }
                                night_day.append(night_day_dict)
                        for key, value_day in day_dict.items():
                            try:
                                night_dict[key]
                            except Exception as _:
                                night_day_dict = {
                                    "tech": f'- | {value_day["tech"]}',
                                    "value": {
                                        "avgloadspeed": f'- | {value_day["value"]["avgloadspeed"]}',
                                        "avgemptyspeed": f'- | {value_day["value"]["avgemptyspeed"]}',
                                        "avgspeed": f'- | {value_day["value"]["avgspeed"]}',
                                        "last_loadspeed": f'- | {value_day["value"]["last_loadspeed"]}',
                                        "last_emptyspeed": f'- | {value_day["value"]["last_emptyspeed"]}',
                                        "last_speed": f'- | {value_day["value"]["last_speed"]}',
                                        "count": f'- | {value_day["value"]["count"]}',
                                        "fio": f'- | {value_day["value"]["fio"]}',
                                        "shovid": f'- | {value_day["value"]["shovid"]}',
                                    },
                                    "key": key,
                                }
                                night_day.append(night_day_dict)
                        night_day = sorted(night_day, key=lambda x: x["key"], reverse=False)
                        return night_day

                    def __query_by_shift(shift) -> list | tuple | dict | str:
                        trips: list[tuple] = Utils.request_to_oracle(
                            query=queries.query_analyse_avg_speed(),
                            args=dict(
                                param_date=param_date,
                                param_shift=shift,
                                param_select_tech_id="Все",
                            ),
                            many=True,
                        )

                        trips_raw = [
                            {
                                "vehid": i[0],
                                "shovid": i[1],
                                "fio": i[2],
                                "taskdate": i[3],
                                "shift": i[4],
                                "trip": i[5],
                                "worktype": i[6],
                                "timeload": i[7],
                                "timeunload": i[8],
                                "avgloadspeed": i[9],
                                "avgemptyspeed": i[10],
                                "avgspeed": i[11],
                            }
                            for i in trips
                        ]

                        # сгруппировать по каждому самосвалу
                        data1 = {}
                        for trip in trips_raw:
                            tech_id = f"{trip['vehid']}"
                            try:
                                data1[tech_id] = [*data1[tech_id], trip]
                            except Exception as _:
                                data1[tech_id] = [trip]

                        def kpd(_trips: list[dict]) -> dict:
                            _avgloadspeed = 0.0
                            _avgemptyspeed = 0.0
                            _avgspeed = 0.0
                            shovid = ""

                            if len(_trips) < 2:
                                return {
                                    "avgloadspeed": _trips[0]["avgloadspeed"],
                                    "avgemptyspeed": _trips[0]["avgemptyspeed"],
                                    "avgspeed": _trips[0]["avgspeed"],
                                    "last_loadspeed": _trips[0]["avgloadspeed"],
                                    "last_emptyspeed": _trips[0]["avgemptyspeed"],
                                    "last_speed": _trips[0]["avgspeed"],
                                    "count": len(_trips),
                                    "fio": _trips[0]["fio"],
                                    "shovid": f'{_trips[0]["shovid"]}(1р)',
                                }

                            count = len(_trips[:-1:])

                            for i in _trips[:-1:]:
                                _avgloadspeed += i["avgloadspeed"] if i["avgloadspeed"] else 0.0
                                _avgemptyspeed += i["avgemptyspeed"] if i["avgemptyspeed"] else 0.0
                                _avgspeed += i["avgspeed"] if i["avgspeed"] else 0.0
                                shovid += f'|{i["shovid"]}'

                            _avgloadspeed = round(_avgloadspeed / count, 2)
                            _avgemptyspeed = round(_avgemptyspeed / count, 2)
                            _avgspeed = round(_avgspeed / count, 2)
                            counter1 = Counter()
                            shovels = [x.strip() for x in str(shovid).split("|") if len(x) > 1]
                            for str1 in shovels:
                                counter1[str1] += 1
                            often = max(counter1, key=counter1.get)
                            shovid = f"{often}({counter1.get(often, 0)}р)"

                            return {
                                "avgloadspeed": _avgloadspeed,
                                "avgemptyspeed": _avgemptyspeed,
                                "avgspeed": _avgspeed,
                                "last_loadspeed": _trips[-1]["avgloadspeed"],
                                "last_emptyspeed": _trips[-1]["avgemptyspeed"],
                                "last_speed": _trips[-1]["avgspeed"],
                                "count": count + 1,
                                "fio": _trips[-1]["fio"],
                                "shovid": shovid,
                            }

                        data2 = []
                        for k, v in data1.items():
                            val = kpd(v)
                            data2.append({"tech": f"{k}({val['count']}р)", "value": val})

                        data3 = sorted(data2, key=lambda x: x["tech"], reverse=False)

                        return data3

                    def __query_speed_by_hours() -> list | tuple | dict | str:
                        trips: list[tuple] = Utils.request_to_oracle(
                            query=queries.query_analyse_avg_speed_by_hours(),
                            args=dict(param_date=param_date),
                            many=True,
                        )

                        trips_raw = [
                            {
                                "vehid": i[0],
                                "time_group": i[1],
                                "скорость груж.": i[2],
                                "empty": i[3],
                                "avg": i[4],
                            }
                            for i in trips
                        ]

                        # сгруппировать по каждому самосвалу
                        data1 = {}
                        for trip in trips_raw:
                            tech_id = f"{trip['vehid']}"
                            try:
                                data1[tech_id] = [*data1[tech_id], trip]
                            except Exception as _:
                                data1[tech_id] = [trip]
                        data2 = []
                        for k, v in data1.items():
                            data2.append({"tech": int(k), "value": v})

                        data3 = sorted(data2, key=lambda x: x["tech"], reverse=False)
                        # print(data1)

                        # for i in data3:
                        #     print(i)

                        return data3

                    return {
                        "data": __query_get_day_and_night(),
                        "lines": __query_speed_by_hours(),
                        "query": {"date": param_date},
                        "parameters": __parameters,
                    }

                return {"response": cache_client.get(query=lambda: __query, request=request, timeout=3)}

        class Stoppages:
            @staticmethod
            @app.get("/api/stoppages/report/aux_dvs", response_class=JSONResponse)
            @Utils.decorator(need_auth=False)
            async def stoppages_report_aux_dvs(request: Request):
                param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDate"], "%Y-%m-%d")
                param_shift: int = int(request.query_params["paramShift"])
                param_target: int = int(request.query_params["paramTarget"])
                param_select_tech_id: str = str(request.query_params["paramSelectTechId"])
                __parameters = {
                    "param_date": param_date,
                    "param_shift": param_shift,
                    "param_target": param_target,
                    "param_select_tech_id": param_select_tech_id,
                }

                def __query_by_aux_id(aux_id: str) -> list | dict | str:
                    _start_at = time.perf_counter()
                    __rows_raw: tuple = Utils.request_to_oracle(
                        query=queries.query_atc_get_aux_stoppages(),
                        args={
                            "param_date": param_date,
                            "param_shift": param_shift,
                            "param_select_tech_id": aux_id,
                        },
                        many=True,
                    )

                    # формирование списка словарей
                    events: list[dict[str, any]] = []
                    for event_raw in __rows_raw:
                        events.append(
                            {
                                "tech": str(event_raw[3]),
                                "date_time": event_raw[0],
                                "speed": float(event_raw[1]),
                                "fuel": int(event_raw[2]),
                            }
                        )

                    # очистка простоев определённой длины
                    stoppages: list[dict] = []
                    while True:
                        try:
                            last_index = 0
                            local_interval = []
                            for event in events:
                                last_index += 1
                                if event["speed"] == 0:
                                    if len(local_interval) != 0:
                                        last: datetime.datetime = local_interval[-1]["date_time"]
                                        cur: datetime.datetime = event["date_time"]
                                        if (last - cur).total_seconds() > 5 * 60:
                                            break
                                    local_interval.append(event)
                                else:
                                    if len(local_interval) == 0:
                                        pass
                                    else:
                                        break
                            high = local_interval[0]["date_time"]
                            lower = local_interval[-1]["date_time"]
                            diff = round((high - lower).total_seconds() / 60, 1)
                            events = events[last_index:]
                            if diff > param_target:
                                stoppages.append(
                                    {
                                        "tech": local_interval[0]["tech"],
                                        "from": lower,
                                        "to": high,
                                        "diff": diff,
                                    }
                                )
                        except Exception as _:
                            break

                    return stoppages[::-1]

                def __query() -> dict:
                    if param_select_tech_id.lower() == "все":
                        auxes: list[str] = Utils.get_all_auxes(["27", "219", "777", "2222", "3333"])
                        all_stoppages: list[dict] = []
                        for aux_id in auxes:
                            all_stoppages.extend(__query_by_aux_id(aux_id=aux_id))
                    else:
                        all_stoppages: list[dict] = __query_by_aux_id(aux_id=param_select_tech_id)
                    return {
                        "data": all_stoppages,
                        "parameters": __parameters,
                    }

                return {"response": cache_client.get(query=lambda: __query, request=request, timeout=10)}

            @staticmethod
            @app.get("/api/stoppages/report/veh_dvs", response_class=JSONResponse)
            @Utils.decorator(need_auth=False)
            async def stoppages_report_veh_dvs(request: Request):
                param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDate"], "%Y-%m-%d")
                param_shift: int = int(request.query_params["paramShift"])
                param_target: int = int(request.query_params["paramTarget"])
                param_select_tech_id: str = str(request.query_params["paramSelectTechId"])
                __parameters = {
                    "param_date": param_date,
                    "param_shift": param_shift,
                    "param_target": param_target,
                    "param_select_tech_id": param_select_tech_id,
                }

                def __query_by_veh_id(veh_id) -> list | dict | str:
                    __rows_raw: tuple = Utils.request_to_oracle(
                        query=queries.query_atc_get_veh_stoppages(),
                        args={
                            "param_date": param_date,
                            "param_shift": param_shift,
                            "param_select_tech_id": veh_id,
                        },
                        many=True,
                    )

                    # формирование списка словарей
                    events: list[dict[str, any]] = []
                    for event_raw in __rows_raw:
                        events.append(
                            {
                                "tech": str(event_raw[3]),
                                "date_time": event_raw[0],
                                "speed": float(event_raw[1]),
                                "fuel": int(event_raw[2]),
                            }
                        )

                    # очистка простоев определённой длины
                    stoppages: list[dict] = []
                    while True:
                        try:
                            last_index = 0
                            local_interval = []
                            for event in events:
                                last_index += 1
                                if event["speed"] == 0:
                                    if len(local_interval) != 0:
                                        last: datetime.datetime = local_interval[-1]["date_time"]
                                        cur: datetime.datetime = event["date_time"]
                                        if (last - cur).total_seconds() > 5 * 60:
                                            break
                                    local_interval.append(event)
                                else:
                                    if len(local_interval) == 0:
                                        pass
                                    else:
                                        break
                            high = local_interval[0]["date_time"]
                            lower = local_interval[-1]["date_time"]
                            diff = round((high - lower).total_seconds() / 60, 1)
                            events = events[last_index:]
                            if diff > param_target:
                                stoppages.append(
                                    {
                                        "tech": local_interval[0]["tech"],
                                        "from": lower,
                                        "to": high,
                                        "diff": diff,
                                        "fuel": f'{local_interval[-10]["fuel"]} - {local_interval[5]["fuel"]} = '
                                        f'{local_interval[-10]["fuel"] - local_interval[5]["fuel"]} ',
                                    }
                                )
                        except Exception as _:
                            break

                    return stoppages[::-1]

                def __query() -> dict:
                    if param_select_tech_id.lower() == "все":
                        dumptrucks: list[str] = Utils.get_all_dumptrucks(["100"])
                        all_stoppages: list[dict] = []
                        for veh_id in dumptrucks:
                            all_stoppages.extend(__query_by_veh_id(veh_id=veh_id))
                    else:
                        all_stoppages: list[dict] = __query_by_veh_id(veh_id=param_select_tech_id)
                    return {
                        "data": all_stoppages,
                        "parameters": __parameters,
                    }

                return {"response": cache_client.get(query=lambda: __query, request=request, timeout=10)}

        @staticmethod
        @app.get("/api", response_class=JSONResponse)
        @Utils.decorator(need_auth=True)
        async def api(request: Request):
            return {"data": f"{request.url} {request.method} OK"}

        @staticmethod
        @app.get("/api/reports/operuchet_dumptrucks", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def reports_operuchet_dumptrucks(request: Request):
            param_date_from = datetime.datetime.strptime(request.query_params.get("dateFrom"), "%Y-%m-%d")
            param_date_to = datetime.datetime.strptime(request.query_params.get("dateTo"), "%Y-%m-%d")
            param_shift_from = int(request.query_params.get("shiftFrom", 1))
            param_shift_to = int(request.query_params.get("shiftTo", 1))
            param_select_tech_id = int(request.query_params.get("selectTechId", 1))
            round_point = int(request.query_params.get("roundPoint", 1))

            def __query() -> list | dict | str:
                rows_raw = Utils.request_to_oracle(
                    query=queries.query_reports_operuchet(),
                    args={
                        "paramDateFrom": param_date_from,
                        "paramShiftFrom": param_shift_from,
                        "paramDateTo": param_date_to,
                        "paramShiftTo": param_shift_to,
                        "paramSelectTechId": param_select_tech_id,
                    },
                    many=True,
                )
                row_instances = []
                for i in rows_raw:
                    tr_mass = round(i[4] / 2.1, round_point)
                    tr_gruz = i[5] * tr_mass
                    sk_mass = round(i[6] / 2.8, round_point)
                    sk_gruz = i[7] * sk_mass
                    rih_mass = round(i[8] / 1.8, round_point)
                    rih_gruz = i[9] * rih_mass
                    prs_mass = round(i[10] / 1.4, round_point)
                    prs_gruz = i[11] * prs_mass
                    rud_mass = round(i[12] / 2.8, round_point)
                    rud_gruz = i[13] * rud_mass
                    summ_mass = round(tr_mass + sk_mass + rih_mass + prs_mass + rud_mass, round_point)
                    summ_proiz_gruz = round(tr_gruz + sk_gruz + rih_gruz + prs_gruz + rud_gruz, round_point)
                    if summ_mass > 0:
                        km = round(summ_proiz_gruz / summ_mass, round_point)
                    else:
                        km = 0
                    temp_row_instances = [
                        [
                            i[3],
                            "тр,км",
                            "тр",
                            round(i[4], round_point),
                            round(i[5], round_point),
                            2.1,
                            tr_mass,
                            summ_proiz_gruz,
                            summ_mass,
                            km,
                        ],
                        [
                            i[1],
                            "ск,км",
                            "ск",
                            round(i[6], round_point),
                            round(i[7], round_point),
                            2.8,
                            sk_mass,
                            "",
                            "",
                            "",
                        ],
                        [
                            "",
                            "рых,км",
                            "рых",
                            round(i[8], round_point),
                            round(i[9], round_point),
                            1.8,
                            rih_mass,
                            "",
                            "",
                            "",
                        ],
                        [
                            "",
                            "ПРС,км",
                            "прс",
                            round(i[10], round_point),
                            round(i[11], round_point),
                            1.4,
                            prs_mass,
                            "",
                            "",
                            "",
                        ],
                        [
                            "",
                            "руда,км",
                            "руд",
                            round(i[12], round_point),
                            round(i[13], round_point),
                            2.8,
                            rud_mass,
                            "",
                            "",
                            "",
                        ],
                    ]
                    row_instances.extend(temp_row_instances)
                return {"data": row_instances}

            _data = cache_client.get(
                key=f"{request.url}_{request.method}_{param_date_from.strftime('%H_%M_%S')}_{param_shift_from}_"
                f"{param_date_to.strftime('%H_%M_%S')}_{param_shift_to}_{param_select_tech_id}_{round_point}",
                timeout=4,
                query=lambda: __query,
            )
            return {"response": _data}

        @staticmethod
        @app.get("/api/reports/time_wait_to_load", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def reports_time_wait_to_load(request: Request):
            param_date = datetime.datetime.strptime(request.query_params.get("date"), "%Y-%m-%d")
            param_shift = int(request.query_params.get("shift", 1))
            param_target = int(request.query_params.get("target", 1))

            def __query() -> list | tuple | dict | str:
                errors: list[tuple] = Utils.request_to_oracle(
                    query=queries.query_errors_asd(),
                    args=dict(
                        param_date=param_date,
                        param_shift=param_shift,
                        param_target=param_target,
                    ),
                    many=True,
                )
                # errors = [
                #     ("ошибка 1", "описание 1"),
                #     ("ошибка 2", "описание 2"),
                #     ("ошибка 3", "описание 3"),
                # ]
                trips_raw = [
                    {
                        "title": i[0],
                        "description": i[1],
                    }
                    for i in errors
                ]
                return {
                    "data": trips_raw,
                    "query": {"date": param_date, "shift": param_shift},
                }

            _data = cache_client.get(
                key=f"{request.url}_{request.method}_{request.query_params.get('date', '-')}_{param_shift}",
                timeout=9,
                query=lambda: __query,
            )
            return {"response": _data}

        @staticmethod
        @app.get("/api/reports/time_wait_to_load", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def reports_time_wait_to_load(request: Request):
            param_date = datetime.datetime.strptime(request.query_params.get("date"), "%Y-%m-%d")
            param_shift = int(request.query_params.get("shift", 1))
            param_select_tech_id = str(request.query_params.get("selectTechId", "Все"))
            param_target = int(request.query_params.get("target", 1))

            def __query() -> list | tuple | dict | str:
                def __query1() -> list | tuple | dict | str:
                    trips: list[tuple] = Utils.request_to_oracle(
                        query=queries.query_time_wait_to_load(),
                        args=dict(
                            param_date=param_date,
                            param_shift=param_shift,
                            param_select_tech_id=param_select_tech_id,
                            param_target=param_target,
                        ),
                        many=True,
                    )

                    trips_raw = [
                        {
                            "vehid": j[0],
                            "timeFrom": j[2],
                            "timeTo": j[3],
                            "time": j[4],
                        }
                        for j in trips
                    ]

                    return trips_raw

                def __query2() -> list | tuple | dict | str:
                    idles: tuple = Utils.request_to_oracle(
                        query=queries.query_time_wait_to_load_avg(),
                        args=dict(
                            param_date=param_date,
                            param_shift=param_shift,
                            param_target=2,
                        ),
                        many=False,
                    )
                    idle = {
                        "sum_idles": idles[0],
                        "trips_idels": idles[1],
                        "trips_all": idles[2],
                        "avg_wait_all": idles[3],
                        "avg_wait": idles[4],
                    }
                    return idle

                try:
                    workbook: Workbook = openpyxl.load_workbook("static_external/grafick.xlsx")
                    worksheet: Worksheet = workbook.active

                    matrix = [x for x in worksheet.iter_rows(values_only=True, min_row=2)]
                    name = "Не назначено"
                    for i in matrix[::-1]:
                        if int(i[1]) == int(param_shift) and (i[0] <= param_date):
                            name = f'{i[2]} [от {i[0].strftime("%d.%m.%Y")} | {i[1]}]'
                            break
                except Exception as error:
                    name = str(error)

                return {
                    "data": __query1(),
                    "idle": __query2(),
                    "query": {"date": param_date, "shift": param_shift, "name": name},
                }

            _data = cache_client.get(
                key=f"{request.url}_{request.method}_{request.query_params.get('get', '-')}_{param_shift}_"
                f"{'All' if param_select_tech_id == 'Все' else param_select_tech_id}_",
                timeout=9,
                query=lambda: __query,
            )
            return {"response": _data}

        @staticmethod
        @app.get("/api/pto/operative_stoppages", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def pto_operative_stoppages(request: Request):
            def __query() -> list | tuple | dict | str:
                trips: list[tuple] = Utils.request_to_oracle(args={}, many=True, query=queries.query_operative_stoppages())

                trips_raw = [
                    {
                        "vehid": i[1],
                        "timestop": "".join(str(i[2]).split("T")),
                        "timego": "".join(str(i[3]).split("T")),
                        "continious": i[4],
                        "type": "длящийся" if str(i[5]) != "68" else "ожидание под погрузку",
                        "description": i[6],
                        "planned": i[7],
                    }
                    for i in trips
                ]
                return {"data": trips_raw}

            _data = cache_client.get(
                key=f"{request.url}_{request.method}",
                timeout=5,
                query=lambda: __query,
            )
            return {"response": _data}

        @staticmethod
        @app.get("/api/pto/analytic_tech", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def pto_analytic_tech(request: Request):
            def __query() -> list | tuple | dict | str:
                trips: list[tuple] = Utils.request_to_oracle(
                    args={"p_minLength": 0.0},
                    many=True,
                    query=queries.query_analyse_predictive(),
                )

                workbook: Workbook = openpyxl.load_workbook("static_external/table_norms.xlsx")
                worksheet: Worksheet = workbook["Самосвалы"]
                matrix_dumptrucks = {}
                for i in range(1 + 1, worksheet.max_row + 1):
                    matrix_dumptrucks[int(worksheet.cell(row=i, column=1).value)] = {
                        "Состояние, %": worksheet.cell(row=i, column=2).value,
                        "Ср. скорость гружённый": worksheet.cell(row=i, column=3).value,
                        "Ср. скорость порожний": worksheet.cell(row=i, column=4).value,
                        "Время разгрузки": worksheet.cell(row=i, column=5).value,
                        "Время ожидания": worksheet.cell(row=i, column=6).value,
                        "Погода, %": worksheet.cell(row=i, column=7).value,
                    }

                worksheet: Worksheet = workbook["Экскаваторы"]
                matrix_shovels = {}
                for i in range(1 + 1, worksheet.max_row + 1):
                    matrix_shovels[str(worksheet.cell(row=i, column=1).value)] = {
                        "Вскрыша скальная": worksheet.cell(row=i, column=2).value,
                        "Вскрыша рыхлая": worksheet.cell(row=i, column=3).value,
                        "Вскрыша транзитная": worksheet.cell(row=i, column=4).value,
                        "Руда скальная": worksheet.cell(row=i, column=5).value,
                        "ВКП скала": worksheet.cell(row=i, column=6).value,
                        "Среднее кол-во ковшей": worksheet.cell(row=i, column=7).value,
                    }

                def kpd_match(j):
                    movetime: datetime.datetime = j[7]
                    elapsed_time1 = round(
                        (movetime - datetime.datetime(2000, 1, 1)).total_seconds() / 60,
                        1,
                    )

                    veh = matrix_dumptrucks.get(
                        int(j[1])
                    )  # {'Состояние, %': 100, 'Ср. скорость гружённый': 20.5, 'Ср. скорость порожний': 22.5, 'Время разгрузки': 0.5, 'Время ожидания': 0}
                    # shov = matrix_shovels[str(j[2])]  # {'Вскрыша скальная': 2.5, 'Вскрыша рыхлая': 2.5, 'Вскрыша транзитная': 2.5, 'Руда скальная': 2.5, 'ВКП Скала': 2.5}

                    path = round(j[11] / veh["Ср. скорость гружённый"] * 60, 1)

                    full = round(path * (100 / veh["Состояние, %"]) * (100 / veh["Погода, %"]), 1)
                    kpd = round((full / elapsed_time1) * 100, 0)

                    return (
                        kpd,
                        f"({path}[путь]) * {veh['Состояние, %']}[состояние самосвала] * {veh['Погода, %']}[погода] = {full}[итого норма] | {round(elapsed_time1, 1)}[факт]",
                        0,
                    )

                trips_raw = [
                    {
                        "vehid": i[1],
                        "shovid": i[2],
                        "unloadid": i[3],
                        "worktype": i[4],
                        "timeload": i[5],
                        "timeunload": i[6],
                        "movetime": i[7],
                        "weigth": i[8],
                        "bucketcount": i[9],
                        "avspeed": i[10],
                        "length": i[11],
                        "unloadlength": i[12],
                        "loadheight": i[13],
                        "unloadheight": i[13],
                        "kpd": kpd_match(i)[0],
                        "detail": kpd_match(i)[1],
                    }
                    for i in trips
                ]

                # print(trips_raw[:5])

                trips_raw = sorted(trips_raw, key=lambda x: x["timeload"], reverse=True)

                # сгруппировать по каждому самосвалу
                data1 = {}
                for trips in trips:
                    tech_id = f"{trips[1]}"
                    try:
                        data1[tech_id] = [*data1[tech_id], trips]
                    except Exception as _:
                        data1[tech_id] = [trips]

                # сгруппировать по трём категориям
                data2 = {}
                for tech_id, trips in data1.items():
                    data2[tech_id] = {}
                    data2[tech_id]["last_trip"] = [trips[-1]]
                    for trip in trips:
                        # CORE
                        if (datetime.datetime.now() - trip[5]).total_seconds() < 1 * 60 * 60:
                            try:
                                data2[tech_id]["last_hour"] = [
                                    *data2[tech_id]["last_hour"],
                                    trip,
                                ]
                            except Exception as _:
                                data2[tech_id]["last_hour"] = [trip]
                        try:
                            data2[tech_id]["last_shift"] = [
                                *data2[tech_id]["last_shift"],
                                trip,
                            ]
                        except Exception as _:
                            data2[tech_id]["last_shift"] = [trip]

                # оценить каждый рейс и записать в новые поля
                data3 = {}
                for tech_id, trips in data2.items():
                    data3[tech_id] = {}
                    data3[tech_id]["ratings"] = {}
                    for type_par in ["last_trip", "last_hour", "last_shift"]:
                        try:
                            count = len(data2[tech_id][type_par])
                            if count > 0:
                                res = 0
                                elapsed_time = 0
                                for trip in data2[tech_id][type_par]:
                                    res += kpd_match(trip)[0]
                                    elapsed_time = ((trip[6] - trip[5]).total_seconds() / 60) + kpd_match(trip)[2]
                                if type_par == "last_trip":
                                    data3[tech_id]["ratings"][type_par] = {
                                        "rating": int(res / count),
                                        "count": int(elapsed_time),
                                    }
                                else:
                                    data3[tech_id]["ratings"][type_par] = {
                                        "rating": int(res / count),
                                        "count": count,
                                    }
                            else:
                                data3[tech_id]["ratings"][type_par] = {
                                    "rating": 0,
                                    "count": 0,
                                }
                        except Exception as _:
                            data3[tech_id]["ratings"][type_par] = {
                                "rating": 0,
                                "count": 0,
                            }

                # конвертация словаря в массив
                data4 = []
                for key, value in data3.items():
                    data4.append({"tech_id": key, **value})

                volumes_by_hours: list[any] = Utils.request_to_oracle(
                    args={},
                    many=True,
                    query=queries.query_get_volumes_by_hours(),
                )
                volumes_by_hours = [{"name": x[0], "объём": float(x[1]), "pv": 2400, "amt": 2400} for x in volumes_by_hours]

                volumes_by_category: list[any] = Utils.request_to_oracle(
                    args={},
                    many=True,
                    query=queries.query_get_volumes_by_category(),
                )
                volumes_by_category = [
                    {
                        "name": x[0] + f" | {x[2]} рейса(-ов)",
                        "объём": float(x[1]),
                        "pv": 2400,
                        "amt": 2400,
                    }
                    for x in volumes_by_category
                ]
                elapsed_time_raw: tuple[any] = Utils.request_to_oracle(args={}, many=False, query=queries.query_get_elapsed_time())

                # TODO #################################################################################################

                now = datetime.datetime.now() - datetime.timedelta(days=1)
                param_shift: int = Utils.get_selected_shift(date_time=now)
                param_date: datetime.datetime = Utils.get_selected_taskdate(date_time=now)
                param_date_shift_begin: datetime.datetime = Utils.get_shift_datetime_begin(date_time=now)
                param_date_shift_end: datetime.datetime = Utils.get_shift_datetime_end(date_time=now)

                list_of_shovels: list[dict] = []
                for shov_id in [
                    "001",
                    "002",
                    "003",
                    "201",
                    "202",
                    "203",
                    "204",
                    "205",
                    "206",
                    "207",
                    "208",
                    "255",
                    # "330",
                    "401",
                    "402",
                    "403",
                ]:
                    """
                    0. Все данные спрятаны циклом по accordion, но есть общий верхний групповой(после настроек)

                    1. Перегрузы и недогрузы - график столбцами, сортированный по времени
                    {
                        "fio_shov": "...",
                        "list": [...{"datetime": "datetime", "fio_veh", "veh": "145", "mass": 102, "volume": "30.0" "material_type": "вскрыша"},
                    ]
                    """

                    # хоз номер экскаватора
                    dict_of_shovel: dict = {"shov_id": shov_id}

                    # фио машиниста
                    fio_shov_r: tuple = Utils.request_to_oracle(
                        args={
                            "param_date": param_date,
                            "param_shift": param_shift,
                            "param_shov_id": shov_id,
                        },
                        many=False,
                        query=queries.query_get_shovel_driver_by_taskdate_and_shift(),
                    )
                    fio_shov = "-" if fio_shov_r is None else fio_shov_r[0]
                    if fio_shov == "-":
                        continue
                    dict_of_shovel["fio_shov"] = fio_shov

                    # рейсы с перегрузами и недогрузами
                    list_peregruz_and_nedogruz_r: list[tuple] = Utils.request_to_oracle(
                        args={
                            "param_shov_id": shov_id,
                            "param_date": param_date,
                            "param_shift": param_shift,
                            "param_shift_begin": param_date_shift_begin,
                            "param_shift_end": param_date_shift_end,
                            "param_low": 92,
                            "param_high": 99,
                        },
                        many=True,
                        query=queries.Target.monitoring_weight_loads(),
                    )
                    list_peregruz_and_nedogruz = [
                        {
                            "vehid": x[0],
                            "worktype": x[1],
                            "timeload": x[2],
                            "weight": x[3],
                            "volume": x[4],
                            "hour_load": x[5],
                            "bucket_count": x[6],
                            "fio": x[7],
                        }
                        for x in list_peregruz_and_nedogruz_r
                    ]
                    dict_of_shovel["list_peregruz_and_nedogruz"] = list_peregruz_and_nedogruz

                    # volumes_by_hours: list[any] = Utils.request_to_oracle(
                    #     args={},
                    #     many=True,
                    #     query=queries.query_get_volumes_by_category(),
                    # )
                    # volumes_by_category = [
                    #     {
                    #         "name": x[0] + f" | {x[2]} рейса(-ов)",
                    #         "объём": float(x[1]),
                    #         "pv": 2400,
                    #         "amt": 2400,
                    #     }
                    #     for x in volumes_by_category
                    # ]

                    list_of_shovels.append(dict_of_shovel)
                # TODO #################################################################################################

                extra = {
                    "volumes_by_hours": volumes_by_hours,
                    "volumes_by_category": volumes_by_category,
                    "trips_count": len(trips_raw),
                    "elapsed_time": elapsed_time_raw[2],
                }
                return {
                    "data": data4,
                    "trips": trips_raw,
                    "extra": extra,
                    "list_of_shovels": list_of_shovels,
                }

            _data = cache_client.get(
                key=f"{request.url}_{request.method}",
                timeout=5,
                query=lambda: __query,
            )
            return {"response": _data}

        @staticmethod
        @app.get("/api/analyse/predictive", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def analyse_predictive(request: Request):
            high_correct = int(request.query_params.get("highCorrect", 0))
            length_correct = int(request.query_params.get("lengthCorrect", 0))
            weather_correct = int(request.query_params.get("whetherCorrect", 100))
            min_length = float(request.query_params.get("minLength", 1.0))

            def __query() -> list | tuple | dict | str:
                trips: list[tuple] = Utils.request_to_oracle(
                    query=queries.query_analyse_predictive(),
                    args={"p_minLength": min_length},
                    many=True,
                )

                workbook: Workbook = openpyxl.load_workbook("static_external/table_norms.xlsx")
                worksheet: Worksheet = workbook["Самосвалы"]
                matrix_dumptrucks = {}
                for i in range(1 + 1, worksheet.max_row + 1):
                    matrix_dumptrucks[int(worksheet.cell(row=i, column=1).value)] = {
                        "Состояние, %": worksheet.cell(row=i, column=2).value,
                        "Ср. скорость гружённый": worksheet.cell(row=i, column=3).value,
                        "Ср. скорость порожний": worksheet.cell(row=i, column=4).value,
                        "Время разгрузки": worksheet.cell(row=i, column=5).value,
                        "Время ожидания": worksheet.cell(row=i, column=6).value,
                        "Погода, %": worksheet.cell(row=i, column=7).value,
                    }

                worksheet: Worksheet = workbook["Экскаваторы"]
                matrix_shovels = {}
                for i in range(1 + 1, worksheet.max_row + 1):
                    matrix_shovels[str(worksheet.cell(row=i, column=1).value)] = {
                        "Вскрыша скальная": worksheet.cell(row=i, column=2).value,
                        "Вскрыша рыхлая": worksheet.cell(row=i, column=3).value,
                        "Вскрыша транзитная": worksheet.cell(row=i, column=4).value,
                        "Руда скальная": worksheet.cell(row=i, column=5).value,
                        "ВКП скала": worksheet.cell(row=i, column=6).value,
                        "Среднее кол-во ковшей": worksheet.cell(row=i, column=7).value,
                    }

                def kpd_match(j):
                    movetime: datetime.datetime = j[7]
                    elapsed_time1 = round(
                        (movetime - datetime.datetime(2000, 1, 1)).total_seconds() / 60,
                        1,
                    )

                    veh = matrix_dumptrucks.get(int(j[1]))
                    # {'Состояние, %': 100, 'Ср. скорость гружённый': 20.5, 'Ср. скорость порожний': 22.5,
                    # 'Время разгрузки': 0.5, 'Время ожидания': 0}

                    # shov = matrix_shovels[str(j[2])]  # {'Вскрыша скальная': 2.5, 'Вскрыша рыхлая': 2.5,
                    # 'Вскрыша транзитная': 2.5, 'Руда скальная': 2.5, 'ВКП Скала': 2.5}

                    path = round(j[11] / veh["Ср. скорость гружённый"] * 60, 1)

                    length = round(j[11] * 10 * length_correct / 60, 1)
                    # print("length: ", j[11], length)

                    high = round(int(abs(int(j[13]) - int(j[14])) / 10) * high_correct / 60, 1)
                    full = round(
                        ((path + high + length) * (100 / veh["Состояние, %"]) * (100 / weather_correct)),
                        1,
                    )
                    # print(weather_correct, full)
                    kpd = round((full / elapsed_time1) * 100, 0)

                    return (
                        kpd,
                        f"({path}[путь] + {length}[корр. манёвров] + {high}[корр. высоты]) * "
                        f"{veh['Состояние, %']}[состояние самосвала] * "
                        f"{veh['Погода, %']}[погода] = {full}[итого норма] | {round(elapsed_time1, 1)}[факт]",
                        0,
                    )

                trips_raw = [
                    {
                        "vehid": i[1],
                        "shovid": i[2],
                        "unloadid": i[3],
                        "worktype": i[4],
                        "timeload": i[5],
                        "timeunload": i[6],
                        "movetime": i[7],
                        "promise": round((i[6] - i[5]).total_seconds() * 2 / 60, 1),
                        # "movetime_duble": round((int(str(i[7]).split(":")[1]) * 60 + int(str(i[7]).split(":")[2])) / 60 * 2, 1),
                        "weigth": i[8],
                        "bucketcount": i[9],
                        "avspeed": i[10],
                        "length": i[11],
                        "unloadlength": i[12],
                        "loadheight": i[13],
                        "unloadheight": i[14],
                        "kpd": kpd_match(i)[0],
                        "detail": kpd_match(i)[1],
                    }
                    for i in trips
                ]

                trips_raw = sorted(trips_raw, key=lambda x: x["timeload"], reverse=True)

                # сгруппировать по каждому самосвалу
                data1 = {}
                for trips in trips:
                    tech_id = f"{trips[1]}"
                    try:
                        data1[tech_id] = [*data1[tech_id], trips]
                    except Exception as _:
                        data1[tech_id] = [trips]

                # сгруппировать по трём категориям
                data2 = {}
                for tech_id, trips in data1.items():
                    data2[tech_id] = {}
                    data2[tech_id]["last_trip"] = [trips[-1]]
                    for trip in trips:
                        # CORE
                        if (datetime.datetime.now() - trip[5]).total_seconds() < 1 * 60 * 60:
                            try:
                                data2[tech_id]["last_hour"] = [
                                    *data2[tech_id]["last_hour"],
                                    trip,
                                ]
                            except Exception as _:
                                data2[tech_id]["last_hour"] = [trip]
                        try:
                            data2[tech_id]["last_shift"] = [
                                *data2[tech_id]["last_shift"],
                                trip,
                            ]
                        except Exception as _:
                            data2[tech_id]["last_shift"] = [trip]

                # оценить каждый рейс и записать в новые поля
                data3 = {}
                for tech_id, trips in data2.items():
                    data3[tech_id] = {}
                    data3[tech_id]["ratings"] = {}
                    for type_par in ["last_trip", "last_hour", "last_shift"]:
                        try:
                            count = len(data2[tech_id][type_par])
                            if count > 0:
                                res = 0
                                elapsed_time = 0
                                for trip in data2[tech_id][type_par]:
                                    res += kpd_match(trip)[0]
                                    elapsed_time = ((trip[6] - trip[5]).total_seconds() / 60) + kpd_match(trip)[2]
                                if type_par == "last_trip":
                                    data3[tech_id]["ratings"][type_par] = {
                                        "rating": int(res / count),
                                        "count": int(elapsed_time),
                                    }
                                else:
                                    data3[tech_id]["ratings"][type_par] = {
                                        "rating": int(res / count),
                                        "count": count,
                                    }
                            else:
                                data3[tech_id]["ratings"][type_par] = {
                                    "rating": 0,
                                    "count": 0,
                                }
                        except Exception as _:
                            data3[tech_id]["ratings"][type_par] = {
                                "rating": 0,
                                "count": 0,
                            }

                # конвертация словаря в массив
                data4 = []
                for key, value in data3.items():
                    data4.append({"tech_id": key, **value})

                return {"data": data4, "trips": trips_raw}

            _data = cache_client.get(
                key=f"{request.url}_{request.method}_{high_correct}_{length_correct}_{weather_correct}_{min_length}",
                timeout=2,
                query=lambda: __query,
            )
            return {"response": _data}

        @staticmethod
        @app.get("/api/events/drainage", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def events_drainage(request: Request):
            time_diff: int = int(request.query_params.get("timeDiff", "10"))

            def __query() -> list | dict | str:
                __rows_raw: tuple = Utils.request_to_oracle(
                    query=queries.query_drainage_status(),
                    args={"timeDiff": time_diff},
                    many=False,
                )
                # time.sleep(0.5)
                # __rows_raw = [
                #     datetime.datetime.now(),
                #     datetime.datetime.now() - datetime.timedelta(minutes=time_diff),
                #     666,
                #     444,
                #     222,
                #     222,
                # ]
                __rows_instances = {
                    "maxtime": __rows_raw[0],
                    "mintime": __rows_raw[1],
                    "maxfuel": __rows_raw[2],
                    "minfuel": __rows_raw[3],
                    "diffuel": int(__rows_raw[4]),
                    "difval": int(__rows_raw[5]),
                }
                return __rows_instances

            _data = cache_client.get(
                key=f"{request.url}_{request.method}",
                timeout=9,
                query=lambda: __query,
            )
            return {"response": {"data": _data}}

        @staticmethod
        @app.get("/api/events/dumptrucks", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def events_dumptrucks(request: Request):
            def __query() -> list | dict | str:
                __rows_raw = Utils.request_to_oracle(query=queries.query_vehtrips_status(), args={}, many=True)
                # time.sleep(0.5)
                # __rows_raw = [
                #     [x, datetime.datetime.now(), x, x, x, x, x] for x in range(130, 146)
                # ]
                __rows_instances = [
                    {
                        "vehid": i[0],
                        "time": i[1],
                        "x": i[2],
                        "y": i[3],
                        "weight": i[4],
                        "fuel": i[5],
                        "speed": i[6],
                    }
                    for i in __rows_raw
                ]
                return __rows_instances

            _data = cache_client.get(
                key=f"{request.url}_{request.method}",
                timeout=4,
                query=lambda: __query,
            )
            return {"response": {"data": _data}}

    class Base:
        @staticmethod
        @app.get("/api/captcha", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def captcha(request: Request):
            await asyncio.sleep(1.0)

            if request.query_params.get("id", None) != "666":
                raise Exception("Captcha error")

            return {"response": "OK"}

        @staticmethod
        @app.post("/api/register", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def register(request: Request):
            # from await axios.post("http://127.0.0.1:8000/api/token", new FormData().append("username", "admin"));
            # form = await request.form()

            # from await axios.post("http://127.0.0.1:8000/api/token", {username: "admin"})
            form = await request.json()
            email: str = str(form.get("email"))
            password: str = str(form.get("password"))
            if not re.match(r"[A-Za-z0-9._-]+@[A-Za-z0-9.-]+\.[A-Z|data-z]{2,}", string=email):
                raise Exception("Email not valid!")
            if not re.match(
                r"^(?=.*?[data-z])(?=.*?[A-Z])(?=.*?[0-9])(?=.*?[#?!@$%^&*-]).{8,}$",
                string=password,
            ):
                raise Exception("Password too weak!")
            Utils.db_query_sqlite(
                query="""INSERT INTO users (email, password) VALUES (?, ?)""",
                args=(email, hashlib.sha256(password.encode()).hexdigest()),
            )
            return "OK"

        @staticmethod
        @app.post("/api/login", response_class=JSONResponse)
        @Utils.decorator(need_auth=False)
        async def login(request: Request, response: Response):
            # get data from form
            form = await request.json()
            email: str = str(form.get("email"))
            password: str = hashlib.sha256(str(form.get("password")).encode()).hexdigest()

            # authenticate user
            user_id: tuple = Utils.db_query_sqlite(
                query="""SELECT id FROM users WHERE email = ? and password = ?""",
                args=(email, password),
                many=False,
            )
            if user_id is None:
                raise Exception("Login or password incorrect!")
            user_id: int = user_id[0]

            # create tokens
            token_access: str = hashlib.sha256(str(email + password + str(user_id)).encode()).hexdigest()
            Utils.db_query_sqlite(
                query="""INSERT INTO tokens (user_id, token_access) VALUES (?, ?)""",
                args=(user_id, token_access),
            )

            # set cookies
            response.set_cookie(
                "token_access",
                token_access,
                max_age=Constants.jwt_token_lifetime_seconds,
                httponly=True,
                secure=True,
            )

            return "OK"

        @staticmethod
        @app.get("/", response_class=HTMLResponse)
        @Utils.decorator(need_auth=False)
        async def root(request: Request):
            try:
                return templates.TemplateResponse("backend/react/build/index.html", {"request": request})
            except Exception as error:
                return templates.TemplateResponse(
                    "backend/templates/error.html",
                    {"request": request, "error": str(error.__str__)},
                )

        @staticmethod
        @app.post("/redirect", response_class=RedirectResponse)
        @Utils.decorator(need_auth=False)
        async def redirect(_: Request):
            return RedirectResponse(url=app.url_path_for("root"), status_code=303)

        # @staticmethod
        # @app.get("/{path:path}", response_class=RedirectResponse)
        # @Utils.decorator(need_auth=False)
        # async def default(request: Request, path: str):
        #     print(f"Path '{path}' is empty!")
        #     return RedirectResponse(url=app.url_path_for("root"), status_code=303)


class Debug:
    @staticmethod
    def get_tech_message_delay():
        """Проверка интервалов сообщений на всей технике"""

        # самосвалы
        def dumptrucks(t_from, t_to, too_low, too_high):
            for i in Utils.get_all_dumptrucks():
                __rows_raw: list[tuple] = Utils.request_to_oracle(
                    query="""
select ev.TIME from EVENTSTATEARCHIVE ev 
where ev.TIME BETWEEN :param_from AND :param_to and ev.VEHID = :param_veh_id
order by time asc
""",
                    args={
                        "param_from": datetime.datetime.strptime(t_from, "%d/%m/%Y %H:%M:%S"),
                        "param_to": datetime.datetime.strptime(t_to, "%d/%m/%Y %H:%M:%S"),
                        "param_veh_id": str(i),
                    },
                    many=True,
                )
                if __rows_raw is None or len(__rows_raw) < 1:
                    continue
                prev: datetime.datetime | None = None
                list_val = []
                for idx, row in enumerate(__rows_raw, 1):
                    curr: datetime.datetime = row[0]
                    if prev:
                        diff = (curr - prev).total_seconds()
                        list_val.append(diff)
                    prev = curr
                list_val = list(filter(lambda x: 30 > x > 0, list_val))
                if len(list_val) < 1:
                    continue
                val = round(sum(list_val) / len(list_val), 1)
                if val < too_low:
                    print(f"{i}: {val} | {len(list_val)} - РЕЖЕ")
                elif val > too_high:
                    print(f"{i}: {val} | {len(list_val)} - ЧАЩЕ")
                else:
                    print(f"{i}: {val} | {len(list_val)}")

        # auxes
        def auxes(t_from, t_to, too_low, too_high):
            for i in Utils.get_all_auxes():
                __rows_raw: list[tuple] = Utils.request_to_oracle(
                    query="""
select ev.TIME from AUXEVENTARCHIVE ev 
where ev.TIME BETWEEN :param_from AND :param_to and ev.AUXID = :param_veh_id
order by time asc
""",
                    args={
                        "param_from": datetime.datetime.strptime(t_from, "%d/%m/%Y %H:%M:%S"),
                        "param_to": datetime.datetime.strptime(t_to, "%d/%m/%Y %H:%M:%S"),
                        "param_veh_id": str(i),
                    },
                    many=True,
                )
                if __rows_raw is None or len(__rows_raw) < 1:
                    continue
                prev: datetime.datetime | None = None
                list_val = []
                for idx, row in enumerate(__rows_raw, 1):
                    curr: datetime.datetime = row[0]
                    if prev:
                        diff = (curr - prev).total_seconds()
                        list_val.append(diff)
                    prev = curr
                list_val = list(filter(lambda x: 30 > x > 0, list_val))
                if len(list_val) < 1:
                    continue
                val = round(sum(list_val) / len(list_val), 1)
                if val < too_low:
                    print(f"{i}: {val} | {len(list_val)} - РЕЖЕ")
                elif val > too_high:
                    print(f"{i}: {val} | {len(list_val)} - ЧАЩЕ")
                else:
                    print(f"{i}: {val} | {len(list_val)}")

        # shov
        def shov(t_from, t_to, too_low, too_high):
            for i in Utils.get_all_shovels():
                __rows_raw: list[tuple] = Utils.request_to_oracle(
                    query="""
select ev.TIME from SHOVEVENTSTATEARCHIVE ev 
where ev.TIME BETWEEN :param_from AND :param_to and ev.SHOVID = :param_veh_id
order by time asc
""",
                    args={
                        "param_from": datetime.datetime.strptime(t_from, "%d/%m/%Y %H:%M:%S"),
                        "param_to": datetime.datetime.strptime(t_to, "%d/%m/%Y %H:%M:%S"),
                        "param_veh_id": str(i),
                    },
                    many=True,
                )
                if __rows_raw is None or len(__rows_raw) < 1:
                    continue
                prev: datetime.datetime | None = None
                list_val = []
                for idx, row in enumerate(__rows_raw, 1):
                    curr: datetime.datetime = row[0]
                    if prev:
                        diff = (curr - prev).total_seconds()
                        list_val.append(diff)
                    prev = curr
                list_val = list(filter(lambda x: 30 > x > 0, list_val))
                if len(list_val) < 1:
                    continue
                val = round(sum(list_val) / len(list_val), 1)
                if val < too_low:
                    print(f"{i}: {val} | {len(list_val)} шт. РЕЖЕ")
                elif val > too_high:
                    print(f"{i}: {val} | {len(list_val)} шт. ЧАЩЕ")
                else:
                    print(f"{i}: {val} | {len(list_val)} шт.")

        t_from = "26/07/2023 20:00:00"
        t_to = "27/07/2023 08:00:00"

        dumptrucks(t_from, t_to, 4.0, 8.0)
        auxes(t_from, t_to, 8.0, 13.0)
        shov(t_from, t_to, 5.0, 15.0)


class Old:
    @staticmethod
    def test_create_new_post():
        async def test():
            async with aiohttp.ClientSession() as session:
                data = {"title": "TEST", "description": "TEST"}
                async with session.post("http://127.0.0.1:8000/create", data=data) as response:
                    print(response)

        asyncio.run(test())

    @staticmethod
    def test_api():
        async def test():
            async with aiohttp.ClientSession() as session:
                async with session.get("http://127.0.0.1:8000/api") as response:
                    data = await response.read()
                    print(response)
                    print(data)

        asyncio.run(test())

    @staticmethod
    def test_comments():
        with sqlite3.connect("database/database.db") as connection:
            cursor = connection.cursor()
            cursor.execute("""SELECT * FROM post_comments""")
            data = cursor.fetchall()
            print(len(data), data)


if __name__ == "__main__":
    """
    1. Написать чистую систему токенов
    2. Переписать систему токенов на ORM tortose or SqlAlchemy
    """

    if not Constants.IS_SERVER:
        Debug.get_tech_message_delay()

    # print(Utils.get_all_dumptrucks())
    # print(Utils.get_all_shovels(["Неопр."]))
    # print(Utils.get_all_auxes(["777"]))

    # DataBase.create_tables()
    # uvicorn.run(app, host="0.0.0.0", port=8000, workers=4)
    # Tests.test_api()
