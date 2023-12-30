# TODO download modules ##############################################################################################################################
import datetime
import hashlib
import json
import math
import os
import re
import time
from collections import Counter
import openpyxl
from fastapi import APIRouter
from fastapi import Request
from fastapi.responses import Response
from openpyxl.styles import Side, Border, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from starlette.responses import HTMLResponse, RedirectResponse, JSONResponse
from starlette.templating import Jinja2Templates
import asyncio
import aiosqlite

# TODO custom modules ################################################################################################################################
from . import constants, utils, queries, debug

#

router = APIRouter()
templates = Jinja2Templates(directory=constants.templates_path)


#


class Events:
    @staticmethod
    @router.get("/api/events/monitoring/drainage", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_events_monitoring_drainage(request: Request):
        param_time_diff: int = int(request.query_params["timeDiff"])
        __parameters = {
            "param_time_diff": param_time_diff,
        }

        def __query() -> list | dict | str:
            __rows_raw: tuple = utils.request_to_oracle(
                query=queries.Events.get_events_monitoring_drainage(),
                args=__parameters,
                many=False,
            )
            __rows_instances = {
                "maxtime": str(__rows_raw[0]),
                "mintime": str(__rows_raw[1]),
                "maxfuel": int(__rows_raw[2]),
                "minfuel": int(__rows_raw[3]),
                "diffuel": int(__rows_raw[4]),
                "difval": int(__rows_raw[5]),
            }
            return {"data": __rows_instances}

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}

    @staticmethod
    @router.get("/api/events/monitoring/dumptrucks", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_events_monitoring_dumptrucks(request: Request):
        def __query() -> dict:
            rows_raw: list[tuple] = utils.request_to_oracle(
                query=queries.Events.get_events_monitoring_dumptrucks(),
                args={},
                many=True,
            )
            rows_dict: list[dict] = [
                {
                    "vehid": i[0],
                    "time": i[1],
                    "x": i[2],
                    "y": i[3],
                    "weight": i[4],
                    "fuel": i[5],
                    "speed": i[6],
                }
                for i in rows_raw
            ]
            return {"data": rows_dict}

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}


class Speed:
    @staticmethod
    @router.get("/api/speed/monitoring/dumptrucks", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_speed_monitoring_dumptrucks(request: Request):
        now = datetime.datetime.now()
        cur_date = utils.get_selected_taskdate(date_time=now)
        cur_shift = utils.get_selected_shift(date_time=now)

        param_target_speed: float = float(request.query_params["param_target_speed"])
        __parameters = {"param_target_speed": param_target_speed}

        def __query() -> dict:
            def __query_speed_dumptrucks_for_now() -> list[dict]:
                # все рейсы за текущую смену сырые
                trips_raw: list[tuple] = utils.request_to_oracle(
                    query=queries.Speed.get_speed_monitoring_dumptrucks_for_now(),
                    args={},
                    many=True,
                )

                # все рейсы за текущую смену словарями
                trips_dict: list[dict] = [
                    {
                        "veh_id": i[0],
                        "fio": i[1],
                        "shov_id": i[2],
                        "avg_speed_full": i[3],
                        "avg_speed_empty": i[4],
                        "avg_speed_all": i[5],
                        "avg_weight": i[6],
                        "avg_length_full": i[7],
                        "avg_length_empty": i[8],
                        "avg_length_all": i[9],
                    }
                    for i in trips_raw
                ]

                # сгруппировать по каждому самосвалу
                data1 = {}
                for trip_dict in trips_dict:
                    veh_id = f"{trip_dict['veh_id']}"
                    if data1.get(veh_id, None) is None:
                        data1[veh_id] = [trip_dict]
                    else:
                        data1[veh_id].append(trip_dict)

                # функция подсчёта всех параметров
                def kpd(_trips: list[dict]) -> dict:
                    comment_description = utils.db_query_sqlite(
                        query=queries.Speed.get_speed_monitoring_dumptrucks_comments(),
                        args={
                            "veh_id": _trips[0]["veh_id"],
                            "taskdate": cur_date,
                            "shift": cur_shift,
                        },
                        many=False,
                    )
                    if comment_description:
                        comment_description = f"{comment_description[0]}"
                    else:
                        comment_description = ""
                    if len(_trips) < 2:
                        return {
                            "avg_speed_full": _trips[0]["avg_speed_full"],
                            "avg_speed_empty": _trips[0]["avg_speed_empty"],
                            "avg_speed_all": _trips[0]["avg_speed_all"],
                            "last_avg_speed_full": _trips[0]["avg_speed_full"],
                            "last_avg_speed_empty": _trips[0]["avg_speed_empty"],
                            "last_avg_speed_all": _trips[0]["avg_speed_all"],
                            "avg_weight": _trips[0]["avg_weight"],
                            "avg_length_full": _trips[0]["avg_length_full"],
                            "avg_length_empty": _trips[0]["avg_length_empty"],
                            "avg_length_all": _trips[0]["avg_length_all"],
                            "sum_trips": len(_trips),
                            "fio": _trips[0]["fio"],
                            "max_shov_id": f'{_trips[0]["shov_id"]}',
                            "max_shov_id_trips": 1,
                            "comment_description": comment_description,
                        }

                    count = len(_trips[:-1:])
                    _avg_speed_full = 0.0
                    _avg_speed_empty = 0.0
                    _avg_speed_all = 0.0
                    _avg_weight = 0.0
                    _avg_length_full = 0.0
                    _avg_length_empty = 0.0
                    _avg_length_all = 0.0
                    shovid = ""

                    for i in _trips[:-1:]:
                        _avg_speed_full += i["avg_speed_full"] if i["avg_speed_full"] else 0.0
                        _avg_speed_empty += i["avg_speed_empty"] if i["avg_speed_empty"] else 0.0
                        _avg_speed_all += i["avg_speed_all"] if i["avg_speed_all"] else 0.0
                        _avg_weight += i["avg_weight"] if i["avg_weight"] else 0.0
                        _avg_length_full += i["avg_length_full"] if i["avg_length_full"] else 0.0
                        _avg_length_empty += i["avg_length_empty"] if i["avg_length_empty"] else 0.0
                        _avg_length_all += i["avg_length_all"] if i["avg_length_all"] else 0.0
                        shovid += f'|{i["shov_id"]}'

                    _avg_speed_full = round(_avg_speed_full / count, 2)
                    _avg_speed_empty = round(_avg_speed_empty / count, 2)
                    _avg_speed_all = round(_avg_speed_all / count, 2)
                    _avg_weight = round(_avg_weight / count, 2)
                    _avg_length_full = round(_avg_length_full / count, 2)
                    _avg_length_empty = round(_avg_length_empty / count, 2)
                    _avg_length_all = round(_avg_length_all / count, 2)
                    counter1 = Counter()
                    shovels = [x.strip() for x in str(shovid).split("|") if len(x) > 1]
                    for str1 in shovels:
                        counter1[str1] += 1
                    often = max(counter1, key=counter1.get)

                    return {
                        "avg_speed_full": _avg_speed_full,
                        "avg_speed_empty": _avg_speed_empty,
                        "avg_speed_all": _avg_speed_all,
                        "last_avg_speed_full": _trips[-1]["avg_speed_full"],
                        "last_avg_speed_empty": _trips[-1]["avg_speed_empty"],
                        "last_avg_speed_all": _trips[-1]["avg_speed_all"],
                        "avg_weight": _avg_weight,
                        "avg_length_full": _avg_length_full,
                        "avg_length_empty": _avg_length_empty,
                        "avg_length_all": _avg_length_all,
                        "sum_trips": count + 1,
                        "fio": _trips[-1]["fio"],
                        "max_shov_id": often,
                        "max_shov_id_trips": counter1.get(often, 0),
                        "comment_description": comment_description,
                    }

                # подсчёт всех параметров
                data2: list[dict] = []
                for k, v in data1.items():
                    val = kpd(v)
                    val["veh_id"] = k
                    data2.append(val)

                # сортировка по номеру

                # отсортировать по эскаваторам, а потом по самосвалам
                data3 = sorted(data2, key=lambda x: (int(x["max_shov_id"]), int(x["veh_id"])), reverse=False)

                return data3

            def __query_speed_rating_for_shovel(
                speeds: list[dict],
            ) -> list[dict]:
                # сгруппировать по каждому экскаватору
                data1 = {}
                for trip_dict in speeds:
                    veh_id = f"{trip_dict['max_shov_id']}"
                    if data1.get(veh_id, None) is None:
                        data1[veh_id] = [trip_dict]
                    else:
                        data1[veh_id].append(trip_dict)

                # подсчёт всех параметров
                data2: list[dict] = []
                for k, v in data1.items():
                    val = sorted(v, key=lambda x: x["avg_speed_full"], reverse=False)[0]
                    comment_description = utils.db_query_sqlite(
                        query=queries.Speed.get_speed_monitoring_dumptrucks_comments(),
                        args={
                            "veh_id": val["veh_id"],
                            "taskdate": cur_date,
                            "shift": cur_shift,
                        },
                        many=False,
                    )
                    if comment_description:
                        val["comment_description"] = f"{comment_description[0]}"
                    else:
                        val["comment_description"] = ""
                    data2.append(val)

                # сортировка по экскаватору
                data3: list[dict] = sorted(data2, key=lambda x: int(x["max_shov_id"]), reverse=False)

                return data3

            table_speed_dumptrucks_for_now: list[dict] = __query_speed_dumptrucks_for_now()

            return {
                "table_speed_dumptrucks_for_now": table_speed_dumptrucks_for_now,
                "rating": __query_speed_rating_for_shovel(speeds=table_speed_dumptrucks_for_now),
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}

    @staticmethod
    @router.post("/api/speed/send_comment/dumptrucks", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def post_speed_send_comment_dumptrucks(request: Request):
        form_data = await request.json()

        veh_id = str(form_data["veh_id"])
        description = str(form_data["description"])

        now = datetime.datetime.now()
        cur_date = utils.get_selected_taskdate(date_time=now)
        cur_shift = utils.get_selected_shift(date_time=now)

        author = str(request.client.host)

        _id = utils.db_query_sqlite(
            query=queries.Speed.post_speed_send_comment_dumptrucks_get_id(),
            args={"veh_id": veh_id, "taskdate": cur_date, "shift": cur_shift},
            many=False,
        )
        if _id is None:
            utils.db_query_sqlite(
                query=queries.Speed.post_speed_send_comment_dumptrucks_insert(),
                args={
                    "veh_id": veh_id,
                    "description": description,
                    "taskdate": cur_date,
                    "shift": cur_shift,
                    "author": author,
                    "updated": now,
                },
                many=False,
            )
        else:
            utils.db_query_sqlite(
                query=queries.Speed.post_speed_send_comment_dumptrucks_update(),
                args={
                    "veh_id": veh_id,
                    "description": description,
                    "taskdate": cur_date,
                    "shift": cur_shift,
                    "author": author,
                    "updated": now,
                },
                many=False,
            )
        return {"response": "OK"}

    @staticmethod
    @router.get("/api/speed/report/dumptrucks", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_target_report_avg_speed(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["param_date"], "%Y-%m-%d")
        __parameters = {
            "param_date": param_date,
        }

        def __query() -> dict:
            def __get_by_shift(shift_num: int) -> dict:
                trips_raw: list[tuple] = utils.request_to_oracle(
                    query=queries.Develop.get_target_report_avg_speed_new(),
                    args={"param_date": param_date, "param_shift": shift_num},
                    many=True,
                )
                trips_dict: list[dict] = [
                    {
                        "veh_id": i[0],
                        "shov_id": i[1],
                        "fio": i[2],
                        "avg_speed_load": i[3],
                        "avg_speed_empty": i[4],
                        "avg_speed": i[5],
                        "avg_weight": i[6],
                        "avg_length_load": i[7],
                        "avg_length_unload": i[8],
                        "avg_length": i[9],
                        "avg_height": i[10],
                        "avg_worktype": i[11],
                        "avg_worktime": i[12],
                    }
                    for i in trips_raw
                ]

                # сгруппировать по каждому самосвалу
                data1 = {}
                for trip in trips_dict:
                    tech_id = int(trip["veh_id"])
                    try:
                        data1[tech_id].append(trip)
                    except Exception as _:
                        data1[tech_id] = [trip]

                def kpd(_trips: list[dict]) -> dict:
                    trips_count = len(_trips)
                    veh_id = _trips[0]["veh_id"]
                    fio = _trips[0]["fio"]

                    comment_description = utils.db_query_sqlite(
                        query=queries.Speed.get_speed_monitoring_dumptrucks_comments(),
                        args={
                            "veh_id": veh_id,
                            "taskdate": param_date,
                            "shift": shift_num,
                        },
                        many=False,
                    )
                    if comment_description:
                        comment_description = f"{comment_description[0]}"
                    else:
                        comment_description = ""

                    if len(_trips) < 2:
                        return {
                            "veh_id": veh_id,
                            "fio": fio,
                            "trips_count": trips_count,
                            "shov_id": f'{_trips[0]["shov_id"]}',
                            "shov_id_trips_count": 1,
                            "avg_speed_load": _trips[0]["avg_speed_load"],
                            "avg_speed_empty": _trips[0]["avg_speed_empty"],
                            "avg_speed": _trips[0]["avg_speed"],
                            "last_avg_speed_load": _trips[0]["avg_speed_load"],
                            "last_avg_speed_empty": _trips[0]["avg_speed_empty"],
                            "last_avg_speed": _trips[0]["avg_speed"],
                            "avg_weight": _trips[0]["avg_weight"],
                            "avg_length_load": _trips[0]["avg_length_load"],
                            "avg_length_unload": _trips[0]["avg_length_unload"],
                            "avg_length": _trips[0]["avg_length"],
                            "avg_height": _trips[0]["avg_height"],
                            "avg_worktype": _trips[0]["avg_worktype"],
                            "avg_worktime": _trips[0]["avg_worktime"],
                            "comment_description": comment_description,
                        }

                    avg_speed_load = 0.0
                    avg_speed_empty = 0.0
                    avg_speed = 0.0
                    avg_weight = 0.0
                    avg_length_load = 0.0
                    avg_length_unload = 0.0
                    avg_length = 0.0
                    avg_height = 0.0
                    avg_worktime = 0.0

                    avg_worktype_counter = Counter()
                    shov_id_counter = Counter()

                    for i in _trips[:-1]:
                        avg_speed_load += i["avg_speed_load"]
                        avg_speed_empty += i["avg_speed_empty"]
                        avg_speed += i["avg_speed"]
                        avg_weight += i["avg_weight"]
                        avg_length_load += i["avg_length_load"]
                        avg_length_unload += i["avg_length_unload"]
                        avg_length += i["avg_length"]
                        avg_height += i["avg_height"]
                        avg_worktime += i["avg_worktime"]
                        shov_id_counter[i["shov_id"]] += 1
                        avg_worktype_counter[i["avg_worktype"]] += 1

                    shov_id = max(shov_id_counter, key=shov_id_counter.get)
                    avg_worktype = max(avg_worktype_counter, key=avg_worktype_counter.get)

                    def avg(val: int | float):
                        return round(val / (trips_count - 1), 2)

                    return {
                        "veh_id": veh_id,
                        "fio": fio,
                        "trips_count": trips_count,
                        "shov_id": shov_id,
                        "shov_id_trips_count": shov_id_counter[shov_id],
                        "avg_speed_load": avg(avg_speed_load),
                        "avg_speed_empty": avg(avg_speed_empty),
                        "avg_speed": avg(avg_speed),
                        "last_avg_speed_load": _trips[-1]["avg_speed_load"],
                        "last_avg_speed_empty": _trips[-1]["avg_speed_empty"],
                        "last_avg_speed": _trips[-1]["avg_speed"],
                        "avg_weight": avg(avg_weight),
                        "avg_length_load": avg(avg_length_load),
                        "avg_length_unload": avg(avg_length_unload),
                        "avg_length": avg(avg_length),
                        "avg_height": avg(avg_height),
                        "avg_worktime": avg(avg_worktime),
                        "avg_worktype": avg_worktype,
                        "comment_description": comment_description,
                    }

                data2 = [kpd(v) for k, v in data1.items()]

                # отсортировать по эскаваторам, а потом по самосвалам
                data3 = sorted(data2, key=lambda x: (int(x["shov_id"]), int(x["veh_id"])), reverse=False)

                to_excel: list[list] = []
                for i in data3:
                    to_excel.append(
                        [
                            param_date,
                            shift_num,
                            i["veh_id"],
                            i["fio"],
                            i["shov_id"],
                            i["avg_speed_load"],
                            i["avg_speed_empty"],
                            i["avg_speed"],
                            i["comment_description"],
                        ]
                    )

                titles: list[str] = [
                    "Дата",
                    "Смена",
                    "Самосвал",
                    "Оператор",
                    "Экскаватор",
                    "Ср. скор. груж.",
                    "Ср. скор. порожн.",
                    "Ср. скор. общ.",
                    "Комментарий",
                ]
                path_to_excel_file: str = utils.export_to_excel(
                    _data=to_excel,
                    _titles=titles,
                    _folder_name="speed",
                    _prefix="средняя_скорость_new",
                )

                return {"data": data3, "path_to_excel_file": path_to_excel_file}

            def __query_speed_by_hours() -> list | tuple | dict | str:
                trips_raw: list[tuple] = utils.request_to_oracle(
                    query=queries.Develop.get_target_report_avg_speed_by_hours(),
                    args={"param_date": param_date},
                    many=True,
                )

                trips_dict = [
                    {
                        "vehid": i[0],
                        "time_group": i[1],
                        "скорость груж.": i[2],
                        "empty": i[3],
                        "avg": i[4],
                    }
                    for i in trips_raw
                ]

                # сгруппировать по каждому самосвалу
                data1 = {}
                for trip in trips_dict:
                    tech_id = f"{trip['vehid']}"
                    try:
                        data1[tech_id] = [*data1[tech_id], trip]
                    except Exception as _:
                        data1[tech_id] = [trip]
                data2 = []
                for k, v in data1.items():
                    data2.append({"tech": int(k), "value": v})

                data3 = sorted(data2, key=lambda x: x["tech"], reverse=False)
                return data3

            return {
                "night": __get_by_shift(1),
                "day": __get_by_shift(2),
                "lines": __query_speed_by_hours(),
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}


class Pto:
    @staticmethod
    @router.get("/api/pto/monitoring/norm_trips", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_pto_monitorin_norm_trips(request: Request):
        param_high_correct: int = int(request.query_params["param_high_correct"])
        param_length_correct: int = int(request.query_params["param_length_correct"])
        param_weather_correct: int = int(request.query_params["param_weather_correct"])
        param_min_length: float = float(request.query_params["param_min_length"])
        __parameters = {
            "param_high_correct": param_high_correct,
            "param_length_correct": param_length_correct,
            "param_weather_correct": param_weather_correct,
            "param_min_length": param_min_length,
        }

        def __query() -> list | tuple | dict | str:
            trips: list[tuple] = utils.request_to_oracle(
                query=queries.Pto.get_pto_monitoring_norm_trips(),
                args={"param_min_length": param_min_length},
                many=True,
            )

            workbook: Workbook = openpyxl.load_workbook("static_external/table_norms.xlsx")
            worksheet: Worksheet = workbook["Самосвалы"]
            matrix_dumptrucks = {}
            for i in range(1 + 1, worksheet.max_row + 1):
                matrix_dumptrucks[int(worksheet.cell(row=i, column=1).value)] = {
                    "Состояние, %": worksheet.cell(row=i, column=2).value,
                    "Ср. скорость гружённый": worksheet.cell(row=i, column=3).value,
                    "Ср. скорость порожний": worksheet.cell(row=i, column=4).value,
                    "Время разгрузки": worksheet.cell(row=i, column=5).value,
                    "Время ожидания": worksheet.cell(row=i, column=6).value,
                    "Погода, %": worksheet.cell(row=i, column=7).value,
                }

            worksheet: Worksheet = workbook["Экскаваторы"]
            matrix_shovels = {}
            for i in range(1 + 1, worksheet.max_row + 1):
                matrix_shovels[str(worksheet.cell(row=i, column=1).value)] = {
                    "Вскрыша скальная": worksheet.cell(row=i, column=2).value,
                    "Вскрыша рыхлая": worksheet.cell(row=i, column=3).value,
                    "Вскрыша транзитная": worksheet.cell(row=i, column=4).value,
                    "Руда скальная": worksheet.cell(row=i, column=5).value,
                    "ВКП скала": worksheet.cell(row=i, column=6).value,
                    "Среднее кол-во ковшей": worksheet.cell(row=i, column=7).value,
                }

            def kpd_match(j):
                movetime: datetime.datetime = j[7]
                elapsed_time1 = round(
                    (movetime - datetime.datetime(2000, 1, 1)).total_seconds() / 60,
                    1,
                )

                veh = matrix_dumptrucks.get(int(j[1]))
                # {'Состояние, %': 100, 'Ср. скорость гружённый': 20.5, 'Ср. скорость порожний': 22.5,
                # 'Время разгрузки': 0.5, 'Время ожидания': 0}

                # shov = matrix_shovels[str(j[2])]  # {'Вскрыша скальная': 2.5, 'Вскрыша рыхлая': 2.5,
                # 'Вскрыша транзитная': 2.5, 'Руда скальная': 2.5, 'ВКП Скала': 2.5}

                path = round(j[11] / veh["Ср. скорость гружённый"] * 60, 1)

                length = round(j[11] * 10 * param_length_correct / 60, 1)
                # print("length: ", j[11], length)

                high = round(int(abs(int(j[13]) - int(j[14])) / 10) * param_high_correct / 60, 1)
                full = round(
                    ((path + high + length) * (100 / veh["Состояние, %"]) * (100 / param_weather_correct)),
                    1,
                )
                # print(weather_correct, full)
                kpd = round((full / elapsed_time1) * 100, 0)

                return (
                    kpd,
                    f"({path}[путь] + {length}[корр. манёвров] + {high}[корр. высоты]) * "
                    f"{veh['Состояние, %']}[состояние самосвала] * "
                    f"{veh['Погода, %']}[погода] = {full}[итого норма] | {round(elapsed_time1, 1)}[факт]",
                    0,
                )

            trips_raw = [
                {
                    "vehid": i[1],
                    "shovid": i[2],
                    "unloadid": i[3],
                    "worktype": i[4],
                    "timeload": i[5],
                    "timeunload": i[6],
                    "movetime": i[7],
                    "promise": round((i[6] - i[5]).total_seconds() * 2 / 60, 1),
                    # "movetime_duble": round((int(str(i[7]).split(":")[1]) * 60 + int(str(i[7]).split(":")[2])) / 60 * 2, 1),
                    "weigth": i[8],
                    "bucketcount": i[9],
                    "avspeed": i[10],
                    "length": i[11],
                    "unloadlength": i[12],
                    "loadheight": i[13],
                    "unloadheight": i[14],
                    "kpd": kpd_match(i)[0],
                    "detail": kpd_match(i)[1],
                }
                for i in trips
            ]

            trips_raw = sorted(trips_raw, key=lambda x: x["timeload"], reverse=True)

            # сгруппировать по каждому самосвалу
            data1 = {}
            for trips in trips:
                tech_id = f"{trips[1]}"
                try:
                    data1[tech_id] = [*data1[tech_id], trips]
                except Exception as _:
                    data1[tech_id] = [trips]

            # сгруппировать по трём категориям
            data2 = {}
            for tech_id, trips in data1.items():
                data2[tech_id] = {}
                data2[tech_id]["last_trip"] = [trips[-1]]
                for trip in trips:
                    # CORE
                    if (datetime.datetime.now() - trip[5]).total_seconds() < 1 * 60 * 60:
                        try:
                            data2[tech_id]["last_hour"] = [
                                *data2[tech_id]["last_hour"],
                                trip,
                            ]
                        except Exception as _:
                            data2[tech_id]["last_hour"] = [trip]
                    try:
                        data2[tech_id]["last_shift"] = [
                            *data2[tech_id]["last_shift"],
                            trip,
                        ]
                    except Exception as _:
                        data2[tech_id]["last_shift"] = [trip]

            # оценить каждый рейс и записать в новые поля
            data3 = {}
            for tech_id, trips in data2.items():
                data3[tech_id] = {}
                data3[tech_id]["ratings"] = {}
                for type_par in ["last_trip", "last_hour", "last_shift"]:
                    try:
                        count = len(data2[tech_id][type_par])
                        if count > 0:
                            res = 0
                            elapsed_time = 0
                            for trip in data2[tech_id][type_par]:
                                res += kpd_match(trip)[0]
                                elapsed_time = ((trip[6] - trip[5]).total_seconds() / 60) + kpd_match(trip)[2]
                            if type_par == "last_trip":
                                data3[tech_id]["ratings"][type_par] = {
                                    "rating": int(res / count),
                                    "count": int(elapsed_time),
                                }
                            else:
                                data3[tech_id]["ratings"][type_par] = {
                                    "rating": int(res / count),
                                    "count": count,
                                }
                        else:
                            data3[tech_id]["ratings"][type_par] = {
                                "rating": 0,
                                "count": 0,
                            }
                    except Exception as _:
                        data3[tech_id]["ratings"][type_par] = {
                            "rating": 0,
                            "count": 0,
                        }

            # конвертация словаря в массив
            data4 = []
            for key, value in data3.items():
                data4.append({"tech_id": key, **value})

            return {"data": data4, "trips": trips_raw, "parameters": __parameters}

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}

    @staticmethod
    @router.get("/api/pto/reports/asd_errors", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def pto_reports_asd_errors(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["param_date"], "%Y-%m-%d")
        param_shift: int = int(request.query_params["param_shift"])
        param_target: str = str(request.query_params["param_target"])
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
            "param_target": param_target,
        }

        def __query() -> dict:
            errors_raw: list[tuple] = utils.request_to_oracle(
                query=queries.Pto.get_pto_report_asd_errors(),
                args=__parameters,
                many=True,
            )
            # print("errors_raw: ", errors_raw)
            errors_dict = [
                {
                    "title": i[0],
                    "description": i[1],
                }
                for i in errors_raw
            ]
            return {"data": errors_dict, "parameters": __parameters}

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}

    @staticmethod
    @router.get("/api/pto/monitoring/oper_stoppages", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def pto_monitoring_oper_stoppages(request: Request):
        def __query() -> dict:
            trips_raw: list[tuple] = utils.request_to_oracle(
                query=queries.Pto.get_pto_monitoring_oper_stoppages(),
                args={},
                many=True,
            )
            trips_dict = [
                {
                    "veh_id": i[1],
                    "timestop": "".join(str(i[2]).split("T")),
                    "timego": "".join(str(i[3]).split("T")),
                    "continious": i[4],
                    "type": "длящийся" if str(i[5]) != "68" else "ожидание под погрузку",
                    "description": i[6],
                    "planned": i[7],
                }
                for i in trips_raw
            ]
            return {"data": trips_dict, "parameters": {}}

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=4)}

    @staticmethod
    @router.get("/api/pto/reports/time_wait_to_load", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def pto_reports_time_wait_to_load(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["param_date"], "%Y-%m-%d")
        param_shift: int = int(request.query_params["param_shift"])
        param_target: int = int(request.query_params["param_target"])
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
            "param_target": param_target,
        }

        def __query() -> list | tuple | dict | str:
            def __query1() -> list | tuple | dict | str | None:
                trips_raw: list[tuple] = utils.request_to_oracle(
                    query=queries.Pto.get_pto_report_time_wait_to_load(),
                    args=__parameters,
                    many=True,
                )
                if len(trips_raw) < 1:
                    return None
                trips_dict = [
                    {
                        "veh_id": j[0],
                        "timeFrom": j[2],
                        "timeTo": j[3],
                        "time": j[4],
                    }
                    for j in trips_raw
                ]
                return trips_dict

            def __query2() -> list | tuple | dict | str | None:
                idle_raw: tuple = utils.request_to_oracle(
                    query=queries.Pto.get_pto_report_time_wait_to_load_avg(),
                    args=__parameters,
                    many=False,
                )

                if len(idle_raw) < 1:
                    return None
                idle_dict = {
                    "sum_idles": idle_raw[0],
                    "trips_idels": idle_raw[1],
                    "trips_all": idle_raw[2],
                    "avg_wait_all": idle_raw[3],
                    "avg_wait": idle_raw[4],
                }
                return idle_dict

            # try:
            #     workbook: Workbook = openpyxl.load_workbook("static_external/grafick.xlsx")
            #     worksheet: Worksheet = workbook.active
            #     matrix = [x for x in worksheet.iter_rows(values_only=True, min_row=2)]
            #     name = "Не назначено"
            #     for i in matrix[::-1]:
            #         if int(i[1]) == int(param_shift) and (i[0] <= param_date):
            #             name = f'{i[2]} [от {i[0].strftime("%d.%m.%Y")} | {i[1]}]'
            #             break
            # except Exception as error:
            #     name = str(error)

            try:
                data = __query1()
            except Exception as error:
                # print("data: ", error)
                data = None
            try:
                idle = __query2()
            except Exception as error:
                # print("idle: ", error)
                idle = None

            return {
                "data": data,
                "idle": idle,
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}

    @staticmethod
    @router.get("/api/pto/reports/sticking", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def pto_reports_sticking(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["param_date"], "%Y-%m-%d")
        param_shift: int = int(request.query_params["param_shift"])
        param_min_speed: int = int(request.query_params["param_min_speed"])
        param_min_weight: int = int(request.query_params["param_min_weight"])
        param_max_weight: int = int(request.query_params["param_max_weight"])

        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
            "param_min_speed": param_min_speed,
            "param_min_weight": param_min_weight,
            "param_max_weight": param_max_weight,
        }

        def __query() -> dict:
            rows_raw: list[tuple] = utils.request_to_oracle(
                query=queries.Pto.get_pto_report_sticking(),
                args=__parameters,
                many=True,
            )
            rows_dict = [
                {
                    "veh_id": i[2],
                    "sum_trips": i[3],
                    "avg_weight": i[4],
                    "sum_weight": i[5],
                    "sum_all_weight": i[6],
                    "percent_weight": 0 if i[6] == 0 else round(i[5] / i[6] * 100, 1),
                }
                for i in rows_raw
            ]
            return {"data": rows_dict, "parameters": __parameters}

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}


class Stoppages:
    @staticmethod
    @router.get("/api/stoppages/report/aux_dvs", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_stoppages_report_aux_dvs(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDate"], "%Y-%m-%d")
        param_shift: int = int(request.query_params["paramShift"])
        param_target: int = int(request.query_params["paramTarget"])
        param_select_tech_id: str = str(request.query_params["paramSelectTechId"])
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
            "param_target": param_target,
            "param_select_tech_id": param_select_tech_id,
        }

        def __query_by_aux_id(aux_id: str) -> list | dict | str:
            _start_at = time.perf_counter()
            __rows_raw: tuple = utils.request_to_oracle(
                query=queries.Stoppages.get_stoppages_report_aux_dvs(),
                args={
                    "param_date": param_date,
                    "param_shift": param_shift,
                    "param_select_tech_id": aux_id,
                },
                many=True,
            )

            # формирование списка словарей
            events: list[dict[str, any]] = []
            for event_raw in __rows_raw:
                events.append(
                    {
                        "tech": str(event_raw[3]),
                        "date_time": event_raw[0],
                        "speed": float(event_raw[1]),
                        "fuel": int(event_raw[2]),
                    }
                )

            # очистка простоев определённой длины
            stoppages: list[dict] = []
            while True:
                try:
                    last_index = 0
                    local_interval = []
                    for event in events:
                        last_index += 1
                        if event["speed"] == 0:
                            if len(local_interval) != 0:
                                last: datetime.datetime = local_interval[-1]["date_time"]
                                cur: datetime.datetime = event["date_time"]
                                if (last - cur).total_seconds() > 5 * 60:
                                    break
                            local_interval.append(event)
                        else:
                            if len(local_interval) == 0:
                                pass
                            else:
                                break
                    high = local_interval[0]["date_time"]
                    lower = local_interval[-1]["date_time"]
                    diff = round((high - lower).total_seconds() / 60, 1)
                    events = events[last_index:]
                    if diff > param_target:
                        stoppages.append(
                            {
                                "tech": local_interval[0]["tech"],
                                "from": lower,
                                "to": high,
                                "diff": diff,
                            }
                        )
                except Exception as _:
                    break

            return stoppages[::-1]

        def __query() -> dict:
            if param_select_tech_id.lower() == "все":
                auxes: list[str] = utils.get_all_auxes(["27", "219", "777", "2222", "3333"])
                all_stoppages: list[dict] = []
                for aux_id in auxes:
                    all_stoppages.extend(__query_by_aux_id(aux_id=aux_id))
            else:
                all_stoppages: list[dict] = __query_by_aux_id(aux_id=param_select_tech_id)
            return {
                "data": all_stoppages,
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=10)}

    @staticmethod
    @router.get("/api/stoppages/report/veh_dvs", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_stoppages_report_veh_dvs(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDate"], "%Y-%m-%d")
        param_shift: int = int(request.query_params["paramShift"])
        param_target: int = int(request.query_params["paramTarget"])
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
            "param_target": param_target,
        }

        def __query() -> dict:
            # беру всех водителей самосвалов за смену
            __rows_raw: list[tuple] = utils.request_to_oracle(
                query=queries.Stoppages.get_stoppages_report_veh_operators(),
                args={
                    "param_date": param_date,
                    "param_shift": param_shift,
                },
                many=True,
            )
            # print(__rows_raw[-1])

            # превращаю массив точек в массив словарей
            __rows_dict: list[dict] = [
                {
                    "veh_id": i[0],
                    "veh_id_fio": i[1],
                }
                for i in __rows_raw
            ]

            # создаю словарь с ключом в виде номера самосвала и значение в виде фио оператора
            operators = {}
            for i in __rows_dict:
                operators[i["veh_id"]] = i["veh_id_fio"]

            # беру все "точки" всех самосвалов за смену
            __rows_raw: list[tuple] = utils.request_to_oracle(
                query=queries.Stoppages.get_stoppages_report_veh_dvs_new(),
                args={
                    "param_date": param_date,
                    "param_shift": param_shift,
                },
                many=True,
            )
            # print(__rows_raw[-1])

            # превращаю массив точек в массив словарей
            __rows_dict: list[dict] = [
                {
                    "veh_id": i[0],
                    "date_time": i[1],
                    "speed": i[2],
                    "motohours": i[3],
                }
                for i in __rows_raw
            ]
            # print(__rows_dict[-1])

            # группирую словари по автосамосвалам
            __rows_dict_by_veh_id: dict[str, list] = {}
            for i in __rows_dict:
                try:
                    __rows_dict_by_veh_id[i["veh_id"]].append(i)
                except Exception as _:
                    __rows_dict_by_veh_id[i["veh_id"]] = [i]
            # print(__rows_dict_by_veh_id)

            __all_stoppages: list[dict] = []
            # запуск цикла по каждому самосвалу
            for k, events_by_veh_id in __rows_dict_by_veh_id.items():
                # очистка простоев определённой длины
                while True:
                    try:
                        last_index = 0
                        local_interval = []
                        for event in events_by_veh_id:
                            last_index += 1
                            if event["speed"] == 0:
                                if len(local_interval) != 0:
                                    last: datetime.datetime = local_interval[-1]["date_time"]
                                    cur: datetime.datetime = event["date_time"]
                                    if (last - cur).total_seconds() > 5 * 60:
                                        break
                                local_interval.append(event)
                            else:
                                if len(local_interval) == 0:
                                    pass
                                else:
                                    break
                        last_point: dict = local_interval[0]
                        first_point: dict = local_interval[-1]
                        difference_time = round((last_point["date_time"] - first_point["date_time"]).total_seconds() / 60, 2)
                        difference_motohours = round(last_point["motohours"] - first_point["motohours"], 2)
                        events_by_veh_id = events_by_veh_id[last_index:]
                        if difference_motohours > param_target:
                            if difference_motohours > difference_time:
                                difference_motohours = difference_time
                            __all_stoppages.append(
                                {
                                    "veh_id": k,
                                    "veh_id_fio": operators[k],
                                    "from": first_point["date_time"],
                                    "to": last_point["date_time"],
                                    "duration": round(difference_motohours / 60, 2),
                                }
                            )
                    except Exception as _:
                        break

            # сортировка простоев по времени начала
            __all_stoppages = sorted(__all_stoppages, key=lambda x: x["from"], reverse=False)
            summary_duration = 0.0
            for i in __all_stoppages:
                summary_duration += float(i["duration"])

            to_excel: list[list] = []
            for i in __all_stoppages:
                to_excel.append(
                    [
                        param_date,
                        param_shift,
                        i["veh_id"],
                        i["veh_id_fio"],
                        i["from"],
                        i["to"],
                        i["duration"],
                    ]
                )
            titles: list[str] = [
                "Дата",
                "Смена",
                "Самосвал",
                "Оператор",
                "Начало простоя",
                "Окончание простоя",
                "Длительность простоя ДВС, часы",
            ]
            path_to_excel_file: str = utils.export_to_excel(
                _data=to_excel,
                _titles=titles,
                _folder_name="stoppages",
                _prefix=f"{param_date.strftime('%Y-%m-%d')}_смена_{param_shift}___холостые_простои_двс_new",
            )

            return {
                "data": __all_stoppages,
                "extra": {
                    "path_to_excel_file": path_to_excel_file,
                    "summary_duration": round(summary_duration, 2),
                },
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=10)}

    @staticmethod
    @router.get("/api/stoppages/report/empty_peregon", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_empty_peregon_report_dumptrucks(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDate"], "%Y-%m-%d")
        param_shift: int = int(request.query_params["paramShift"])
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
        }

        def __query() -> dict:
            # все точки списком
            _events_list = utils.request_to_oracle(
                query=queries.Stoppages.get_empty_peregon_report_dumptrucks_new(),
                args={
                    "param_date": param_date,
                    "param_shift": param_shift,
                },
                many=True,
            )
            # все точки за смену словарями
            _events_list_dict = [
                {
                    "veh_id": i[0],
                    "date_time": i[1],
                    "x": int(i[2]),
                }
                for i in _events_list
            ]
            # все точки сгруппированные по хоз.номеру
            _events_dict_by_veh_id: dict[str, list[dict]] = {}
            for i in _events_list_dict:
                try:
                    _events_dict_by_veh_id[i["veh_id"]].append(i)
                except Exception as _:
                    _events_dict_by_veh_id[i["veh_id"]] = [i]

            # все перегоны
            _result_peregons: list[dict] = []
            line = 86250
            for _veh_id, _events in _events_dict_by_veh_id.items():
                vector = not (_events[0]["x"] > line)
                prev: datetime.datetime = None
                for i in _events:
                    curr = i["date_time"]
                    diff = 1 * 60 * 60 if prev is None else (curr - prev).total_seconds()
                    is_abk = i["x"] > line
                    if diff > 15 * 60:
                        if is_abk:
                            if vector:
                                vector = False
                                _result_peregons.append({"veh_id": f"{_veh_id}", "date_time": curr, "target": "На АБК"})
                                prev = curr
                        else:
                            if vector is False:
                                vector = True
                                _result_peregons.append({"veh_id": f"{_veh_id}", "date_time": curr, "target": "На Карьер"})
                                prev = curr
            _result_peregons = sorted(_result_peregons, key=lambda x: (x["veh_id"], x["date_time"]), reverse=False)
            to_excel: list[list] = []
            for i in _result_peregons:
                to_excel.append(
                    [
                        param_date,
                        param_shift,
                        i["veh_id"],
                        i["date_time"],
                        i["target"],
                    ]
                )
            titles: list[str] = [
                "Дата",
                "Смена",
                "Самосвал",
                "Время пересечения",
                "Направление",
            ]
            path_to_excel_file: str = utils.export_to_excel(
                _data=to_excel,
                _titles=titles,
                _folder_name="peregons",
                _prefix=f"{param_date.strftime('%Y-%m-%d')}_смена_{param_shift}___холостые_перегоны_new",
            )
            return {
                "data": _result_peregons,
                "extra": {
                    "path_to_excel_file": path_to_excel_file,
                },
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=10)}


class Gto:
    @staticmethod
    @router.get("/api/gto/report/dumptrucks", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_gto_report_dumptrucks(request: Request):
        param_date: datetime.datetime = datetime.datetime.strptime(request.query_params["param_date"], "%Y-%m-%d")
        param_shift: int = int(request.query_params["param_shift"])
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
        }

        def __query() -> dict:
            def query_speed_lenght_height_trips() -> list[dict]:
                workbook: Workbook = openpyxl.load_workbook("static_external/gto_norms_dumptrucks.xlsx")
                worksheet: Worksheet = workbook.active
                norms_dumptrucks: dict[str, dict] = {}
                for i in range(1 + 1, worksheet.max_row + 1):
                    norms_dumptrucks[str(worksheet.cell(row=i, column=1).value)] = {
                        "plan_speed_load": float(worksheet.cell(row=i, column=2).value),
                        "plan_speed_unload": float(worksheet.cell(row=i, column=3).value),
                        "plan_speed_all": float(worksheet.cell(row=i, column=4).value),
                        "plan_weight": float(worksheet.cell(row=i, column=5).value),
                        "plan_min_weight": float(worksheet.cell(row=i, column=6).value),
                        "plan_max_weight": float(worksheet.cell(row=i, column=7).value),
                        "plan_length_all": float(worksheet.cell(row=i, column=8).value),
                        "plan_time_unloading": float(worksheet.cell(row=i, column=9).value),
                        "plan_time_move_to_lenght": float(worksheet.cell(row=i, column=10).value),
                        "plan_avg_volume": float(worksheet.cell(row=i, column=11).value),
                    }

                workbook: Workbook = openpyxl.load_workbook("static_external/gto_norms_shovels.xlsx")
                worksheet: Worksheet = workbook.active
                norms_shovels: dict[str, dict] = {}
                for i in range(1 + 1, worksheet.max_row + 1):
                    norms_shovels[str(worksheet.cell(row=i, column=1).value)] = {
                        "plan_time_loading": float(worksheet.cell(row=i, column=2).value),
                        "plan_length_all": float(worksheet.cell(row=i, column=3).value),
                    }

                # todo комментарии по скорости самосвалов ###########################################
                comments_raw = utils.db_query_sqlite(
                    query=queries.Speed.get_speed_monitoring_all_dumptrucks_comments(),
                    args={
                        "taskdate": param_date,
                        "shift": param_shift,
                    },
                    many=True,
                )
                comments_dict: dict[str, str] = {}
                for i in comments_raw:
                    comments_dict[str(i[0])] = str(i[1])

                # todo фамилии операторов и максимум рейсов под кем ################################
                trips_raw: list[tuple] = utils.request_to_oracle(
                    query=queries.Gto.get_gto_trips_with_drivers(),
                    args={"param_date": param_date, "param_shift": param_shift},
                    many=True,
                )
                trips_dict: list[dict] = [
                    {
                        "veh_id": i[0],
                        "veh_id_fio": i[1],
                        "shov_id": i[2],
                        "shov_id_fio": i[3],
                    }
                    for i in trips_raw
                ]

                # сгруппировать по каждому самосвалу
                trips_by_dumptrucks: dict[str, list[dict]] = {}
                for trip in trips_dict:
                    veh_id = str(trip["veh_id"])
                    try:
                        trips_by_dumptrucks[veh_id].append(trip)
                    except Exception as _:
                        trips_by_dumptrucks[veh_id] = [trip]
                drivers = {}
                for veh_id, trips in trips_by_dumptrucks.items():
                    shovels = {}
                    for trip in trips:
                        try:
                            shovels[trip["shov_id"]] = shovels[trip["shov_id"]] + 1
                        except:
                            shovels[trip["shov_id"]] = 1
                    shov_id = max(shovels, key=shovels.get)
                    shov_id_fio = ""
                    veh_id_fio = ""
                    for trip in trips:
                        if trip["shov_id"] == shov_id:
                            shov_id_fio = trip["shov_id_fio"]
                            veh_id_fio = trip["veh_id_fio"]
                    drivers[veh_id] = {
                        "veh_id": veh_id,
                        "veh_id_fio": veh_id_fio,
                        "shov_id": shov_id,
                        "shov_id_fio": shov_id_fio,
                    }
                #####################################################################################

                trips_raw: list[tuple] = utils.request_to_oracle(
                    query=queries.Gto.get_gto_report_dumptrucks_speed_lenght_height_trips(),
                    args={"param_date": param_date, "param_shift": param_shift},
                    many=True,
                )

                def get_percent(cur_val, all_val, digit=0):
                    return round((cur_val / all_val * 100) - 100, digit)

                trips_dict: list[dict] = []
                for i in trips_raw:
                    veh_id = i[0]
                    shov_id = drivers[veh_id]["shov_id"]

                    _avg_speed_all = i[5]
                    _plan_speed_all = norms_dumptrucks[veh_id]["plan_speed_all"]
                    _diff_speed_all = get_percent(_avg_speed_all, _plan_speed_all)  # round((_avg_speed_all / _plan_speed_all * 100) - 100, 0)
                    _speed_comment = comments_dict.get(str(veh_id), "-")

                    _avg_length_all = i[8]
                    _plan_length_all = norms_shovels[shov_id]["plan_length_all"]
                    _diff_length_all = get_percent(_avg_length_all, _plan_length_all) * -1

                    _avg_weight = i[10]
                    _plan_weight = norms_dumptrucks[veh_id]["plan_weight"]
                    _diff_avg_weight = get_percent(_avg_weight, _plan_weight)

                    _avg_volume = i[13]
                    _plan_avg_volume = norms_dumptrucks[veh_id]["plan_avg_volume"]
                    _diff_avg_volume = get_percent(_avg_volume, _plan_avg_volume)

                    _avg_time_loading = i[15]
                    _plan_time_loading = norms_shovels[shov_id]["plan_time_loading"]
                    _diff_avg_time_loading = get_percent(_avg_time_loading, _plan_time_loading) * -1

                    _avg_time_unloading = i[17]
                    _plan_time_unloading = norms_dumptrucks[veh_id]["plan_time_unloading"]
                    _diff_avg_time_unloading = get_percent(_avg_time_unloading, _plan_time_unloading) * -1

                    _avg_length_load = i[6]  # 2.2 ср. расстояние
                    _avg_time_move = i[19]  # 9минут
                    _plan_time_move_to_lenght = norms_dumptrucks[veh_id]["plan_time_move_to_lenght"]  # 4минуты / 1км
                    _diff_avg_time_move_to_lenght = get_percent(_avg_time_move / _avg_length_load, _plan_time_move_to_lenght) * -1

                    new_dict = {
                        "veh_id": veh_id,
                        "veh_id_fio": drivers[veh_id]["veh_id_fio"],
                        "shov_id": shov_id,
                        "shov_id_fio": drivers[veh_id]["shov_id_fio"],
                        "sum_trips": i[2],
                        #
                        "avg_speed_load": i[3],
                        "avg_speed_unload": i[4],
                        "avg_speed_all": _avg_speed_all,
                        "plan_speed_all": _plan_speed_all,
                        "diff_speed_all": _diff_speed_all,
                        "speed_comment": _speed_comment,
                        #
                        "avg_length_load": _avg_length_load,
                        "avg_length_unload": i[7],
                        "avg_length_all": _avg_length_all,
                        "plan_length_all": _plan_length_all,
                        "diff_length_all": _diff_length_all,
                        #
                        "diff_height": i[9],
                        #
                        "sum_weight": i[11],
                        "avg_weight": _avg_weight,
                        "plan_weight": _plan_weight,
                        "diff_avg_weight": _diff_avg_weight,
                        #
                        "sum_volume": i[12],
                        "avg_volume": _avg_volume,
                        "plan_avg_volume": _plan_avg_volume,
                        "diff_avg_volume": _diff_avg_volume,
                        #
                        "sum_time_loading": i[14],
                        "avg_time_loading": _avg_time_loading,
                        "plan_time_loading": _plan_time_loading,
                        "diff_avg_time_loading": _diff_avg_time_loading,
                        #
                        "sum_time_unloading": i[16],
                        "avg_time_unloading": _avg_time_unloading,
                        "plan_time_unloading": _plan_time_unloading,
                        "diff_avg_time_unloading": _diff_avg_time_unloading,
                        #
                        "sum_time_move": i[18],
                        "avg_time_move": _avg_time_move,
                        "plan_time_move_to_lenght": _plan_time_move_to_lenght,
                        "diff_avg_time_move_to_lenght": _diff_avg_time_move_to_lenght,
                        #
                        "sum_time_trip": i[20],
                        "avg_time_trip": i[21],
                    }
                    trips_dict.append(new_dict)

                trips = sorted(trips_dict, key=lambda x: (x["shov_id"], x["veh_id"]), reverse=False)

                return trips

            return {
                "data": query_speed_lenght_height_trips(),
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}


class Develop:
    @staticmethod
    @router.get("/api/target/monitoring/weight_loads", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_target_monitoring_weight_loads(request: Request):
        now = datetime.datetime.now() - datetime.timedelta(days=1)
        param_shift: int = utils.get_selected_shift(date_time=now)
        param_date: datetime.datetime = utils.get_selected_taskdate(date_time=now)
        param_date_shift_begin: datetime.datetime = utils.get_shift_datetime_begin(date_time=now)

        __parameters = {
            "param_shift": param_shift,
            "param_date": param_date,
        }

        def __query() -> list | tuple | dict | str:
            # todo ###################################################################################################################################
            workbook: Workbook = openpyxl.load_workbook("static_external/norms_dumptrucks.xlsx")
            worksheet: Worksheet = workbook.active
            norms_dumptrucks: dict[str, dict] = {}
            for j in range(1 + 1, worksheet.max_row + 1):
                norms_dumptrucks[str(worksheet.cell(row=j, column=1).value)] = {
                    "min_weight": worksheet.cell(row=j, column=2).value,
                    "max_weight": worksheet.cell(row=j, column=3).value,
                }

            trips_raw: list[tuple] = utils.request_to_oracle(
                query=queries.Develop.get_target_monitoring_loads_new(),
                args={},
                many=True,
            )
            trips_dict: list[dict] = [
                {
                    "shov_id": i[0],
                    "veh_id": i[1],
                    "time_load": i[2],
                    "time_unload": i[3],
                    "weight": i[4],
                    "length_load": i[5],
                    "length_unload": i[6],
                    "height": i[7],
                    "worktype": i[8],
                    "volume": i[11],
                    "fio": i[-1],
                }
                for i in trips_raw
            ]

            # сгруппировать по каждому экскаватору
            trips_by_shovels: dict[str, list[dict]] = {}
            for trip in trips_dict:
                shov_id = str(trip["shov_id"])
                try:
                    trips_by_shovels[shov_id].append(trip)
                except Exception as _:
                    trips_by_shovels[shov_id] = [trip]

            def kpd(_trips: list[dict]) -> dict:
                trips_count = len(_trips)
                _shov_id = _trips[0]["shov_id"]
                _fio = _trips[0]["fio"]

                lower_trips = []
                higher_trips = []
                normal_trips = []

                volume_by_worktype = {}

                for i in _trips:
                    weight = int(i["weight"])
                    veh_id = str(i["veh_id"])
                    volume = float(i["volume"])

                    volume_by_worktype[i["worktype"]] = volume_by_worktype.get(i["worktype"], 0) + volume

                    if weight < norms_dumptrucks[veh_id]["min_weight"]:
                        lower_trips.append(i)
                    elif weight > norms_dumptrucks[veh_id]["max_weight"]:
                        higher_trips.append(i)
                    else:
                        normal_trips.append(i)

                volumes_by_worktype: list[dict] = []
                for k, v in volume_by_worktype.items():
                    volumes_by_worktype.append({"name": k, "value": round(v, 2)})

                lower_count = len(lower_trips)
                lower = {"trips_count": lower_count, "percent": int(lower_count / trips_count * 100), "negative_weight": 0}
                for i in lower_trips:
                    lower["negative_weight"] += 91 - int(i["weight"])
                lower["negative_trips"] = lower["negative_weight"] // 91 if lower_count > 0 else 0

                higher_count = len(higher_trips)
                higher = {"trips_count": higher_count, "percent": int(higher_count / trips_count * 100), "negative_weight": 0}
                for i in higher_trips:
                    higher["negative_weight"] += int(i["weight"]) - 91
                higher["negative_trips"] = higher["negative_weight"] // 91 if higher_count > 0 else 0

                normal_count = len(normal_trips)
                normal = {"trips_count": normal_count, "percent": int(normal_count / trips_count * 100), "negative_weight": 0}
                for i in normal_trips:
                    normal["negative_weight"] += int(i["weight"]) - 91
                normal["negative_trips"] = normal["negative_weight"] // 91 if normal_count > 0 else 0

                hour_from_start_shift = round((now - param_date_shift_begin).seconds / 60 / 60, 1)

                return {
                    "shov_id": _shov_id,
                    "fio": _fio,
                    "lower": lower,
                    "higher": higher,
                    "normal": normal,
                    "hour_from_start_shift": hour_from_start_shift,
                    "volumes_by_worktype": volumes_by_worktype,
                }

            # оценка показателей каждого экскаватора
            shovels = [kpd(v) for k, v in trips_by_shovels.items()]

            # очистка только перегрузов и недогрузов
            trips_warning_dict: list[dict] = [
                {
                    "shov_id": i[0],
                    "veh_id": i[1],
                    "time_load": i[2],
                    "weight": i[4],
                    "fio": i[-1],
                }
                for i in trips_raw
                if (int(i[4]) > norms_dumptrucks[str(i[1])]["max_weight"]) or (int(i[4]) < norms_dumptrucks[str(i[1])]["min_weight"])
            ]

            # сгруппировать по каждому экскаватору
            trips_by_shovels_warning = {}
            for trip in trips_warning_dict:
                shov_id = str(trip["shov_id"])
                try:
                    trips_by_shovels_warning[shov_id].append(trip)
                except Exception as _:
                    trips_by_shovels_warning[shov_id] = [trip]

            # оценка показателей каждого экскаватора
            list_of_shovels = [
                {"shov_id": k, "fio": v[0]["fio"], "trips": sorted(v, key=lambda x: x["time_load"], reverse=False)}
                for k, v in trips_by_shovels_warning.items()
            ]
            return {
                "data": sorted(list_of_shovels, key=lambda x: x["shov_id"], reverse=False),
                "shovels": sorted(shovels, key=lambda x: x["shov_id"], reverse=False),
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=5)}

    @staticmethod
    @router.get("/api/target/report/weight_loads", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_target_report_weight_loads(request: Request):
        param_date_from: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDateFrom"], "%Y-%m-%d")
        param_date_to: datetime.datetime = datetime.datetime.strptime(request.query_params["paramDateTo"], "%Y-%m-%d")
        param_shift_from: int = int(request.query_params["paramShiftFrom"])
        param_shift_to: int = int(request.query_params["paramShiftTo"])
        __parameters = {
            "param_date_from": param_date_from,
            "param_date_to": param_date_to,
            "param_shift_from": param_shift_from,
            "param_shift_to": param_shift_to,
            "param_weight_low": 91,
            "param_weight_high": 100,
        }

        def __query() -> dict:
            _trips_raw: list[tuple] = utils.request_to_oracle(
                args=__parameters,
                many=True,
                query=queries.Develop.get_target_report_weight_loads(),
            )
            _trips_dict: list[dict] = [
                {
                    "date": i[0],
                    "shift": i[1],
                    "shov_id": i[2],
                    "shov_driver": i[3],
                    "veh_id": i[4],
                    "veh_driver": i[5],
                    "area": i[6],
                    "worktype": i[7],
                    "unloadid": i[8],
                    "timeload": i[9],
                    "avspeed": i[10],
                    "weigth": i[11],
                }
                for i in _trips_raw
            ]

            def export_to_excel(_data: list[dict]) -> str:
                key = utils.create_encrypted_password(
                    _random_chars="abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890",
                    _length=8,
                )
                date = f"{time.strftime('%Y-%m-%d')}"
                path = "media/data/temp/target_report_weight_loads"
                file_name = f"недогрузы_и_перегрузы_{date}_{key}.xlsx"
                workbook: Workbook = openpyxl.Workbook()
                worksheet: Worksheet = workbook.active

                # Delete old files
                for root, dirs, files in os.walk(path, topdown=True):
                    for file in files:
                        try:
                            date_file = str(file).strip().split(".")[0].strip().split("_")[-1]
                            if date != date_file:
                                os.remove(f"{path}/{file}")
                        except Exception as _:
                            pass
                # worksheet.cell(row=1, column=1, value="РАБОЧИЙ ЛИСТ")
                # worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

                titles = [
                    "Дата",
                    "Смена",
                    "Экскаватор",
                    "Машинист",
                    "Самосвал	",
                    "Оператор",
                    "Горизонт",
                    "Материал",
                    "Место разгрузки",
                    "Время погрузки",
                    "Ср. скорость",
                    "Масса",
                ]
                for col_idx, value in enumerate(titles, 1):
                    worksheet.cell(row=1, column=col_idx, value=value)

                for row_idx, row in enumerate(_data, 2):
                    for col_idx, value in enumerate(row.values(), 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)

                worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{len(_data) + 1}"

                side_think = Side(border_style="thin", color="FF808080")
                # {'mediumDashDotDot', 'thin', 'dashed', 'mediumDashed', 'dotted', 'double', 'thick',
                # 'medium', 'dashDot','dashDotDot', 'hair', 'mediumDashDot', 'slantDashDot'}
                border_think = Border(
                    top=side_think,
                    left=side_think,
                    right=side_think,
                    bottom=side_think,
                )
                aligm_think = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                    shrink_to_fit=True,
                )
                font_header = Font(name="Arial", size=12, bold=True)

                # headers
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = font_header
                    cell.border = border_think

                font_body = Font(name="Arial", size=10, bold=False)
                green_fill = PatternFill(fgColor="F04141", fill_type="solid")

                # body
                for row_idx in range(2, worksheet.max_row + 1):
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = font_body
                        cell.alignment = aligm_think
                        cell.border = border_think
                        if col_idx == 12 and cell.value >= 100:
                            cell.fill = green_fill

                # set column width
                for col_idx in range(1, worksheet.max_column + 1):
                    width: int = 1
                    for row_idx in range(1, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        value: int = len(str(cell.value))
                        if value > width:
                            width = value
                    worksheet.column_dimensions[get_column_letter(col_idx)].height = 1
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = round((width * 1.25), 3)
                    # worksheet.column_dimensions[get_column_letter(col_idx)].auto_size = True
                    # worksheet.column_dimensions[get_column_letter(col_idx)].bestFit = True

                worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
                worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LETTER
                worksheet.page_margins.left = 0.05
                worksheet.page_margins.right = 0.05
                worksheet.page_margins.header = 0.1
                worksheet.page_margins.bottom = 0.2
                worksheet.page_margins.footer = 0.2
                worksheet.page_margins.top = 0.1
                worksheet.print_options.horizontalCentered = True
                # sheet.print_options.verticalCentered = True
                worksheet.page_setup.fitToHeight = 1
                worksheet.page_setup.scale = 100
                worksheet.page_setup.fitToHeight = 1
                worksheet.page_setup.fitToWidth = 1
                # worksheet.protection.password = key + "_1"
                # worksheet.protection.sheet = True
                # worksheet.protection.enable()
                full_path = f"{path}/{file_name}"
                workbook.save(full_path)
                return full_path

            return {
                "data": _trips_dict,
                "path_to_excel_file": export_to_excel(_trips_dict),
                "parameters": __parameters,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=5)}

    @staticmethod
    @router.get("/api/pto/analytic_tech", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_pto_analytic_tech(request: Request):
        def __query() -> dict:
            volumes_by_hours: list[any] = utils.request_to_oracle(
                args={},
                many=True,
                query=queries.Develop.get_pto_analytic_tech_volumes_by_hours(),
            )
            volumes_by_hours = [{"name": x[0], "объём": float(x[1]), "pv": 2400, "amt": 2400} for x in volumes_by_hours]

            volumes_by_category: list[any] = utils.request_to_oracle(
                args={},
                many=True,
                query=queries.Develop.get_pto_analytic_tech_volumes_by_category(),
            )
            volumes_by_category = [
                {
                    "name": x[0] + f" | {x[2]} рейса(-ов)",
                    "объём": float(x[1]),
                    "pv": 2400,
                    "amt": 2400,
                }
                for x in volumes_by_category
            ]
            elapsed_time_raw: tuple[any] = utils.request_to_oracle(args={}, many=False, query=queries.Develop.query_get_elapsed_time())

            # TODO #################################################################################################

            now = datetime.datetime.now() - datetime.timedelta(days=1)
            param_shift: int = utils.get_selected_shift(date_time=now)
            param_date: datetime.datetime = utils.get_selected_taskdate(date_time=now)
            param_date_shift_begin: datetime.datetime = utils.get_shift_datetime_begin(date_time=now)
            param_date_shift_end: datetime.datetime = utils.get_shift_datetime_end(date_time=now)

            list_of_shovels: list[dict] = []
            for shov_id in [
                "001",
                "002",
                "003",
                "201",
                "202",
                "203",
                "204",
                "205",
                "206",
                "207",
                "208",
                "255",
                # "330",
                "401",
                "402",
                "403",
            ]:
                """
                0. Все данные спрятаны циклом по accordion, но есть общий верхний групповой(после настроек)

                1. Перегрузы и недогрузы - график столбцами, сортированный по времени
                {
                    "fio_shov": "...",
                    "list": [...{"datetime": "datetime", "fio_veh", "veh": "145", "mass": 102, "volume": "30.0" "material_type": "вскрыша"},
                ]
                """

                # хоз номер экскаватора
                dict_of_shovel: dict = {"shov_id": shov_id}

                # фио машиниста
                fio_shov_r: tuple = utils.request_to_oracle(
                    args={
                        "param_date": param_date,
                        "param_shift": param_shift,
                        "param_shov_id": shov_id,
                    },
                    many=False,
                    query=queries.Develop.get_target_monitoring_weight_loads(),
                )
                fio_shov = "-" if fio_shov_r is None else fio_shov_r[0]
                if fio_shov == "-":
                    continue
                dict_of_shovel["fio_shov"] = fio_shov

                # рейсы с перегрузами и недогрузами
                list_peregruz_and_nedogruz_r: list[tuple] = utils.request_to_oracle(
                    args={
                        "param_shov_id": shov_id,
                        "param_date": param_date,
                        "param_shift": param_shift,
                        "param_shift_begin": param_date_shift_begin,
                        "param_shift_end": param_date_shift_end,
                        "param_low": 92,
                        "param_high": 99,
                    },
                    many=True,
                    query=queries.Develop.get_pto_analytic_tech_weight_loads(),
                )
                list_peregruz_and_nedogruz = [
                    {
                        "vehid": x[0],
                        "worktype": x[1],
                        "timeload": x[2],
                        "weight": x[3],
                        "volume": x[4],
                        "hour_load": x[5],
                        "bucket_count": x[6],
                        "fio": x[7],
                    }
                    for x in list_peregruz_and_nedogruz_r
                ]
                dict_of_shovel["list_peregruz_and_nedogruz"] = list_peregruz_and_nedogruz

                # volumes_by_hours: list[any] = Utils.request_to_oracle(
                #     args={},
                #     many=True,
                #     query=queries.query_get_volumes_by_category(),
                # )
                # volumes_by_category = [
                #     {
                #         "name": x[0] + f" | {x[2]} рейса(-ов)",
                #         "объём": float(x[1]),
                #         "pv": 2400,
                #         "amt": 2400,
                #     }
                #     for x in volumes_by_category
                # ]

                list_of_shovels.append(dict_of_shovel)
            # TODO #################################################################################################

            return {
                "data": None,
                "trips": None,
                "extra": {
                    "volumes_by_hours": volumes_by_hours,
                    "volumes_by_category": volumes_by_category,
                    "trips_count": -1,
                    "elapsed_time": elapsed_time_raw[2],
                },
                "list_of_shovels": list_of_shovels,
            }

        return {"response": constants.cache.get(query=lambda: __query, request=request, timeout=9)}


class Base:
    @staticmethod
    @router.get("/", response_class=HTMLResponse)
    @utils.decorator(need_auth=False)
    async def root(request: Request):
        try:
            headers = {"Cache-Control": "max-age=0"}
            return templates.TemplateResponse("backend/react/build/index.html", {"request": request}, headers=headers)
        except Exception as error:
            return templates.TemplateResponse("backend/templates/error.html", {"request": request, "error": str(error.__str__)})

    @staticmethod
    @router.get("/api", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def api(request: Request):
        debug.get_tech_message_delay()
        # import requests
        #
        # bot_token = "6619466319:AAFQ5XcFnDIR1PWCVrT4GONq6LiD_ceYkZk"
        # message_text = "Всем\nпривет"
        # r = requests.post(
        #     url=f"https://api.telegram.org/bot{bot_token}/sendMessage",
        #     data={"chat_id": "1289279426", "text": message_text},
        # )
        # print(r.status_code)

        #         while True:
        #             date_time = datetime.datetime.now().strftime("%d.%m.%Y %H")
        #             message_text = f"""Данные за {date_time}:00
        # Рейсы: 40 (ск:12 | рыхл:8 | руд:7)
        # Объём:  (ск:12 | рыхл:8 | руд:7)
        #
        # СКОРОСТЬ:
        # Ср.скорость: 18.6 (груж:16.6 | порожн:20.2)
        # Лучший(145): 20.6 (груж:16.6 | порожн:20.2)
        # Худший(132): 18.6 (груж:16.6 | порожн:20.2)
        # """
        #             r = requests.post(
        #                 url=f"https://api.telegram.org/bot{bot_token}/sendMessage",
        #                 data={"chat_id": "1289279426", "text": message_text},
        #             )
        #             await asyncio.sleep(60 * 60 * 0.5)

        # chats = requests.get(f"https://api.telegram.org/bot{bot_token}/getUpdates").json()
        # print("chats: ", chats)
        # for chat in chats["result"]:
        #     r = requests.post(
        #         url=f"https://api.telegram.org/bot{bot_token}/sendMessage",
        #         data={"chat_id": chat["message"]["chat"]["id"], "text": message_text},
        #     )
        #     print(r.status_code)

        return {"data": f"OK"}

    @staticmethod
    @router.get("/api/communicator", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def get_communicator(request: Request):
        async def __query() -> dict:
            async with aiosqlite.connect("database_centr.db") as connection:
                cursor = await connection.cursor()
                query1 = """
CREATE TABLE IF NOT EXISTS messages (
id INTEGER PRIMARY KEY AUTOINCREMENT,
subsystem TEXT UNIQUE,
message TEXT,
date_time_subsystem DATETIME,
date_time_server DATETIME
)
"""
                await cursor.execute(query1)
                await connection.commit()
                query2 = """
SELECT subsystem, message, date_time_subsystem, date_time_server  
FROM messages
"""
                async with connection.execute(query2) as cursor:
                    data = await cursor.fetchall()
                    json_data = []
                    for i in data:
                        json_data.append(
                            {
                                "subsystem": str(i[0]),
                                "message": json.loads(str(i[1])),
                                "date_time_subsystem": str(i[2]),
                                "date_time_server": str(i[3]),
                            }
                        )
                    group_by_subsystem = {}
                    for j in json_data:
                        group_by_subsystem[j["subsystem"]] = j
                    await asyncio.sleep(1)
                    print(f"group_by_subsystem\n: {group_by_subsystem}")
                    return {"data": group_by_subsystem}

        return {"response": await constants.cache.async_get(query=lambda: __query, request=request, timeout=5)}

    @staticmethod
    @router.post("/api/communicator", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def post_communicator(request: Request):
        if str(request.headers.get("Authorization", "")) != "Token=auth_token":
            raise Exception("Неверный токен доступа!")

        form_data = await request.json()
        print(f"form_data\n: {form_data}")

        subsystem = form_data["subsystem"]
        data = form_data["data"]
        date_time_subsystem = form_data["date_time_subsystem"]
        date_time_server = datetime.datetime.now()
        async with aiosqlite.connect("database_centr.db") as connection:
            cursor = await connection.cursor()
            query1 = """
CREATE TABLE IF NOT EXISTS messages (
id INTEGER PRIMARY KEY AUTOINCREMENT,
subsystem TEXT UNIQUE,
message TEXT,
date_time_subsystem DATETIME,
date_time_server DATETIME
)
"""
            await cursor.execute(query1)
            await connection.commit()
            query2 = """
INSERT OR REPLACE INTO messages 
(subsystem, message, date_time_subsystem, date_time_server)
VALUES 
(?, ?, ?, ?)
"""
            await cursor.execute(query2, (subsystem, json.dumps(data), date_time_subsystem, date_time_server))
            await connection.commit()

        return {"data": "OK"}

    @staticmethod
    @router.post("/redirect", response_class=RedirectResponse)
    @utils.decorator(need_auth=False)
    async def redirect(_: Request):
        return RedirectResponse(url=router.url_path_for("root"), status_code=303)

    @staticmethod
    @router.get("/{path:path}", response_class=RedirectResponse)
    @utils.decorator(need_auth=False)
    async def default(request: Request, path: str):
        print(f"Path '{path}' is empty!")
        return RedirectResponse(url=router.url_path_for("root"), status_code=303)


class Old:
    @staticmethod
    @router.get("/api/captcha", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def captcha(request: Request):
        await asyncio.sleep(1.0)

        if request.query_params.get("id", None) != "666":
            raise Exception("Captcha error")

        return {"response": "OK"}

    @staticmethod
    @router.post("/api/register", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def register(request: Request):
        # from await axios.post("http://127.0.0.1:8000/api/token", new FormData().append("username", "admin"));
        # form = await request.form()

        # from await axios.post("http://127.0.0.1:8000/api/token", {username: "admin"})
        form = await request.json()
        email: str = str(form.get("email"))
        password: str = str(form.get("password"))
        if not re.match(r"[A-Za-z0-9._-]+@[A-Za-z0-9.-]+\.[A-Z|data-z]{2,}", string=email):
            raise Exception("Email not valid!")
        if not re.match(
            r"^(?=.*?[data-z])(?=.*?[A-Z])(?=.*?[0-9])(?=.*?[#?!@$%^&*-]).{8,}$",
            string=password,
        ):
            raise Exception("Password too weak!")
        utils.db_query_sqlite(
            query="""INSERT INTO users (email, password) VALUES (?, ?)""",
            args=(email, hashlib.sha256(password.encode()).hexdigest()),
            many=False,
        )
        return "OK"

    @staticmethod
    @router.post("/api/login", response_class=JSONResponse)
    @utils.decorator(need_auth=False)
    async def login(request: Request, response: Response):
        # get data from form
        form = await request.json()
        email: str = str(form.get("email"))
        password: str = hashlib.sha256(str(form.get("password")).encode()).hexdigest()

        # authenticate user
        user_id: tuple = utils.db_query_sqlite(
            query="""SELECT id FROM users WHERE email = ? and password = ?""",
            args=(email, password),
            many=False,
        )
        if user_id is None:
            raise Exception("Login or password incorrect!")
        user_id: int = user_id[0]

        # create tokens
        token_access: str = hashlib.sha256(str(email + password + str(user_id)).encode()).hexdigest()
        utils.db_query_sqlite(
            query="""INSERT INTO tokens (user_id, token_access) VALUES (?, ?)""",
            args=(user_id, token_access),
            many=False,
        )

        # set cookies
        response.set_cookie(
            "token_access",
            token_access,
            max_age=constants.jwt_token_lifetime_seconds,
            httponly=True,
            secure=True,
        )

        return "OK"
