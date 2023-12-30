# TODO download modules ##############################################################################################################################
import datetime
import os
import random
import sqlite3
import threading
import time
from functools import wraps
from typing import Coroutine

import aiofiles
import openpyxl
import oracledb
from fastapi import Request
from openpyxl.styles import Side, Alignment, Font, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# TODO custom modules ################################################################################################################################
from . import constants

#


class CacheServer:
    def __init__(self):
        self.data: dict[str, any] = {}
        self.clear_tasks: dict[str, datetime.datetime] = {}

    def get(
        self,
        key: str = None,
        request: Request = None,
        query: callable = None,
        timeout: float = 1.0,
    ):
        if key is None:
            key = f"{request.url}_{request.method}_{''.join([str(x) for x in request.query_params.values()])}"
        value = self.data.get(key, None)
        if value is None and query:
            value = query()()
            self.set(key=key, value=value, timeout=timeout)

        return value

    async def async_get(
        self,
        key: str = None,
        request: Request = None,
        query: Coroutine = None,
        timeout: float = 1.0,
    ):
        if key is None:
            key = f"{request.url}_{request.method}_{''.join([str(x) for x in request.query_params.values()])}"
        value = self.data.get(key, None)
        if value is None and query:
            value = await query()()
            self.set(key=key, value=value, timeout=timeout)

        return value

    def set(self, key: str, value: any, timeout: float = 1.0):
        # Задача очистки кэша не останавливается при установке нового кэша

        self.data[key] = value
        threading.Thread(
            target=self.clear_cache,
            args=(),
            kwargs={
                "key": key,
                "timeout": timeout,
            },
        ).start()

    def clear(self, key: str):
        if self.data.get(key, None) is not None:
            del self.data[key]

    def reset(self):
        self.data = {}

    def clear_cache(self, key: str, timeout: float = 1.0):
        if 0.0 >= timeout > 99999:
            return

        # запоминаю время начала этой задачи
        date_time = datetime.datetime.now()
        self.clear_tasks[key] = date_time

        time.sleep(timeout)

        # отменяю задачу очистки кэша, если он был перезаписан
        if self.clear_tasks.get(key, None) != date_time:
            return

        if self.data.get(key, None) is not None:
            del self.data[key]


def decorator(need_auth=False):
    def decorator_inline(function: callable):
        @wraps(function)
        async def decorated_function(request: Request, *args, **kwargs):
            _error = None
            response = None
            # try:
            time_start_func = time.perf_counter()

            client_ip = request.client.host
            client_ip = "127.0.0.1"
            if client_ip != "127.0.0.1":
                try:
                    subnet = client_ip.split(".")[-2]
                    if subnet not in ["16", "17", "23", "200"]:
                        with open("clients.txt", "a") as file:
                            file.write(f"{client_ip}\n")
                        print("Subnet: ", subnet, " Нет доступа!")
                        raise Exception("Нет доступа!")
                except Exception as error:
                    raise Exception("Нет доступа!")

            response = await function(request, *args, **kwargs)
            elapsed_time = round((time.perf_counter() - time_start_func), 2)
            user_id = -1
            if need_auth:
                token_access: str | None = request.cookies.get("token_access", None)
                if token_access is None:
                    raise Exception("Need Authorization!")
                token: tuple = db_query_sqlite(
                    query="""SELECT user_id, datetime_elapsed FROM tokens WHERE token_access = ?""",
                    args=(token_access,),
                    many=False,
                )
                user_id: int = token[0]
                datetime_elapsed: datetime.datetime = datetime.datetime.strptime(token[1], "%Y-%m-%d %H:%M:%S")
                if datetime.datetime.now() > datetime_elapsed + datetime.timedelta(hours=6, seconds=constants.jwt_token_lifetime_seconds):
                    raise Exception("Token is terminate")
                request.user_id = user_id
            text = (
                f"{client_ip} | {str(datetime.datetime.now())[0:-5:1]}({elapsed_time}s) ({request.method})/{'/'.join(str(request.url).split('/')[3:])} || "
                f"{response if constants.LOGGING_RESPONSE else '[successfully response disabled]'}"
            )
            # except Exception as error:
            #     _error = str(error)
            #     user_id = -1
            #     text = f"{str(datetime.datetime.now())[0:-5:1]} ({request.method})/{'/'.join(str(request.url).split('/')[3:])} || {error}"

            if constants.LOGGING_TO_CONSOLE:
                print("\n" + text)
            if constants.LOGGING_TO_FILE:
                async with aiofiles.open("static/log.txt", mode="a", encoding="utf-8") as file:
                    await file.write(text + "\n")
            if constants.LOGGING_TO_DATABASE:
                db_query_sqlite(
                    query="""INSERT INTO logs (user_id, ip, path, method, data) VALUES (?, ?, ?, ?, ?)""",
                    args=(
                        user_id,
                        request.headers.get("referer", "-"),
                        "/".join(str(request.url).split("/")[3:]),
                        request.method,
                        text.split(" || ")[-1],
                    ),
                    many=False,
                )
            if _error is not None:
                raise Exception(_error)
            return response

        return decorated_function

    return decorator_inline


def db_query_sqlite(query: str, many: bool, args=(), source="database.db") -> tuple | list[tuple] | None:
    with sqlite3.connect(source, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as connection:
        cursor = connection.cursor()
        cursor.execute(query, args)
        try:
            if many:
                return cursor.fetchall()
            return cursor.fetchone()
        except Exception as _:
            return None


def request_to_oracle(query: str, args: dict = None, many: bool = True) -> tuple | list[tuple] | None:
    _ = """
sudo su
mkdir -p /opt/oracle
cd /opt/oracle
wget https://download.oracle.com/otn_software/linux/instantclient/214000/instantclient-basic-linux.x64-21.4.0.0.0dbru.zip
unzip instantclient_21_4
apt install libaio1
echo /opt/oracle/instantclient_21_4 > /etc/old.so.conf.d/oracle-instantclient.conf
ldconfig
pip install cx_Oracle
exit
"""
    try:
        oracledb.init_oracle_client(
            lib_dir=r"C:\ADDITIONAL\web_platform\instantclient_21_9_lite"
            if constants.IS_SERVER
            else r"C:\development\projects\instantclient_21_9_lite"
        )
    except Exception as err:
        print(err)
        pass
    try:
        with oracledb.connect(constants.oracle_db_connection_string) as connection:
            with connection.cursor() as cursor:
                if args is None:
                    args = {}
                cursor.execute(query, args)
                if many:
                    return cursor.fetchall()
                return cursor.fetchone()
    except Exception as err:
        raise err


def request_to_oracle_emulate(path_to_file: str, with_headers: bool) -> tuple[tuple]:
    workbook: Workbook = openpyxl.load_workbook(path_to_file)
    worksheet: Worksheet = workbook.active
    matrix: tuple[tuple] = tuple(
        worksheet.iter_rows(
            min_row=1 if with_headers else 2,
            min_col=1,
            max_row=worksheet.max_row,
            max_col=worksheet.max_column,
            values_only=True,
        )
    )
    workbook.close()
    return matrix


def get_selected_shift(date_time: datetime.datetime) -> int:
    return 1 if date_time.hour < 8 or date_time.hour >= 20 else 2


def get_selected_taskdate(
    date_time: datetime.datetime,
) -> datetime.datetime:
    # now1 = datetime.datetime.strptime("2023-08-24T19:00:00.0", "%Y-%m-%dT%H:%M:%S.%f")
    # now2 = datetime.datetime.strptime("2023-08-24T23:00:00.0", "%Y-%m-%dT%H:%M:%S.%f")
    # now3 = datetime.datetime.strptime("2023-08-25T00:30:00.0", "%Y-%m-%dT%H:%M:%S.%f")
    # now4 = datetime.datetime.strptime("2023-08-25T01:30:00.0", "%Y-%m-%dT%H:%M:%S.%f")
    # now5 = datetime.datetime.strptime("2023-08-25T07:00:00.0", "%Y-%m-%dT%H:%M:%S.%f")
    # now6 = datetime.datetime.strptime("2023-08-25T08:30:00.0", "%Y-%m-%dT%H:%M:%S.%f")
    # print(now1, get_selected_taskdate(date_time=now1), get_selected_shift(date_time=now1))
    # print(now2, get_selected_taskdate(date_time=now2), get_selected_shift(date_time=now2))
    # print(now3, get_selected_taskdate(date_time=now3), get_selected_shift(date_time=now3))
    # print(now4, get_selected_taskdate(date_time=now4), get_selected_shift(date_time=now4))
    # print(now5, get_selected_taskdate(date_time=now5), get_selected_shift(date_time=now5))
    # print(now6, get_selected_taskdate(date_time=now6), get_selected_shift(date_time=now6))
    if get_selected_shift(date_time) == 2:
        return date_time.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        date_time = date_time + datetime.timedelta(days=1 if date_time.hour >= 20 else 0)
        return date_time.replace(hour=0, minute=0, second=0, microsecond=0)


def get_shift_datetime_begin(
    date_time: datetime.datetime,
) -> datetime.datetime:
    taskdate = get_selected_taskdate(date_time)
    if get_selected_shift(date_time) == 2:
        return taskdate.replace(hour=8, minute=0, second=0, microsecond=0)
    else:
        return taskdate.replace(hour=20, minute=0, second=0, microsecond=0)


def get_shift_datetime_end(
    date_time: datetime.datetime,
) -> datetime.datetime:
    taskdate = get_selected_taskdate(date_time)
    if get_selected_shift(date_time) == 2:
        return taskdate.replace(hour=20, minute=0, second=0, microsecond=0)
    else:
        return taskdate.replace(hour=8, minute=0, second=0, microsecond=0)


def get_all_dumptrucks(exclude: list[str] | None = None) -> list[str]:
    __rows_raw: list[tuple] = request_to_oracle(
        query="""SELECT VEHID FROM DUMPTRUCKS""",
        args={},
        many=True,
    )
    if exclude is None:
        return [str(x[0]) for x in __rows_raw]
    return [str(x[0]) for x in __rows_raw if str(x[0]) not in exclude]
    # with open(
    #     "../frontend/src/data/dumptrucks.json", mode="r", encoding="utf-8"
    # ) as f:
    #     _file: dict = json.load(f)
    #     _list: list[str] = [
    #         str(x["tech"]) for x in _file["list"] if str(x["type"]) == "самосвал"
    #     ]
    # if exclude is None:
    #     return _list
    # return [x for x in _list if x not in exclude]


def get_all_shovels(exclude: list[str] | None = None) -> list[str]:
    __rows_raw: list[tuple] = request_to_oracle(
        query="""SELECT SHOVID FROM SHOVELS""",
        args={},
        many=True,
    )
    if exclude is None:
        return [str(x[0]) for x in __rows_raw]
    return [str(x[0]) for x in __rows_raw if str(x[0]) not in exclude]
    # with open("../frontend/src/data/shovels.json", mode="r", encoding="utf-8") as f:
    #     _file: dict = json.load(f)
    #     _list: list[str] = [
    #         str(x["tech"])
    #         for x in _file["list"]
    #         if str(x["type"]) == "экскаватор" or x["type"] == "экскаватор ЭШ"
    #     ]
    # if exclude is None:
    #     return _list
    # return [x for x in _list if x not in exclude]


def get_all_auxes(exclude: list[str] | None = None) -> list[str]:
    __rows_raw: list[tuple] = request_to_oracle(
        query="""SELECT AUXID FROM AUXTECHNICS""",
        args={},
        many=True,
    )
    if exclude is None:
        return [str(x[0]) for x in __rows_raw]
    return [str(x[0]) for x in __rows_raw if str(x[0]) not in exclude]
    # with open("../frontend/src/data/auxes.json", mode="r", encoding="utf-8") as f:
    #     _file: dict = json.load(f)
    #     _list: list[str] = [str(x["tech"]) for x in _file["list"]]
    # if exclude is None:
    #     return _list
    # return [x for x in _list if x not in exclude]


def export_to_excel(
    _data: list[list],
    _titles: list[str],
    _folder_name: str,
    _prefix: str,
) -> str:
    key = create_encrypted_password(
        _random_chars="abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890",
        _length=16,
    )
    date = f"{time.strftime('%Y-%m-%d')}"
    path = f"media/data/temp/{_folder_name}"
    file_name = f"{_prefix}_{date}_{key}.xlsx"
    workbook: Workbook = openpyxl.Workbook()
    worksheet: Worksheet = workbook.active

    # Delete old files
    for root, dirs, files in os.walk(path, topdown=True):
        for file in files:
            try:
                date_file = str(file).strip().split("_")[-2].strip()
                if date != date_file:
                    os.remove(f"{path}/{file}")
            except Exception as _:
                pass
    # worksheet.cell(row=1, column=1, value="РАБОЧИЙ ЛИСТ")
    # worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    for col_idx, value in enumerate(_titles, 1):
        worksheet.cell(row=1, column=col_idx, value=value)

    for row_idx, row in enumerate(_data, 2):
        for col_idx, value in enumerate(row, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{len(_data) + 1}"

    side_think = Side(border_style="thin", color="FF808080")
    # {'mediumDashDotDot', 'thin', 'dashed', 'mediumDashed', 'dotted', 'double', 'thick',
    # 'medium', 'dashDot','dashDotDot', 'hair', 'mediumDashDot', 'slantDashDot'}
    border_think = Border(
        top=side_think,
        left=side_think,
        right=side_think,
        bottom=side_think,
    )
    aligm_think = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
        shrink_to_fit=True,
    )
    font_header = Font(name="Arial", size=12, bold=True)

    # headers
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = font_header
        cell.border = border_think

    font_body = Font(name="Arial", size=10, bold=False)
    green_fill = PatternFill(fgColor="F04141", fill_type="solid")

    # body
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.alignment = aligm_think
            cell.border = border_think
            # if col_idx == 4 + 1 and cell.value < param_target_speed:
            #     cell.fill = green_fill

    # set column width
    for col_idx in range(1, worksheet.max_column + 1):
        width: int = 1
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value: int = len(str(cell.value))
            if value > width:
                width = value
        worksheet.column_dimensions[get_column_letter(col_idx)].height = 1
        worksheet.column_dimensions[get_column_letter(col_idx)].width = round((width * 1.25), 3)
        # worksheet.column_dimensions[get_column_letter(col_idx)].auto_size = True
        # worksheet.column_dimensions[get_column_letter(col_idx)].bestFit = True

    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LETTER
    worksheet.page_margins.left = 0.05
    worksheet.page_margins.right = 0.05
    worksheet.page_margins.header = 0.1
    worksheet.page_margins.bottom = 0.2
    worksheet.page_margins.footer = 0.2
    worksheet.page_margins.top = 0.1
    worksheet.print_options.horizontalCentered = True
    # sheet.print_options.verticalCentered = True
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.scale = 100
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.fitToWidth = 1
    # worksheet.protection.password = key + "_1"
    # worksheet.protection.sheet = True
    # worksheet.protection.enable()
    full_path = f"{path}/{file_name}"
    workbook.save(full_path)
    return full_path


def create_encrypted_password(
    _random_chars="abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890",
    _length=8,
):
    password = ""
    for i in range(1, _length + 1):
        password += random.choice(_random_chars)
    return password


def raw_tuple_list_to_list_dictionaries(titles: list[str], trips_raw: list[tuple]) -> list[dict]:
    trips_dict: list[dict] = []
    for trip in trips_raw:
        for index, title in enumerate(titles, 0):
            trips_dict.append({f"{title}": trip[index]})
    return trips_dict


def print_time_beauty(t_time: datetime.datetime) -> str:
    return f"{t_time.hour}:{t_time.minute}:{t_time.second}"


if __name__ == "__main__":
    pass
