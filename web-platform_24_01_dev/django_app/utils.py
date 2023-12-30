import contextlib
import datetime
import json
import os
import random
import sqlite3
import time
from functools import wraps

import openpyxl
from django.core.cache import caches
from django.http import HttpResponse, JsonResponse, HttpRequest
import oracledb
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from django_settings import settings

LocMemCache = caches["default"]


def decorator_json(need_auth=False, is_cache: bool = True, cache_timeout: int = 1):
    def decorator_inline(function: callable):
        @wraps(function)
        def decorated_function(request: HttpRequest, *args, **kwargs):
            if is_cache:
                return JsonResponse(
                    data={"response": Cache.caching(query=lambda: function(request, *args, **kwargs), request=request, timeout=cache_timeout)},
                    status=200,
                )
            # TODO request2 = Request2(request)
            response = function(request, *args, **kwargs)
            return response

        return decorated_function

    return decorator_inline


class Request:
    class CustomRequest:
        def __init__(self, request: HttpRequest):
            self.request = request
            self.params = {}
            for k, v in request.GET.items():
                self.params[k] = v
            try:
                for k, v in json.loads(request.body).items():
                    self.params[k] = v
            except Exception as error:
                pass
            for k, v in request.POST.items():
                self.params[k] = v
            for k, v in request.FILES.items():
                self.params[k] = v

        def param(self, name: str, default: any = None, is_safe: bool = False):
            if is_safe:
                return self.params.get(name, default)
            return self.params[name]

    @staticmethod
    def decorator_json_response(need_auth=False, is_cache: bool = True, cache_timeout: int = 1) -> any:
        def decorator_inline(function: callable) -> any:
            @wraps(function)
            def decorated_function(request: HttpRequest, *args, **kwargs) -> any:
                _error = None
                response = None
                # try:
                # TODO замер производительности
                time_start_func = time.perf_counter()

                # TODO проверка подсети
                client_ip = request.META.get("HTTP_HOST", "")
                # if settings.IS_ASD_SERVER:
                #     try:
                #         subnet = client_ip.split(".")[-2]
                #         if subnet not in ["16", "17", "23", "200"]:
                #             with open("static/clients.txt", "a") as file:
                #                 file.write(f"{client_ip}\n")
                #             print("Subnet: ", subnet, " Нет доступа!")
                #             raise Exception("Нет доступа!")
                #     except Exception as error:
                #         raise Exception("Нет доступа!")

                # TODO проверка токена
                # print("headers: ", request.headers)
                # if request.get_host() != "127.0.0.1:81":
                # if str(request.headers.get("Authorization", "")) != "Token=auth_token":
                #     raise Exception("Неверный токен доступа!")

                c_request = Request.CustomRequest(request=request)
                if is_cache and request.method == "GET":
                    return JsonResponse(
                        data={"response": Cache.caching(query=lambda: function(c_request, *args, **kwargs), request=request, timeout=cache_timeout)},
                        status=200,
                    )
                response = function(c_request, *args, **kwargs)
                elapsed_time = round((time.perf_counter() - time_start_func), 2)
                user_id = -1
                if need_auth:
                    # token_access: str | None = request.cookies.get("token_access", None)
                    # if token_access is None:
                    #     raise Exception("Need Authorization!")
                    # token: tuple = db_query_sqlite(
                    #     query="""SELECT user_id, datetime_elapsed FROM tokens WHERE token_access = ?""",
                    #     args=(token_access,),
                    #     many=False,
                    # )
                    # user_id: int = token[0]
                    # datetime_elapsed: datetime.datetime = datetime.datetime.strptime(token[1], "%Y-%m-%d %H:%M:%S")
                    # if datetime.datetime.now() > datetime_elapsed + datetime.timedelta(hours=6, seconds=constants.jwt_token_lifetime_seconds):
                    #     raise Exception("Token is terminate")
                    # request.user_id = user_id
                    pass
                text = (
                    f"{client_ip} | {str(datetime.datetime.now())[0:-5:1]}({elapsed_time}s) ({request.method})/{'/'.join(str(request.path).split('/')[3:])} || "
                    f"{response if settings.C_LOGGING_RESPONSE else '[successfully response disabled]'}"
                )
                # except Exception as error:
                #     _error = str(error)
                #     user_id = -1
                #     text = f"{str(datetime.datetime.now())[0:-5:1]} ({request.method})/{'/'.join(str(request.url).split('/')[3:])} || {error}"

                if settings.C_LOGGING_TO_CONSOLE:
                    print("\n" + text)
                if settings.C_LOGGING_TO_FILE:
                    with open("static/log.txt", mode="a", encoding="utf-8") as file:
                        file.write(text + "\n")
                # TODO
                if settings.C_LOGGING_TO_DATABASE:
                    # db_query_sqlite(
                    #     query="""INSERT INTO logs (user_id, ip, path, method, data) VALUES (?, ?, ?, ?, ?)""",
                    #     args=(
                    #         user_id,
                    #         request.headers.get("referer", "-"),
                    #         "/".join(str(request.url).split("/")[3:]),
                    #         request.method,
                    #         text.split(" || ")[-1],
                    #     ),
                    #     many=False,
                    # )
                    pass
                if _error is not None:
                    raise Exception(_error)

                return JsonResponse(
                    data={"response": response},
                    status=200,
                )

            return decorated_function

        return decorator_inline


class Except:
    @staticmethod
    def decorator_http_exception(func: callable):
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as error:
                print("error: ", error)
                return HttpResponse(str(error))

        return wrapper

    @staticmethod
    def decorator_json_exception(func: callable):
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as error:
                print("error: ", error)
                return JsonResponse(data={"error": str(error)}, status=500, safe=False)

        return wrapper


class Cache:
    @staticmethod
    def caching(
        key: str = None,
        request: HttpRequest = None,
        query: callable = None,
        timeout: float = 1.0,
    ) -> any:
        if key is None:
            key = f"{request.path}_{request.method}_{''.join([str(x) for x in request.GET.values()])}"
        value = LocMemCache.get(key, None)
        if value is None and query:
            value = query()
            LocMemCache.set(key=key, value=value, timeout=timeout)
        else:
            # print("use CACHE")
            pass
        return value

    @staticmethod
    def caching_simple(key: str, timeout: int, query: any = lambda: any) -> any:
        data = LocMemCache.get(key)
        if not data:
            data = query()
            LocMemCache.set(key, data, timeout=timeout)
        return data


class Sql:
    @staticmethod
    def execute_sql(_source: str, _query: str, _kwargs: dict, is_many: bool):
        with contextlib.closing(sqlite3.connect(f"database/{_source}")) as connection:
            cursor = connection.cursor()
            cursor.execute(_query, _kwargs)
            connection.commit()
            if is_many:
                return cursor.fetchall()
            return cursor.fetchone()

    @staticmethod
    def request_to_oracle(
        query: str, args: dict = None, many: bool = True, is_fake: bool = False, fake_file_path: str = ""
    ) -> tuple | list[tuple] | None:
        if is_fake:
            try:
                workbook: Workbook = openpyxl.load_workbook(fake_file_path)
                worksheet: Worksheet = workbook.active
                matrix: list[tuple] = list(
                    (x for x in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column, values_only=True))
                )
                data = []
                l_data = []
                for i in range(0, len(matrix[0])):
                    l_data.append(i)
                data.append(l_data)
                data.extend(matrix)
                time.sleep(1.0)
                return data
            except Exception as error:
                raise error

        _ = """
    sudo su
    mkdir -p /opt/oracle
    cd /opt/oracle
    wget https://download.oracle.com/otn_software/linux/instantclient/214000/instantclient-basic-linux.x64-21.4.0.0.0dbru.zip
    unzip instantclient_21_4
    apt install libaio1
    echo /opt/oracle/instantclient_21_4 > /etc/old.so.conf.d/oracle-instantclient.conf
    ldconfig
    pip install cx_Oracle
    exit
    """
        try:
            oracledb.init_oracle_client(
                lib_dir=r"C:\ADDITIONAL\web_platform\instantclient_21_9_lite"
                if settings.DEBUG
                else r"C:\development\projects\instantclient_21_9_lite"
            )
        except Exception as err:
            print(err)
            pass
        try:
            with oracledb.connect(settings.ORACLE_CONN_STR) as connection:
                with connection.cursor() as cursor:
                    if args is None:
                        args = {}
                    cursor.execute(query, args)
                    if many:
                        return cursor.fetchall()
                    return cursor.fetchone()
        except Exception as err:
            raise err


class DateTime:
    format = "%d.%m.%Y %H:%M:%S"

    @staticmethod
    def now_str():
        return datetime.datetime.now().strftime(DateTime.format)

    @staticmethod
    def convert_from_str(_datetime_str: str) -> datetime.datetime:
        return datetime.datetime.strptime(_datetime_str, DateTime.format)

    @staticmethod
    def convert_from_datetime(_datetime: datetime.datetime) -> str:
        return _datetime.strftime(DateTime.format)


class Excel:
    @staticmethod
    def export_to_excel(
        _data: list[list],
        _titles: list[str],
        _folder_name: str,
        _prefix: str,
    ) -> str:
        key = create_encrypted_password(
            _random_chars="abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890",
            _length=16,
        )
        date = f"{time.strftime('%Y-%m-%d')}"
        path = f"static/media/data/temp/{_folder_name}"
        file_name = f"{_prefix}_{date}_{key}.xlsx"
        workbook: Workbook = openpyxl.Workbook()
        worksheet: Worksheet = workbook.active

        # Delete old files
        for root, dirs, files in os.walk(path, topdown=True):
            for file in files:
                try:
                    date_file = str(file).strip().split("_")[-2].strip()
                    if date != date_file:
                        os.remove(f"{path}/{file}")
                except Exception as _:
                    pass
        # worksheet.cell(row=1, column=1, value="РАБОЧИЙ ЛИСТ")
        # worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

        for col_idx, value in enumerate(_titles, 1):
            worksheet.cell(row=1, column=col_idx, value=value)

        for row_idx, row in enumerate(_data, 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{len(_data) + 1}"

        side_think = Side(border_style="thin", color="FF808080")
        # {'mediumDashDotDot', 'thin', 'dashed', 'mediumDashed', 'dotted', 'double', 'thick',
        # 'medium', 'dashDot','dashDotDot', 'hair', 'mediumDashDot', 'slantDashDot'}
        border_think = Border(
            top=side_think,
            left=side_think,
            right=side_think,
            bottom=side_think,
        )
        aligm_think = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
            shrink_to_fit=True,
        )
        font_header = Font(name="Arial", size=12, bold=True)

        # headers
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = font_header
            cell.border = border_think

        font_body = Font(name="Arial", size=10, bold=False)
        green_fill = PatternFill(fgColor="F04141", fill_type="solid")

        # body
        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.alignment = aligm_think
                cell.border = border_think
                # if col_idx == 4 + 1 and cell.value < param_target_speed:
                #     cell.fill = green_fill

        # set column width
        for col_idx in range(1, worksheet.max_column + 1):
            width: int = 1
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value: int = len(str(cell.value))
                if value > width:
                    width = value
            worksheet.column_dimensions[get_column_letter(col_idx)].height = 1
            worksheet.column_dimensions[get_column_letter(col_idx)].width = round((width * 1.25), 3)
            # worksheet.column_dimensions[get_column_letter(col_idx)].auto_size = True
            # worksheet.column_dimensions[get_column_letter(col_idx)].bestFit = True

        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LETTER
        worksheet.page_margins.left = 0.05
        worksheet.page_margins.right = 0.05
        worksheet.page_margins.header = 0.1
        worksheet.page_margins.bottom = 0.2
        worksheet.page_margins.footer = 0.2
        worksheet.page_margins.top = 0.1
        worksheet.print_options.horizontalCentered = True
        # sheet.print_options.verticalCentered = True
        worksheet.page_setup.fitToHeight = 1
        worksheet.page_setup.scale = 100
        worksheet.page_setup.fitToHeight = 1
        worksheet.page_setup.fitToWidth = 1
        # worksheet.protection.password = key + "_1"
        # worksheet.protection.sheet = True
        # worksheet.protection.enable()
        full_path = f"{path}/{file_name}"
        workbook.save(full_path)
        return full_path


def create_encrypted_password(
    _random_chars="abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890",
    _length=8,
):
    password = ""
    for i in range(1, _length + 1):
        password += random.choice(_random_chars)
    return password
