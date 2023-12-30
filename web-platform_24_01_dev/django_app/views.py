import datetime
import json
import time
import socket

import openpyxl
import requests
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from django_app import utils, queries
from django_settings import settings


@utils.Except.decorator_http_exception
def index(request):
    # print("PATH: ", request.path)
    # return HttpResponse("ok")
    # return render(request, "build/index.html", context={"path": request.path})
    return render(request, "build/index.html")


class Idea:
    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def api_idea_create(req: utils.Request.CustomRequest) -> dict:
        param_author: str = str(req.param("param_author", is_safe=True, default=""))
        param_subdivision: str = str(req.param("param_subdivision", is_safe=True, default=""))
        param_position: str = str(req.param("param_position", is_safe=True, default=""))
        param_phone: str = str(req.param("param_phone", is_safe=True, default=""))
        param_email: str = str(req.param("param_email", is_safe=True, default=""))
        param_title: str = str(req.param("param_title", is_safe=True, default=""))
        param_description: str = str(req.param("param_description", is_safe=True, default=""))
        param_place: str = str(req.param("param_place", is_safe=True, default=""))
        param_effect: str = str(req.param("param_effect", is_safe=True, default=""))
        param_need: str = str(req.param("param_need", is_safe=True, default=""))
        param_is_feedback: bool = True if str(req.param("param_is_feedback", is_safe=True, default="")) == "да" else False
        param_link: str = str(req.param("param_link", is_safe=True, default=""))

        # create table
        utils.Sql.execute_sql(
            _source="idea.db",
            _query=queries.Idea.table(),
            _kwargs={},
            is_many=True,
        )
        # insert new row
        now = utils.DateTime.now_str()
        utils.Sql.execute_sql(
            _source="idea.db",
            _query=queries.Idea.insert(),
            _kwargs={
                "author": param_author,
                "subdivision": param_subdivision,
                "position": param_position,
                "phone": param_phone,
                "email": param_email,
                "title": param_title,
                "description": param_description,
                "place": param_place,
                "effect": param_effect,
                "need": param_need,
                "is_feedback": 1 if param_is_feedback else 0,
                "link": param_link,
                "created": now,
            },
            is_many=True,
        )

        return {"data": "Успешно отправлено(сәтті жіберілді)"}

    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=1)
    def api_idea_export(req: utils.Request.CustomRequest) -> dict:
        # create table
        utils.Sql.execute_sql(
            _source="idea.db",
            _query=queries.Idea.table(),
            _kwargs={},
            is_many=True,
        )
        # select all rows
        _raw_rows = utils.Sql.execute_sql(
            _source="idea.db",
            _query=queries.Idea.select_all(),
            _kwargs={},
            is_many=True,
        )
        # serialize
        _ideas = []
        for i in _raw_rows:
            _ideas.append(
                {
                    "id": int(i[0]),
                    "author": str(i[1]),
                    "subdivision": str(i[2]),
                    "position": str(i[3]),
                    "phone": str(i[4]),
                    "email": str(i[5]),
                    "title": str(i[6]),
                    "description": str(i[7]),
                    "place": str(i[8]),
                    "effect": str(i[9]),
                    "need": str(i[10]),
                    "is_feedback": True if int(i[11]) == 1 else False,
                    "link": str(i[12]),
                    "created": utils.DateTime.convert_from_str(str(i[13])),
                }
            )
        # sorting
        _ideas = sorted(_ideas, key=lambda x: x["created"], reverse=True)
        workbook: Workbook = openpyxl.Workbook()
        worksheet: Worksheet = workbook.active
        titles = [
            "ФИО(полностью)",
            "Подразделение",
            "Профессия, должность",
            "Номер телефона",
            "Электронная почта",
            "Предложение(наименование)",
            "Описание",
            "Краткое описание объекта",
            "Ожидаемый эффект",
            "Необходимые ТМЦ",
            "Нужно ли связаться с работником",
            "Ссылка",
            "Дата и время создания",
        ]
        for column, title in enumerate(titles, 1):
            worksheet.cell(row=1, column=column, value=str(title))
        for row, idea in enumerate(_ideas, 2):
            worksheet.cell(row=row, column=1, value=str(idea.get("author", "")))
            worksheet.cell(row=row, column=2, value=str(idea.get("subdivision", "")))
            worksheet.cell(row=row, column=3, value=str(idea.get("position", "")))
            worksheet.cell(row=row, column=4, value=str(idea.get("phone", "")))
            worksheet.cell(row=row, column=5, value=str(idea.get("email", "")))
            worksheet.cell(row=row, column=6, value=str(idea.get("title", "")))
            worksheet.cell(row=row, column=7, value=str(idea.get("description", "")))
            worksheet.cell(row=row, column=8, value=str(idea.get("place", "")))
            worksheet.cell(row=row, column=9, value=str(idea.get("effect", "")))
            worksheet.cell(row=row, column=10, value=str(idea.get("need", "")))
            worksheet.cell(row=row, column=11, value="нужно связаться" if idea.get("is_feedback", True) else "не нужно связываться")
            worksheet.cell(row=row, column=12, value=str(idea.get("link", "")))
            worksheet.cell(row=row, column=13, value=utils.DateTime.convert_from_datetime(idea.get("created", "")))
        filename = f"static/media/idea/export/{datetime.datetime.now().strftime('%m_%d_%Y__%H_%M_%S')}.xlsx"
        workbook.save(filename)
        return {"data": {"path_to_excel_file": filename}}


class Claim:
    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def api_claim_create(req: utils.Request.CustomRequest) -> dict:
        param_author: str = str(req.param("param_author", is_safe=True, default=""))
        param_description: str = str(req.param("param_description", is_safe=True, default=""))
        param_tech: str = str(req.param("param_tech", is_safe=True, default=""))

        # create table
        utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.table(),
            _kwargs={},
            is_many=True,
        )
        # insert new row
        now = utils.DateTime.now_str()
        utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.insert(),
            _kwargs={"author": param_author, "description": param_description, "tech": param_tech, "is_active": 1, "updated": now, "created": now},
            is_many=True,
        )
        # count active
        claim_count = utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.select_active(),
            _kwargs={},
            is_many=False,
        )[0]

        # send notifications to telegram
        text = f"Новая заявка:\n{param_author}({param_tech})\n{param_description}\n\nКоличество активных заявок: {claim_count}\nСписок всех заявок: https://kgp.lol/claim/list/"
        bot_token = "6619466319:AAFQ5XcFnDIR1PWCVrT4GONq6LiD_ceYkZk"
        # "1289279426(Андриенко),5171879758(Оспанов),806309122(Ивайкин),458411477(Кувалдин),5546367644(горный),5452471486(горный),5659943979(горный),6701232381(горный)"
        user_list = "1289279426,458411477,5546367644,5452471486,5659943979,6701232381"
        # user_list = "1289279426"
        for usr in [x.strip() for x in user_list.split(",")]:
            try:
                response = requests.post(url=f"https://api.telegram.org/bot{bot_token}/sendMessage", json={"chat_id": usr, "text": text})
                if response.status_code not in (200, 201):
                    raise Exception(response.status_code)
            except Exception as error:
                print(error)
                time.sleep(1.0)
                response = requests.post(url=f"https://api.telegram.org/bot{bot_token}/sendMessage", json={"chat_id": usr, "text": text})
                if response.status_code not in (200, 201):
                    raise Exception(response.status_code)
        return {"data": "Успешно отправлено(сәтті жіберілді)"}

    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=1)
    def api_claim_list(req: utils.Request.CustomRequest) -> dict:
        # create table
        utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.table(),
            _kwargs={},
            is_many=True,
        )
        # select all rows
        _raw_rows = utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.select_all(),
            _kwargs={},
            is_many=True,
        )
        # serialize
        _claims = []
        for i in _raw_rows:
            _claims.append(
                {
                    "id": int(i[0]),
                    "author": str(i[1]),
                    "description": str(i[2]),
                    "tech": str(i[3]),
                    "is_active": True if int(i[4]) == 1 else False,
                    "updated": utils.DateTime.convert_from_str(_datetime_str=str(i[5])),
                    "created": utils.DateTime.convert_from_str(_datetime_str=str(i[6])),
                }
            )
        # sorting
        _claims = sorted(_claims, key=lambda x: (x["is_active"], x["created"]), reverse=True)
        return {"data": _claims}

    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def api_claim_update(req: utils.Request.CustomRequest) -> dict:
        param_id: int = int(req.param("param_id", is_safe=False))

        # create table
        utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.table(),
            _kwargs={},
            is_many=True,
        )
        # update old row
        _raw_rows = utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.update(),
            _kwargs={"updated": utils.DateTime.now_str(), "id": param_id},
            is_many=True,
        )
        # select one row
        _raw_row = utils.Sql.execute_sql(
            _source="claim.db",
            _query=queries.Claim.select_one(),
            _kwargs={"id": param_id},
            is_many=False,
        )
        i = _raw_row
        claim = {
            "id": int(i[0]),
            "author": str(i[1]),
            "description": str(i[2]),
            "tech": str(i[3]),
            "is_active": True if int(i[4]) == 1 else False,
            "updated": utils.DateTime.convert_from_str(_datetime_str=str(i[5])),
            "created": utils.DateTime.convert_from_str(_datetime_str=str(i[6])),
        }
        if not claim.get("is_active", None):
            # count active
            claim_count = utils.Sql.execute_sql(
                _source="claim.db",
                _query=queries.Claim.select_active(),
                _kwargs={},
                is_many=False,
            )[0]
            text = f"Заявка выполнена:\n{claim.get('author', '')}({claim.get('tech', '')})\n\nКоличество активных заявок: {claim_count}\nСписок всех заявок: https://kgp.lol/claim/list/"
            bot_token = "6619466319:AAFQ5XcFnDIR1PWCVrT4GONq6LiD_ceYkZk"
            # "1289279426(Андриенко),5171879758(Оспанов),806309122(Ивайкин),458411477(Кувалдин),5546367644(горный),5452471486(горный),5659943979(горный),6701232381(горный)"
            user_list = "1289279426,458411477,5546367644,5452471486,5659943979,6701232381"
            # user_list = "1289279426"
            for usr in [x.strip() for x in user_list.split(",")]:
                try:
                    response = requests.post(url=f"https://api.telegram.org/bot{bot_token}/sendMessage", json={"chat_id": usr, "text": text})
                    if response.status_code not in (200, 201):
                        raise Exception(response.status_code)
                except Exception as error:
                    print(error)
                    time.sleep(1.0)
                    response = requests.post(url=f"https://api.telegram.org/bot{bot_token}/sendMessage", json={"chat_id": usr, "text": text})
                    if response.status_code not in (200, 201):
                        raise Exception(response.status_code)
        return {"data": "OK"}


@csrf_exempt
@utils.Except.decorator_json_exception
def api_proxy_bot1(request):
    if request.method != "POST":
        raise Exception("Method not allowed")
    if str(request.headers.get("Authorization", "")) != "Token=auth_token":
        raise Exception("Неверный токен доступа!")
    form_data = json.loads(request.body)
    bot_token = str(form_data["bot_token"])
    user_list = [x.strip() for x in str(form_data["user_list"]).split(",")]
    text = str(form_data["text"])
    for user in user_list:
        try:
            response = requests.post(url=f"https://api.telegram.org/bot{bot_token}/sendMessage", json={"chat_id": user, "text": text})
            if response.status_code not in (200, 201):
                raise Exception(response.status_code)
        except Exception as error:
            print(error)
            response = requests.post(url=f"https://api.telegram.org/bot{bot_token}/sendMessage", json={"chat_id": user, "text": text})
            if response.status_code not in (200, 201):
                raise Exception(response.status_code)
    return JsonResponse({"message": "OK"}, status=201)


@csrf_exempt
@utils.Except.decorator_json_exception
def api_proxy_sql(request):
    try:
        # Получаем SQL-запрос из запроса
        query = str(request.POST["query"]).encode(encoding="utf-8")

        # Создаем сокет
        server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

        # Привязываем сокет
        server.bind(("192.168.1.204", 81))
        server.listen(5)

        # Ожидаем подключения клиента
        client_socket, addr = server.accept()

        try:
            # Отправляем запрос на сервер
            client_socket.sendall(query)

            # Читаем ответ от сервера
            response = b""
            while True:
                data = client_socket.recv(1024)
                print("d: ", data)
                if not data:
                    break
                response += data
            print("response: ", response)

            # Декодируем и разбираем JSON
            response_json = json.loads(response.decode(encoding="utf-8"))

            return JsonResponse({"message": response_json}, status=201)
        finally:
            # Закрываем соединение с клиентом
            client_socket.close()
    except Exception as error:
        return JsonResponse({"error": str(error)}, status=500)
    finally:
        # Закрываем серверный сокет
        server.close()


class Center:
    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def api_communicator_center_monitoring(req: utils.Request.CustomRequest) -> dict:
        if req.request.method == "GET":
            # create table
            utils.Sql.execute_sql(
                _source="/center/event_last.db",
                _query=queries.Center.table_event_last(),
                _kwargs={},
                is_many=True,
            )
            # select all rows
            _raw_rows = utils.Sql.execute_sql(
                _source="/center/event_last.db",
                _query=queries.Center.select_all_event_last(),
                _kwargs={},
                is_many=True,
            )
            # serialize
            _events = []
            for i in _raw_rows:
                _events.append(
                    {
                        "id": str(i[0]),
                        "subsystem": str(i[1]),
                        "message": json.loads(str(i[2])),
                        "date_time_subsystem": str(i[3]),
                        "date_time_server": str(i[4]),
                    }
                )
            # groupping
            group_by_subsystem = {}
            for j in _events:
                group_by_subsystem[j["subsystem"]] = j
            return group_by_subsystem
        elif req.request.method == "POST":
            form_data = json.loads(req.request.body)
            subsystem = form_data["subsystem"]
            data = form_data["data"]
            date_time_subsystem = form_data["date_time_subsystem"]
            date_time_server = utils.DateTime.now_str()

            # create table
            utils.Sql.execute_sql(
                _source="/center/event_last.db",
                _query=queries.Center.table_event_last(),
                _kwargs={},
                is_many=True,
            )
            # update
            utils.Sql.execute_sql(
                _source="/center/event_last.db",
                _query=queries.Center.insert_or_replace_event_last(),
                _kwargs={
                    "subsystem": str(subsystem),
                    "message": str(json.dumps(data, ensure_ascii=False)),
                    "date_time_subsystem": str(date_time_subsystem),
                    "date_time_server": str(date_time_server),
                },
                is_many=True,
            )

            database_name = datetime.datetime.now().strftime("%Y_%m_%d")
            # create table
            utils.Sql.execute_sql(
                _source=f"/center/event_history_{database_name}.db",
                _query=queries.Center.table_event_history(),
                _kwargs={},
                is_many=True,
            )
            # insert
            utils.Sql.execute_sql(
                _source=f"/center/event_history_{database_name}.db",
                _query=queries.Center.insert_event_history(),
                _kwargs={
                    "subsystem": str(subsystem),
                    "message": str(json.dumps(data, ensure_ascii=False)),
                    "date_time_subsystem": str(date_time_subsystem),
                    "date_time_server": str(date_time_server),
                },
                is_many=True,
            )

            return {"message": "ok"}

    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def api_communicator_center_sticking(req: utils.Request.CustomRequest) -> dict:
        if req.request.method == "GET":
            # create table
            utils.Sql.execute_sql(
                _source="/center/sticking.db",
                _query=queries.CenterSticking.table(),
                _kwargs={},
                is_many=True,
            )
            # select all rows
            _raw_rows = utils.Sql.execute_sql(
                _source="/center/sticking.db",
                _query=queries.CenterSticking.select_all(),
                _kwargs={},
                is_many=True,
            )
            # serialize
            i = _raw_rows[0]
            _trips = {
                "id": str(i[0]),
                "subsystem": str(i[1]),
                "message": json.loads(str(i[2])),
                "date_time_subsystem": str(i[3]),
                "date_time_server": str(i[4]),
            }
            return {"data": _trips}
        elif req.request.method == "POST":
            form_data = json.loads(req.request.body)
            subsystem = form_data["subsystem"]
            data = form_data["data"]
            date_time_subsystem = form_data["date_time_subsystem"]
            date_time_server = utils.DateTime.now_str()

            # create table
            utils.Sql.execute_sql(
                _source="/center/sticking.db",
                _query=queries.CenterSticking.table(),
                _kwargs={},
                is_many=True,
            )
            # update
            utils.Sql.execute_sql(
                _source="/center/sticking.db",
                _query=queries.CenterSticking.insert_or_replace(),
                _kwargs={
                    "subsystem": str(subsystem),
                    "message": str(json.dumps(data, ensure_ascii=False)),
                    "date_time_subsystem": str(date_time_subsystem),
                    "date_time_server": str(date_time_server),
                },
                is_many=True,
            )

            return {"message": "ok"}


class Stoppages:
    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def get_stoppages_report_veh_dvs(req: utils.Request.CustomRequest) -> dict:
        param_date: datetime.datetime = datetime.datetime.strptime(req.param("param_date", is_safe=False), "%Y-%m-%d")
        param_shift: int = int(req.param("param_shift", is_safe=False))
        param_target: int = int(req.param("param_target", is_safe=False))
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
            "param_target": param_target,
        }

        is_fake = not settings.IS_ASD_SERVER

        # беру всех водителей самосвалов за смену
        __rows_raw: list[tuple] = utils.Sql.request_to_oracle(
            query=queries.Stoppages.get_stoppages_report_veh_operators(),
            args={
                "param_date": param_date,
                "param_shift": param_shift,
            },
            many=True,
            is_fake=is_fake,
            fake_file_path="static/fake/get_stoppages_report_veh_operators.xlsx",
        )
        # print(__rows_raw[-1])

        # превращаю массив точек в массив словарей
        __rows_dict: list[dict] = [
            {
                "veh_id": i[0],
                "veh_id_fio": i[1],
            }
            for i in __rows_raw[2 if is_fake else 0 :]
        ]

        # создаю словарь с ключом в виде номера самосвала и значение в виде фио оператора
        operators = {}
        for i in __rows_dict:
            operators[i["veh_id"]] = i["veh_id_fio"]

        # беру все "точки" всех самосвалов за смену
        __rows_raw: list[tuple] = utils.Sql.request_to_oracle(
            query=queries.Stoppages.get_stoppages_report_veh_dvs_new(),
            args={
                "param_date": param_date,
                "param_shift": param_shift,
            },
            many=True,
            is_fake=is_fake,
            fake_file_path="static/fake/get_stoppages_report_veh_dvs_new.xlsx",
        )
        # print(__rows_raw[-1])

        # превращаю массив точек в массив словарей
        __rows_dict: list[dict] = [
            {
                "veh_id": i[0],
                "date_time": i[1],
                "speed": i[2],
                "motohours": i[3],
            }
            for i in __rows_raw[2 if is_fake else 0 :]
        ]
        # print(__rows_dict[-1])

        # группирую словари по автосамосвалам
        __rows_dict_by_veh_id: dict[str, list] = {}
        for i in __rows_dict:
            try:
                __rows_dict_by_veh_id[i["veh_id"]].append(i)
            except Exception as _:
                __rows_dict_by_veh_id[i["veh_id"]] = [i]
        # print(__rows_dict_by_veh_id)

        __all_stoppages: list[dict] = []
        # запуск цикла по каждому самосвалу
        for k, events_by_veh_id in __rows_dict_by_veh_id.items():
            # очистка простоев определённой длины
            while True:
                try:
                    last_index = 0
                    local_interval = []
                    for event in events_by_veh_id:
                        last_index += 1
                        if event["speed"] == 0:
                            if len(local_interval) != 0:
                                last: datetime.datetime = local_interval[-1]["date_time"]
                                cur: datetime.datetime = event["date_time"]
                                if (last - cur).total_seconds() > 5 * 60:
                                    break
                            local_interval.append(event)
                        else:
                            if len(local_interval) == 0:
                                pass
                            else:
                                break
                    last_point: dict = local_interval[0]
                    first_point: dict = local_interval[-1]
                    difference_time = round((last_point["date_time"] - first_point["date_time"]).total_seconds() / 60, 2)
                    difference_motohours = round(last_point["motohours"] - first_point["motohours"], 2)
                    events_by_veh_id = events_by_veh_id[last_index:]
                    if difference_motohours > param_target:
                        if difference_motohours > difference_time:
                            difference_motohours = difference_time
                        __all_stoppages.append(
                            {
                                "veh_id": k,
                                "veh_id_fio": operators[k],
                                "from": first_point["date_time"],
                                "to": last_point["date_time"],
                                "duration": round(difference_motohours / 60, 2),
                            }
                        )
                except Exception as _:
                    break

        # сортировка простоев по времени начала
        __all_stoppages = sorted(__all_stoppages, key=lambda x: x["from"], reverse=False)
        summary_duration = 0.0
        for i in __all_stoppages:
            summary_duration += float(i["duration"])

        to_excel: list[list] = []
        for i in __all_stoppages:
            to_excel.append(
                [
                    param_date,
                    param_shift,
                    i["veh_id"],
                    i["veh_id_fio"],
                    i["from"],
                    i["to"],
                    i["duration"],
                ]
            )
        titles: list[str] = [
            "Дата",
            "Смена",
            "Самосвал",
            "Оператор",
            "Начало простоя",
            "Окончание простоя",
            "Длительность простоя ДВС, часы",
        ]
        path_to_excel_file: str = utils.Excel.export_to_excel(
            _data=to_excel,
            _titles=titles,
            _folder_name="stoppages",
            _prefix=f"{param_date.strftime('%Y-%m-%d')}_смена_{param_shift}___холостые_простои_двс_new",
        )

        return {
            "data": __all_stoppages,
            "extra": {
                "path_to_excel_file": path_to_excel_file,
                "summary_duration": round(summary_duration, 2),
            },
            "parameters": __parameters,
        }

    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def get_empty_peregon_report_dumptrucks(req: utils.Request.CustomRequest) -> dict:
        param_date: datetime.datetime = datetime.datetime.strptime(req.param("param_date", is_safe=False), "%Y-%m-%d")
        param_shift: int = int(req.param("param_shift", is_safe=False))
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
        }

        is_fake = not settings.IS_ASD_SERVER

        # все точки списком
        _events_list = utils.Sql.request_to_oracle(
            query=queries.Stoppages.get_empty_peregon_report_dumptrucks_new(),
            args={
                "param_date": param_date,
                "param_shift": param_shift,
            },
            many=True,
            is_fake=is_fake,
            fake_file_path="static/fake/get_empty_peregon_report_dumptrucks_new.xlsx",
        )
        # все точки за смену словарями
        _events_list_dict = [
            {
                "veh_id": i[0],
                "date_time": i[1],
                "x": int(i[2]),
            }
            for i in _events_list[2 if is_fake else 0 :]
        ]
        # все точки сгруппированные по хоз.номеру
        _events_dict_by_veh_id: dict[str, list[dict]] = {}
        for i in _events_list_dict:
            try:
                _events_dict_by_veh_id[i["veh_id"]].append(i)
            except Exception as _:
                _events_dict_by_veh_id[i["veh_id"]] = [i]

        # все перегоны
        _result_peregons: list[dict] = []
        line = 86250
        for _veh_id, _events in _events_dict_by_veh_id.items():
            vector = not (_events[0]["x"] > line)
            prev: datetime.datetime = None
            for i in _events:
                curr = i["date_time"]
                diff = 1 * 60 * 60 if prev is None else (curr - prev).total_seconds()
                is_abk = i["x"] > line
                if diff > 15 * 60:
                    if is_abk:
                        if vector:
                            vector = False
                            _result_peregons.append({"veh_id": f"{_veh_id}", "date_time": curr, "target": "На АБК"})
                            prev = curr
                    else:
                        if vector is False:
                            vector = True
                            _result_peregons.append({"veh_id": f"{_veh_id}", "date_time": curr, "target": "На Карьер"})
                            prev = curr
        _result_peregons = sorted(_result_peregons, key=lambda x: (x["veh_id"], x["date_time"]), reverse=False)
        to_excel: list[list] = []
        for i in _result_peregons:
            to_excel.append(
                [
                    param_date,
                    param_shift,
                    i["veh_id"],
                    i["date_time"],
                    i["target"],
                ]
            )
        titles: list[str] = [
            "Дата",
            "Смена",
            "Самосвал",
            "Время пересечения",
            "Направление",
        ]
        path_to_excel_file: str = utils.Excel.export_to_excel(
            _data=to_excel,
            _titles=titles,
            _folder_name="peregons",
            _prefix=f"{param_date.strftime('%Y-%m-%d')}_смена_{param_shift}___холостые_перегоны_new",
        )
        return {
            "data": _result_peregons,
            "extra": {
                "path_to_excel_file": path_to_excel_file,
            },
            "parameters": __parameters,
        }


class Speed:
    @staticmethod
    @csrf_exempt
    @utils.Request.decorator_json_response(need_auth=False, is_cache=True, cache_timeout=10)
    def get_target_report_avg_speed_custom(req: utils.Request.CustomRequest) -> dict:
        # params
        param_date: datetime.datetime = datetime.datetime.strptime(req.param("param_date", is_safe=False), "%Y-%m-%d")
        param_shift: int = int(req.param("param_shift", is_safe=False))
        __parameters = {
            "param_date": param_date,
            "param_shift": param_shift,
        }

        is_fake = not settings.IS_ASD_SERVER

        # raw
        trips_raw: list[tuple] = utils.Sql.request_to_oracle(
            query=queries.Speed.get_target_report_avg_speed_custom(),
            args={"param_date": param_date, "param_shift": param_shift},
            many=True,
            is_fake=is_fake,
            fake_file_path="static/fake/get_target_report_avg_speed_custom.xlsx",
        )
        # print("trips_raw: ", trips_raw[0:3])
        """
[
[0,         1,          2,      3,      4,                      5,                  6,              7,              8,              9,              10,                 11,                     12,             13,                     14,                      15,            16,                 17,                 18], 
('SHOVID', 'WORKTYPE', 'VEHID', 'FIO', 'WEIGHT_MOVE_LOAD', 'LENGTH_MOVE_LOAD', 'TIME_LOAD', 'TIME_MOVE_LOAD', 'TIME_UNLOAD', 'TIME_MOVE_UNLOAD', 'TIME_LOAD_NEXT', 'DURATION_LOADING', 'DURATION_MOVE_LOAD', 'DURATION_UNLOADING', 'DURATION_MOVE_UNLOAD', 'DURATION_TRIP', 'SPEED_LOAD_ASD', 'SPEED_LOAD_CUSTOM', 'SPEED_DIFF'), 
('203', 'Вскрыша скальная', '142', 'Рецлов Олег Николаевич', 85, 3, datetime.datetime(2023, 12, 11, 8, 6, 50), datetime.datetime(2023, 12, 11, 8, 9, 4), datetime.datetime(2023, 12, 11, 8, 18, 30), datetime.datetime(2023, 12, 11, 8, 19, 3), datetime.datetime(2023, 12, 11, 8, 29, 48), 2.23, 9.43, 0.55, 10.75, 22.97, 19.96, 19.08, 0.88)
]"""

        # serialize
        trips_dict = [
            {
                "SHOVID": str(i[0]),
                "WORKTYPE": str(i[1]),
                "VEHID": str(i[2]),
                "FIO": str(i[3]),
                "WEIGHT_MOVE_LOAD": int(i[4]),
                "LENGTH_MOVE_LOAD": float(i[5]),
                "TIME_LOAD": i[6],
                "TIME_MOVE_LOAD": i[7],
                "TIME_UNLOAD": i[8],
                "TIME_MOVE_UNLOAD": i[9],
                "TIME_LOAD_NEXT": i[10],
                "DURATION_LOADING": float(i[11]) if i[11] else 0.0,
                "DURATION_MOVE_LOAD": float(i[12]) if i[12] else 0.0,
                "DURATION_UNLOADING": float(i[13]),
                "DURATION_MOVE_UNLOAD": float(i[14]),
                "DURATION_TRIP": float(i[15]),
                "SPEED_LOAD_ASD": float(i[16]),
                "SPEED_LOAD_CUSTOM": float(i[17]) if i[17] else 0.0,
                "SPEED_DIFF": float(i[18]) if i[18] else 0.0,
            }
            for i in trips_raw[2 if is_fake else 0 :]
        ]

        return {"data": trips_dict, "parameters": __parameters}
